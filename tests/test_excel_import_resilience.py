"""
Suite de Pruebas de Resiliencia del Importador Inteligente de Excel (Smart Scanner)
==================================================================================

Valida:
1. Limpieza de celdas con errores de fórmulas (#REF!, #N/A, #VALUE!).
2. Detección dinámica de cabeceras en filas intermedias (fila 4, 6, 7, 8).
3. Soporte de cabeceras multinivel (combinación de dos filas consecutivas).
4. Selección inteligente de hojas en libros multi-hojas.
5. Parseo integral de los archivos Excel reales de la UAGRM si están presentes en Downloads.
"""

import os
import sys
import glob
import pytest
import openpyxl
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.student_service import (
    _clean_text,
    _normalize_header,
    _split_carnet,
    _find_headers_smart,
    _pick_best_sheet_for_students,
)


class TestCellSanitization:
    """Valida la sanitización de celdas y errores de fórmula."""

    def test_clean_text_formula_errors(self):
        assert _clean_text("#REF!") is None
        assert _clean_text("#N/A") is None
        assert _clean_text("#VALUE!") is None
        assert _clean_text("#DIV/0!") is None
        assert _clean_text("NULL") is None
        assert _clean_text("NaN") is None
        assert _clean_text("   ") is None
        assert _clean_text(None) is None

    def test_clean_text_valid_numbers_and_strings(self):
        assert _clean_text(2969698.0) == "2969698"
        assert _clean_text("3897608.0") == "3897608"
        assert _clean_text("Juan Perez") == "Juan Perez"
        assert _clean_text(12345) == "12345"

    def test_split_carnet_variations(self):
        assert _split_carnet("2726683 - 1J") == ("2726683", "1J")
        assert _split_carnet("1313665-1D") == ("1313665", "1D")
        assert _split_carnet("2969698") == ("2969698", None)
        assert _split_carnet("#REF!") == (None, None)
        assert _split_carnet(None) == (None, None)


class TestSmartHeaderDetection:
    """Valida la detección de cabeceras en distintas posiciones y multinivel."""

    def test_header_at_row_1(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombres", "Apellidos", "CI", "Registro"])
        ws.append(["Juan", "Perez", "123456", "20201234"])
        
        h_row, start_data_row, col_map, _ = _find_headers_smart(ws)
        assert h_row == 1
        assert start_data_row == 2
        assert col_map["col_nombre"] == 1
        assert col_map["col_apellido"] == 2
        assert col_map["col_carnet"] == 3
        assert col_map["col_registro"] == 4

    def test_header_at_row_6_with_titles_above(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["UNIDAD DE POSTGRADO"])
        ws.append(["MAESTRIA EN AUDITORIA"])
        ws.append([""])
        ws.append(["DOCENTE: MSC."])
        ws.append(["FECHA: 2026"])
        ws.append(["Nº", "APELLIDOS Y NOMBRES", "REGISTRO", "C.I.", "EXP."])
        ws.append([1, "ALVAREZ ESTELA", 213182696, 3897608, "SCZ"])
        ws.append([2, "ARTEAGA HENRY", 951005642, 3935402, "SCZ"])

        h_row, start_data_row, col_map, _ = _find_headers_smart(ws)
        assert h_row == 6
        assert start_data_row == 7
        assert col_map["col_nombre"] == 2
        assert col_map["col_registro"] == 3
        assert col_map["col_carnet"] == 4
        assert col_map["col_extension"] == 5

    def test_multilevel_headers_row_8_and_9(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(7):
            ws.append(["TITULO INSTITUCIONAL"])
        ws.append(["Nº", "APELLIDOS Y NOMBRES", None, None])
        ws.append([None, None, "Registro", "C.I."])
        ws.append([1, "ALVAREZ ESTELA", 213182696, 3897608])

        h_row, start_data_row, col_map, _ = _find_headers_smart(ws)
        assert h_row == 8
        assert start_data_row == 10
        assert col_map["col_nombre"] == 2
        assert col_map["col_registro"] == 3
        assert col_map["col_carnet"] == 4


class TestRealFilesCompatibility:
    """Pruebas opcionales directas contra los archivos descargados de la UAGRM."""

    def test_parse_real_downloads_if_present(self):
        downloads = os.path.expanduser(r"C:\Users\Usuario\Downloads")
        if not os.path.exists(downloads):
            pytest.skip("Carpeta Downloads no accesible")

        # 1. Probar archivo MGTAF
        mgtaf_files = glob.glob(os.path.join(downloads, "*MGTAF*.xlsx"))
        if mgtaf_files:
            wb = openpyxl.load_workbook(mgtaf_files[0], data_only=True)
            sheet, start_row, col_map, _ = _pick_best_sheet_for_students(wb)
            assert sheet is not None
            assert col_map["col_nombre"] > 0
            assert col_map["col_carnet"] > 0

            students = []
            for r in range(start_row, sheet.max_row + 1):
                nombre = _clean_text(sheet.cell(row=r, column=col_map["col_nombre"]).value)
                carnet, _ = _split_carnet(sheet.cell(row=r, column=col_map["col_carnet"]).value)
                if not nombre and not carnet:
                    continue
                if nombre and any(tok in nombre.upper() for tok in ("ELABORADO POR", "FIRMA", "TOTAL", "PROMEDIO")):
                    continue
                students.append((nombre, carnet))

            assert len(students) >= 80, f"Debe extraer al menos 80 estudiantes de MGTAF (obtenidos: {len(students)})"

        # 2. Probar archivo DIPLO EDUCACION SUPERIOR
        diplo_files = glob.glob(os.path.join(downloads, "*DIPLO. EDU. SUP*.xlsx"))
        if diplo_files:
            wb = openpyxl.load_workbook(diplo_files[0], data_only=True)
            sheet, start_row, col_map, _ = _pick_best_sheet_for_students(wb)
            assert sheet is not None
            assert col_map["col_nombre"] > 0
            assert col_map["col_carnet"] > 0

            students = []
            for r in range(start_row, sheet.max_row + 1):
                nombre = _clean_text(sheet.cell(row=r, column=col_map["col_nombre"]).value)
                carnet, _ = _split_carnet(sheet.cell(row=r, column=col_map["col_carnet"]).value)
                if not nombre and not carnet:
                    continue
                students.append((nombre, carnet))

            assert len(students) >= 40, f"Debe extraer al menos 40 estudiantes del Diplomado (obtenidos: {len(students)})"
