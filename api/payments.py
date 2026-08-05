# -*- coding: utf-8 -*-
"""
API de Pagos (Payments)
=======================

Endpoints para gestionar pagos de estudiantes, incluyendo 
rollback financiero y control de Caja/Bancos.
"""

from typing import List, Any, Optional, Union
import asyncio
import json
import re
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, status
from pydantic import BaseModel, Field # IMPORTACIÓN REQUERIDA
from models.course import Course
from models.payment import Payment
from models.student import Student
from models.user import User
from models.enums import EstadoPago
from schemas.payment import (
    PaymentCreate,
    PaymentResponse,
    PaymentApproval,
    PaymentRejection,
    PaymentReversion,
    PaymentWithDetails
)
from services import payment_service
from beanie import PydanticObjectId
from beanie.operators import In
from bson import ObjectId
from fastapi.encoders import jsonable_encoder

from api.dependencies import require_cobranza, require_staff, require_superadmin, get_current_user, filtro_cursos_por_rol, puede_ver_economico
from schemas.common import PaginatedResponse, PaginationMeta
import math

router = APIRouter()


@router.post(
    "/",
    response_model=PaymentResponse,
    status_code=201,
    summary="Registrar Pago"
)
async def create_payment(
    *,
    inscripcion_id: str = Form(..., description="ID de la inscripción"),
    metodo_pago: str = Form(default="Transferencia", description="Transferencia, Depósito o Caja"),
    monto_comprobante: float = Form(...),
    concepto: Optional[str] = Form(None),

    # Datos opcionales según el método de pago
    numero_transaccion: Optional[str] = Form(None),
    remitente: Optional[str] = Form(None),
    banco: Optional[str] = Form(None),
    fecha_comprobante: Optional[str] = Form(None),
    cuenta_destino: Optional[str] = Form(None),
    # F-COBRANZA-026 (2026-07-22): Kevin: "el sistema no debe dejar que se suba
    # una solicitud de pago sin el comprobante no importa que tipo de pago sea
    # caja o todo lo demas". El comprobante es OBLIGATORIO para todos los
    # metodos, incluyendo Caja.
    file: UploadFile = File(..., description="Comprobante obligatorio (imagen o PDF)"),

    # F-SYNC-PAGOS-MODULOS (2026-08-04, Kevin): sincronizar con la logica
    # del modal CargaInicialModal. Si viene, se aplica directo a los modulos
    # en vez de prorratear. Formato JSON: '{"2": 294}' = paga modulo 3 (Bs 294).
    pagos_modulos_json: Optional[str] = Form(
        default=None,
        description='JSON con Dict[str, float]. Ej: \'{"2": 294}\' = paga modulo 3.'
    ),
    detalle: Optional[str] = Form(
        default=None,
        description="Detalle desglosado del pago (opcional, se genera auto si vienen pagos_modulos)."
    ),

    current_user: Union[User, Student] = Depends(get_current_user)
) -> Any:
    """
    Registrar un nuevo pago.
    Soporta pagos digitales (exige voucher/número) y pagos físicos (Caja).
    - Si lo sube un Estudiante: entra obligatoriamente como PENDIENTE.
    - Si lo registra un Perfil Autorizado (User/Staff): se AUTO-VALIDA de inmediato.
    """
    from core.cloudinary_utils import upload_image, upload_pdf
    from schemas.payment import PaymentCreate

    is_staff = isinstance(current_user, User)

    # 1. Validaciones rígidas según el Método de Pago
    if not file:
        raise HTTPException(status_code=400, detail="El comprobante es obligatorio (imagen o PDF) para todos los métodos de pago.")
    if metodo_pago != "Caja":
        if not numero_transaccion:
            raise HTTPException(status_code=400, detail="El número de transacción es obligatorio para este método de pago.")
        if not banco:
            raise HTTPException(status_code=400, detail="Debe especificar el banco emisor.")
    
    comprobante_url = None
    
    try:
        # 2. Subida de Archivo a la Nube
        student_id_for_folder = current_user.id if not is_staff else "staff_upload"
        if file:
            folder = f"payments/{student_id_for_folder}"
            safe_transaction = (numero_transaccion or "caja_pago").replace(' ', '_').replace('/', '_')
            public_id = f"voucher_{safe_transaction}"
            
            image_types = ["image/jpeg", "image/jpg", "image/png", "image/webp"]
            pdf_type = "application/pdf"
            
            if file.content_type in image_types:
                comprobante_url = await upload_image(file, folder, public_id)
            elif file.content_type == pdf_type:
                comprobante_url = await upload_pdf(file, folder, public_id)
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Formato no permitido: {file.content_type}. Use imagen o PDF"
                )
        
        # 3. Ensamblaje del Payload Relajado
        # F-SYNC-PAGOS-MODULOS (2026-08-04, Kevin): parsear pagos_modulos_json
        # si viene (formato JSON string). Si no, queda None.
        pagos_modulos_dict = None
        if pagos_modulos_json:
            try:
                pagos_modulos_dict = json.loads(pagos_modulos_json)
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"pagos_modulos_json invalido (debe ser JSON valido): {e}"
                )
        payment_in = PaymentCreate(
            inscripcion_id=inscripcion_id,
            metodo_pago=metodo_pago,
            monto_comprobante=monto_comprobante,
            concepto=concepto,
            cantidad_pago=monto_comprobante,
            numero_transaccion=numero_transaccion,
            remitente=remitente,
            banco=banco,
            fecha_comprobante=fecha_comprobante,
            cuenta_destino=cuenta_destino,
            comprobante_url=comprobante_url,
            pagos_modulos=pagos_modulos_dict,
            detalle=detalle,
        )
        
        if is_staff:
            # Obtener el student_id desde la inscripción
            from models.enrollment import Enrollment
            from beanie import PydanticObjectId as _POI
            enr_oid = _POI(inscripcion_id) if not isinstance(inscripcion_id, _POI) else inscripcion_id
            enrollment_obj = await Enrollment.get(enr_oid)
            if not enrollment_obj:
                raise HTTPException(status_code=404, detail="Inscripción no encontrada.")
            target_student_id = enrollment_obj.estudiante_id
            
            payment = await payment_service.create_payment(
                payment_in=payment_in,
                student_id=target_student_id,
                auto_approve=True,
                approved_by=current_user.username,
                skip_ownership_check=True,
                # F-087: staff subiendo via /payments/ → subido_por="estudiante"
                # (semánticamente: el staff está subiendo en nombre del estudiante).
                # Para distinguir esto del caso upload-by-encargado (que se
                # llama desde otro endpoint y se setea allí), usamos "estudiante"
                # porque el endpoint es el genérico de creación.
                subido_por="estudiante",
            )
        else:
            payment = await payment_service.create_payment(
                payment_in=payment_in,
                student_id=current_user.id,
                auto_approve=False,
                # F-087: estudiante subiendo su propio comprobante.
                subido_por="estudiante",
            )
        
        return await payment_service.enrich_payment_with_details(payment)

    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear pago: {str(e)}")


@router.post(
    "/by-staff",
    response_model=PaymentResponse,
    summary="Registrar Pago en nombre de Estudiante (Staff: cobranza/admin/superadmin)"
)
async def create_payment_by_staff(
    *,
    inscripcion_id: str = Form(..., description="ID de la inscripción"),
    estudiante_id: str = Form(..., description="ID del estudiante (para asociar el comprobante en Cloudinary)"),
    metodo_pago: str = Form(default="Transferencia", description="Transferencia, Depósito o Caja"),
    monto_comprobante: float = Form(..., gt=0, description="Monto del pago en BOB (>0)"),
    concepto: Optional[str] = Form(None, description="Concepto (opcional; si vacío, backend calcula glosa detallada)"),

    numero_transaccion: Optional[str] = Form(None),
    remitente: Optional[str] = Form(None),
    banco: Optional[str] = Form(None),
    fecha_comprobante: Optional[str] = Form(None),
    cuenta_destino: Optional[str] = Form(None),
    # F-COBRANZA-026 (2026-07-22): comprobante obligatorio para todos los métodos
    file: UploadFile = File(..., description="Comprobante obligatorio (imagen o PDF)"),

    current_user: User = Depends(require_staff)
) -> Any:
    """
    F-COBRANZA-017 (2026-07-22): cuando el estudiante no pudo subir su
    comprobante desde su perfil (por problemas técnicos, falta de acceso,
    etc.), el personal de COBRANZA puede REGISTRAR el pago completo en
    nombre del estudiante desde la gestion de pagos.

    Decisión Joel 2026-07-22 22:25: el botón vive en la parte superior
    derecha de /app/payments (no en el menú de 3 puntos de cada pago).
    Roles permitidos: superadmin, admin, cobranza. NO cpd, NO coordinador,
    NO encargado_curso.

    Diferencias con `create_payment` (estudiante):
    - Acepta `estudiante_id` (no lo toma del current_user).
    - El pago nace APROBADO (auto-aprobación como F-COBRANZA-004).
    - `verificado_por` = nombre del usuario staff.
    - Glosa: si el frontend manda placeholder genérico ("Matrícula" /
      "Módulo"), se regenera con detalle de módulos cubiertos.
    - F-COBRANZA-026: comprobante OBLIGATORIO (no opcional para Caja).
    """
    from core.cloudinary_utils import upload_image, upload_pdf
    from schemas.payment import PaymentCreate

    # RBAC: solo roles económicos (no cpd, no coordinador, no docente)
    if current_user.rol not in ["superadmin", "admin", "cobranza"]:
        raise HTTPException(
            status_code=403,
            detail="Solo cobranza/admin/superadmin pueden registrar pagos en nombre de estudiantes."
        )

    # F-COBRANZA-026: comprobante obligatorio siempre
    if not file:
        raise HTTPException(status_code=400, detail="El comprobante es obligatorio (imagen o PDF) para todos los métodos de pago.")
    if metodo_pago != "Caja":
        if not numero_transaccion:
            raise HTTPException(status_code=400, detail="El número de transacción es obligatorio para este método.")
        if not banco:
            raise HTTPException(status_code=400, detail="Debe especificar el banco emisor.")

    comprobante_url = None
    if file:
        folder = f"payments/{estudiante_id}"
        if numero_transaccion:
            public_id = f"voucher_{numero_transaccion}"
        else:
            public_id = f"staff_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        image_types = ["image/jpeg", "image/jpg", "image/png", "image/webp"]
        pdf_type = "application/pdf"

        if file.content_type in image_types:
            comprobante_url = await upload_image(file, folder, public_id)
        elif file.content_type == pdf_type:
            comprobante_url = await upload_pdf(file, folder, public_id)
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Formato no permitido: {file.content_type}. Use imagen o PDF"
            )

    payment_in = PaymentCreate(
        inscripcion_id=inscripcion_id,
        metodo_pago=metodo_pago,
        monto_comprobante=monto_comprobante,
        concepto=concepto,
        cantidad_pago=monto_comprobante,
        numero_transaccion=numero_transaccion,
        remitente=remitente,
        banco=banco,
        fecha_comprobante=fecha_comprobante,
        cuenta_destino=cuenta_destino,
        comprobante_url=comprobante_url,
    )  # /payments/by-staff ya no soporta pagos_modulos (no es Form-friendly); se usa /payments/ con pagos_modulos_json.

    try:
        # F-COBRANZA-017: el pago se crea APROBADO al registrarlo desde
        # cobranza (no pasa por el flujo pendiente → revisar → aprobar).
        # Esto es consistente con F-COBRANZA-004 (auto-aprobación cuando
        # el estudiante sube su comprobante). Joel decidió 22:25 que sea
        # automático para no demorar al estudiante.
        #
        # F-COBRANZA-034 (2026-07-22): skip_ownership_check=True porque el
        # check de "la inscripcion pertenece al estudiante" es solo para
        # el endpoint del estudiante (evitar que un estudiante pague la
        # inscripcion de otro). El staff (cobranza/admin/superadmin) está
        # autorizado a registrar pagos en nombre de cualquier estudiante
        # del sistema. Bug reportado por Lic. Sandra Zabala: el check
        # enrollment.estudiante_id != student_id siempre fallaba porque
        # estudiante_id llega como string del Form y enrollment.estudiante_id
        # es PydanticObjectId.
        from beanie import PydanticObjectId as _POI
        student_oid = estudiante_id if isinstance(estudiante_id, _POI) else _POI(estudiante_id)
        payment = await payment_service.create_payment(
            payment_in=payment_in,
            student_id=student_oid,
            auto_approve=True,
            approved_by=current_user.username,
            skip_ownership_check=True,
            # F-087: staff registrando el pago COMPLETO en nombre del estudiante
            # (no es el caso "estudiante subió comprobante"). Es el encargado
            # quien hizo todo el registro, así que subido_por="encargado".
            subido_por="encargado",
        )

        # F-COBRANZA-014: el saldo del enrollment se actualiza dentro de
        # create_payment (vía actualizar_saldo_enrollment). No hace falta
        # hacer nada más aquí.

        return await payment_service.enrich_payment_with_details(payment)

    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al registrar pago por staff: {str(e)}")


@router.get(
    "/",
    response_model=PaginatedResponse[PaymentResponse],
    summary="Listar Pagos"
)
async def list_payments(
    *,
    page: int = Query(1, ge=1, description="Número de página"),
    per_page: int = Query(10, ge=1, le=500, description="Elementos por página"),
    q: Optional[str] = Query(None, description="Búsqueda por transacción o comprobante"),
    estado: Optional[str] = Query(None, description="Filtrar por estado"),
    curso_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por Curso ID"),
    estudiante_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por Estudiante ID"),
    tipo_concepto: Optional[str] = Query(None, description="Filtrar por tipo de concepto (matricula, colegiatura)"),
    current_user: User | Student = Depends(get_current_user)
) -> Any:
    
    if isinstance(current_user, User):
        # AUDITORÍA (CRÍTICO #1): sin este guard, cualquier rol autenticado
        # (docente, encargado_curso, coordinador) caía en ningún filtro y veía
        # TODOS los pagos/comprobantes del sistema. Solo el personal financiero
        # y de gestión académica tiene algún tipo de acceso a esta vista.
        if current_user.rol not in ["superadmin", "admin", "mae", "cpd", "cobranza"]:
            raise HTTPException(status_code=403, detail="No autorizado para ver pagos")

        # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo ve pagos de esos cursos.
        filtro_rol = filtro_cursos_por_rol(current_user)
        cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

        payments, total_count = await payment_service.get_all_payments(
            page=page,
            per_page=per_page,
            q=q,
            estado=estado,
            curso_id=curso_id,
            estudiante_id=estudiante_id,
            cursos_permitidos=cursos_permitidos,
            tipo_concepto=tipo_concepto
        )
        
        # Filtrado de RBAC (CPD vs Cobranza)
        filtered_payments = []
        for p in payments:
            concepto_lower = (p.concepto or "").lower().strip()
            is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
            if current_user.rol == "cpd" and not is_matricula:
                continue
            filtered_payments.append(p)
            
        payments = filtered_payments
        total_count = len(filtered_payments)
        
    elif isinstance(current_user, Student):
        all_payments = await payment_service.get_payments_by_student(
            student_id=current_user.id
        )
        if estado and estado != "Todos los estados":
            all_payments = [p for p in all_payments if p.estado_pago.value == estado]
            
        if tipo_concepto:
            if tipo_concepto == "matricula":
                all_payments = [p for p in all_payments if "matricula" in (p.concepto or "").lower() or "matrícula" in (p.concepto or "").lower()]
            elif tipo_concepto == "colegiatura":
                all_payments = [p for p in all_payments if "matricula" not in (p.concepto or "").lower() and "matrícula" not in (p.concepto or "").lower()]
            
        total_count = len(all_payments)
        start = (page - 1) * per_page
        end = start + per_page
        payments = all_payments[start:end]
    else:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    total_pages = math.ceil(total_count / per_page) if total_count > 0 else 0
    has_next = page < total_pages
    has_prev = page > 1
    
    enriched_payments = await payment_service.enrich_payments_with_details_bulk(payments)
    
    return {
        "data": enriched_payments,
        "meta": PaginationMeta(
            page=page,
            limit=per_page,
            totalItems=total_count,
            totalPages=total_pages,
            hasNextPage=has_next,
            hasPrevPage=has_prev
        )
    }


@router.get("/pendientes/list", response_model=List[PaymentResponse])
async def get_payments_pendientes(
    *,
    current_user: User = Depends(require_staff)
) -> Any:
    if current_user.rol not in ["superadmin", "admin", "cpd", "cobranza"]:
        raise HTTPException(status_code=403, detail="No autorizado para listar pagos pendientes")

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo ve pendientes de esos cursos.
    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    payments = await payment_service.get_payments_pendientes(cursos_permitidos=cursos_permitidos)
    
    filtered_payments = []
    for p in payments:
        concepto_lower = (p.concepto or "").lower().strip()
        is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
        
        if current_user.rol == "cpd":
            if is_matricula:
                filtered_payments.append(p)
        else:
            filtered_payments.append(p)
            
    return await payment_service.enrich_payments_with_details_bulk(filtered_payments)


def _parse_rango_fechas(fecha_desde: Optional[str], fecha_hasta: Optional[str]):
    from datetime import datetime, date
    if not fecha_desde:
        fecha_desde = date.today().isoformat()
    if not fecha_hasta:
        fecha_hasta = fecha_desde
    try:
        fecha_desde_dt = datetime.fromisoformat(fecha_desde)
        fecha_hasta_dt = datetime.fromisoformat(fecha_hasta).replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usar YYYY-MM-DD")
    return fecha_desde, fecha_hasta, fecha_desde_dt, fecha_hasta_dt


@router.get(
    "/dashboard/resumen-economico",
    summary="Resumen Económico del Dashboard (Cobranza / Coordinador Financiero)"
)
async def get_resumen_economico_endpoint(
    *,
    current_user: User = Depends(require_staff)
) -> Any:
    """
    ISSUE-P-DASHBOARD-COBRANZA: tarjetas de resumen económico del dashboard.
    Incluye el ingreso por matrícula como dato contable (aunque Cobranza no
    apruebe matrículas, sí debe verlas recaudadas porque genera los informes
    económicos). Mismo conjunto de roles económicos que los reportes de caja.
    """
    # ISSUE-R-PERFIL-GENERICO: económico = superadmin/admin/cobranza/mae + coordinador FINANCIERO.
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver el resumen económico")

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo ve su(s) curso(s).
    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    return await payment_service.get_resumen_economico(cursos_permitidos=cursos_permitidos)


@router.get(
    "/reportes/caja",
    summary="Reporte de Caja por Fechas (Tabla Interactiva)"
)
async def get_reporte_caja_endpoint(
    *,
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    curso_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por curso"),
    estudiante_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por estudiante (F-COBRANZA-003)"),
    estado: Optional[str] = Query(None, description="Filtrar por estado del pago"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=200),
    current_user: User = Depends(require_staff)
) -> Any:
    """
    ISSUE-P-REPORTE: tabla interactiva de ingresos filtrable por rango de
    fechas (fecha real del pago), curso y edición de programa. Devuelve la
    página solicitada + totales agregados de TODO el rango filtrado (no solo
    la página actual) para el resumen visual encima de la tabla.

    F-COBRANZA-003 (2026-07-21): filtro opcional por estudiante_id.
    Permite ver todos los pagos de un estudiante específico en el rango.
    """
    # CPD excluido: los reportes de caja son económicos (regla del usuario:
    # "económico solo cobranza y el coordinador financiero"; CPD solo audita la
    # matrícula desde Gestión de Pagos, no ve reportes de caja).
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver reportes de caja")

    _, _, fecha_desde_dt, fecha_hasta_dt = _parse_rango_fechas(fecha_desde, fecha_hasta)

    concepto_regex = None
    if current_user.rol == "cpd":
        concepto_regex = {"concepto": {"$regex": r"^matr[ií]cula$", "$options": "i"}}

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo ve su(s) curso(s).
    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    resultado = await payment_service.get_reporte_caja(
        fecha_desde_dt=fecha_desde_dt,
        fecha_hasta_dt=fecha_hasta_dt,
        page=page,
        per_page=per_page,
        curso_id=curso_id,
        estudiante_id=estudiante_id,  # F-COBRANZA-003
        estado=estado,
        concepto_regex=concepto_regex,
        cursos_permitidos=cursos_permitidos
    )

    total_pages = math.ceil(resultado["total_count"] / per_page) if resultado["total_count"] > 0 else 0
    enriched = await payment_service.enrich_payments_with_details_bulk(resultado["payments"])

    # Los pagos enriquecidos vienen de model_dump() y conservan campos PyObjectId
    # F-075-FIX-7 (2026-07-23): faltaba el `return` (bug pre-existente). El
    # endpoint respondia 200 con body null porque no retornaba nada. Aqui se
    # retorna la estructura completa: lista paginada + resumen + total.
    has_next = page < total_pages
    has_prev = page > 1

    return {
        "data": enriched,
        "meta": PaginationMeta(
            page=page,
            limit=per_page,
            totalItems=resultado["total_count"],
            totalPages=total_pages,
            hasNextPage=has_next,
            hasPrevPage=has_prev
        ),
        "resumen": resultado["resumen"],
    }


@router.get(
    "/reportes/lista-habilitados",
    summary="F-075: Lista de Postgraduantes Habilitados (informe acta de notas)"
)
async def get_lista_habilitados(
    *,
    curso_id: PydanticObjectId = Query(..., description="ID del curso (obligatorio)"),
    modulo_index: Optional[int] = Query(
        None,
        ge=0,
        description=(
            "0 = solo matrícula, 1..N = solo ese módulo, None/omitido = TODOS los módulos "
            "(un registro por estudiante-módulo)"
        ),
    ),
    current_user: User = Depends(require_staff),
) -> Any:
    """
    F-075 (2026-07-23): Genera la 'Lista de Postgraduantes Habilitados' para
    aprobación de acta de notas. Formato estilo papel Sandra:

    Encabezado:
      LISTA DE POSTGRADUANTES HABILITADOS
      (MAESTRÍA / DIPLOMADO / DOCTORADO) <nombre del programa>
      MÓDULO: <Módulo N: nombre>
      PERÍODO: <rango de fechas de los pagos>
      DOCENTE: <nombre del docente del módulo>

    Filas: N° | Apellido y Nombre | C.I. | Fecha | N° Boleta | Importe Bs. | Beca %

    Solo se listan estudiantes que TIENEN al menos un pago APROBADO
    aplicado al módulo (o matrícula) pedido. Si el estudiante pagó en
    varios pagos parciales, se SUMAN en un solo registro.

    El sistema incluye el % de beca del estudiante para justificar por qué
    unos pagan más que otros.
    """
    return await payment_service.generar_lista_habilitados(
        curso_id=curso_id,
        modulo_index=modulo_index,
    )
    # (inscripcion_id, estudiante_id, curso_id, _id) como objetos ObjectId. Este
    # endpoint devuelve un dict crudo (sin response_model que los coaccione, a
    # diferencia de list_payments), por lo que hay que serializarlos a string
    # explícitamente o FastAPI falla con PydanticSerializationError.
    return jsonable_encoder(
        {
            "data": enriched,
            "resumen": resultado["resumen"],
            "meta": PaginationMeta(
                page=page, limit=per_page, totalItems=resultado["total_count"], totalPages=total_pages,
                hasNextPage=(page < total_pages), hasPrevPage=(page > 1)
            )
        },
        custom_encoder={ObjectId: str}
    )


# =============================================================================
# F-078 (2026-07-24): Exportar Lista de Postgraduantes Habilitados como XLSX/PDF
# =============================================================================
# Kevin pidio: "mejores los excel como excel no como csv, y aparte que
# puedas sacar PDF, directamente los mejores pdf". El frontend antes generaba
# un CSV manual con BOM. Ahora el backend genera XLSX (openpyxl) y PDF
# (reportlab) nativos, con formato "papel Sandra":
#
# - Encabezado: titulo + tipo + programa + modulo + periodo + docente
# - Tabla con todas las columnas (N°, Nombre, CI, Estado, Fecha, N° Boleta,
#   Importe, Pendiente, Beca %)
# - Totales al pie: cant. estudiantes + total importe + total pendiente
# - PDF landscape A4 con estilos (colores por estado)
# =============================================================================
@router.get(
    "/reportes/lista-habilitados/xlsx",
    summary="F-078: Lista de Postgraduantes Habilitados - Export XLSX (estilo papel Sandra)",
)
async def get_lista_habilitados_xlsx(
    *,
    curso_id: PydanticObjectId = Query(..., description="ID del curso (obligatorio)"),
    modulo_index: Optional[int] = Query(
        None,
        ge=0,
        description=(
            "0 = solo matrícula, 1..N = solo ese módulo, None/omitido = TODOS los módulos"
        ),
    ),
    current_user: User = Depends(require_staff),
) -> Any:
    """F-078: Genera el archivo XLSX nativo (openpyxl) con el formato papel Sandra."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver informes")

    data = await payment_service.generar_lista_habilitados(
        curso_id=curso_id, modulo_index=modulo_index
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista Habilitados"

    # ====== ESTILOS ======
    title_font = Font(bold=True, size=14, color="1F4E78")
    subtitle_font = Font(bold=True, size=11, color="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell_font = Font(size=9)
    pagado_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # verde
    parcial_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amarillo
    pendiente_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # rojo
    total_font = Font(bold=True, size=10, color="1F4E78")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(border_style="thin", color="A0A0A0")
    cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # ====== ENCABEZADO ======
    enc = data["encabezado"]
    ws.merge_cells("A1:I1")
    ws["A1"] = enc["titulo"]
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:I2")
    ws["A2"] = f"{enc['programa_tipo']}: {enc['programa_nombre']}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:I3")
    ws["A3"] = f"MÓDULO: {enc['modulo']}"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = center

    ws.merge_cells("A4:I4")
    ws["A4"] = f"PERÍODO: {enc['periodo'] or 'N/A'}"
    ws["A4"].font = subtitle_font
    ws["A4"].alignment = center

    ws.merge_cells("A5:I5")
    ws["A5"] = f"DOCENTE: {enc['docente'] or 'N/A'}"
    ws["A5"].font = subtitle_font
    ws["A5"].alignment = center

    # Línea vacía
    ws.row_dimensions[6].height = 5

    # ====== TABLA DE ESTUDIANTES ======
    header_row = 7
    headers = ["N°", "Apellido y Nombre", "C.I.", "Estado", "Fecha", "N° Boleta", "Importe Bs.", "Pendiente Bs.", "Beca"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border

    rows = data.get("rows", [])
    for i, r in enumerate(rows, 1):
        row = header_row + i
        fecha = r.get("fecha_pago")
        fecha_str = ""
        if fecha:
            try:
                d = datetime.fromisoformat(fecha.replace("Z", "+00:00") if isinstance(fecha, str) else fecha.isoformat())
                fecha_str = d.strftime("%d/%m/%Y")
            except Exception:
                fecha_str = str(fecha)
        beca_str = f"{r.get('beca')} ({r.get('beca_porcentaje', 0)}%)" if r.get('beca') else "Sin beca"

        values = [
            i,
            r.get("nombre", ""),
            r.get("ci", ""),
            r.get("estado_pago", "PENDIENTE"),
            fecha_str,
            r.get("numero_boleta") or "",
            float(r.get("importe") or 0),
            float(r.get("monto_pendiente") or 0),
            beca_str,
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = cell_font
            cell.border = cell_border
            if col in (1, 4, 5, 6):
                cell.alignment = center
            elif col in (7, 8):
                cell.alignment = right
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = left
            # Color de fila segun estado
            estado = r.get("estado_pago", "")
            if estado == "PAGADO":
                cell.fill = pagado_fill
            elif estado == "PARCIAL":
                cell.fill = parcial_fill
            elif estado == "PENDIENTE":
                cell.fill = pendiente_fill

    # ====== TOTALES AL PIE ======
    total_row = header_row + len(rows) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(row=total_row, column=1, value="TOTALES").font = total_font
    ws.cell(row=total_row, column=1).alignment = right
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=7, value=float(data.get("total_importe") or 0))
    ws.cell(row=total_row, column=7).font = total_font
    ws.cell(row=total_row, column=7).fill = total_fill
    ws.cell(row=total_row, column=7).number_format = "#,##0.00"
    ws.cell(row=total_row, column=7).alignment = right
    ws.cell(row=total_row, column=7).border = cell_border
    ws.cell(row=total_row, column=8, value=float(data.get("total_pendiente") or 0))
    ws.cell(row=total_row, column=8).font = total_font
    ws.cell(row=total_row, column=8).fill = total_fill
    ws.cell(row=total_row, column=8).number_format = "#,##0.00"
    ws.cell(row=total_row, column=8).alignment = right
    ws.cell(row=total_row, column=8).border = cell_border
    ws.cell(row=total_row, column=9, value=f"{data.get('total_estudiantes', 0)} est.")
    ws.cell(row=total_row, column=9).font = total_font
    ws.cell(row=total_row, column=9).fill = total_fill
    ws.cell(row=total_row, column=9).alignment = center
    ws.cell(row=total_row, column=9).border = cell_border

    # ====== ANCHOS DE COLUMNA ======
    widths = [5, 35, 15, 12, 12, 18, 14, 14, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # Generar archivo
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Nombre del archivo
    modulo_safe = enc["modulo"].replace(" ", "_").replace(":", "").replace("/", "_")[:40]
    fname = f"lista_habilitados_{modulo_safe}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@router.get(
    "/reportes/lista-habilitados/pdf",
    summary="F-078: Lista de Postgraduantes Habilitados - Export PDF (estilo papel Sandra, landscape)",
)
async def get_lista_habilitados_pdf(
    *,
    curso_id: PydanticObjectId = Query(..., description="ID del curso (obligatorio)"),
    modulo_index: Optional[int] = Query(
        None,
        ge=0,
        description=(
            "0 = solo matrícula, 1..N = solo ese módulo, None/omitido = TODOS los módulos"
        ),
    ),
    current_user: User = Depends(require_staff),
) -> Any:
    """F-078: Genera el PDF landscape A4 con el formato papel Sandra."""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver informes")

    data = await payment_service.generar_lista_habilitados(
        curso_id=curso_id, modulo_index=modulo_index
    )

    pdf_file = BytesIO()
    doc = SimpleDocTemplate(
        pdf_file,
        pagesize=landscape(A4),
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=10*mm, bottomMargin=10*mm,
        title="Lista de Postgraduantes Habilitados - KYC DataHub",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#0c4a6e"),
        spaceAfter=2*mm, alignment=1,  # center
    )
    subtitle_style = ParagraphStyle(
        "SubtitleStyle", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#475569"),
        spaceAfter=1*mm, alignment=1,  # center
    )

    enc = data["encabezado"]
    elements = []
    elements.append(Paragraph(enc["titulo"], title_style))
    elements.append(Paragraph(
        f"<b>{enc['programa_tipo']}:</b> {enc['programa_nombre']}", subtitle_style
    ))
    elements.append(Paragraph(f"<b>MÓDULO:</b> {enc['modulo']}", subtitle_style))
    elements.append(Paragraph(f"<b>PERÍODO:</b> {enc['periodo'] or 'N/A'}", subtitle_style))
    elements.append(Paragraph(f"<b>DOCENTE:</b> {enc['docente'] or 'N/A'}", subtitle_style))
    elements.append(Spacer(1, 4*mm))

    # Tabla
    header_row = ["N°", "Apellido y Nombre", "C.I.", "Estado", "Fecha", "N° Boleta", "Importe", "Pendiente", "Beca"]
    table_data = [header_row]

    rows = data.get("rows", [])
    for i, r in enumerate(rows, 1):
        fecha = r.get("fecha_pago")
        fecha_str = ""
        if fecha:
            try:
                d = datetime.fromisoformat(
                    fecha.replace("Z", "+00:00") if isinstance(fecha, str) else fecha.isoformat()
                )
                fecha_str = d.strftime("%d/%m/%Y")
            except Exception:
                fecha_str = str(fecha)[:10]
        beca_str = f"{r.get('beca')} ({r.get('beca_porcentaje', 0)}%)" if r.get('beca') else ""

        table_data.append([
            str(i),
            r.get("nombre", ""),
            r.get("ci", ""),
            r.get("estado_pago", "PENDIENTE"),
            fecha_str,
            r.get("numero_boleta") or "",
            f"{float(r.get('importe') or 0):.2f}",
            f"{float(r.get('monto_pendiente') or 0):.2f}",
            beca_str,
        ])

    # Fila de totales
    table_data.append([
        "", "", "", "", "", "TOTALES:",
        f"{float(data.get('total_importe') or 0):.2f}",
        f"{float(data.get('total_pendiente') or 0):.2f}",
        f"{data.get('total_estudiantes', 0)} est.",
    ])

    # Anchos de columna (landscape A4 ~ 277mm utiles)
    col_widths = [
        10*mm,  # N°
        70*mm,  # Nombre
        22*mm,  # CI
        18*mm,  # Estado
        20*mm,  # Fecha
        30*mm,  # N° Boleta
        20*mm,  # Importe
        22*mm,  # Pendiente
        50*mm,  # Beca
    ]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Body
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),  # N° centrado
        ("ALIGN", (3, 1), (3, -1), "CENTER"),  # Estado centrado
        ("ALIGN", (4, 1), (4, -1), "CENTER"),  # Fecha centrado
        ("ALIGN", (5, 1), (5, -1), "CENTER"),  # N° Boleta centrado
        ("ALIGN", (6, 1), (6, -1), "RIGHT"),   # Importe right
        ("ALIGN", (7, 1), (7, -1), "RIGHT"),   # Pendiente right
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A0A0A0")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#1F4E78")),
    ]
    # Color de filas por estado
    for i, r in enumerate(rows, 1):
        estado = r.get("estado_pago", "")
        if estado == "PAGADO":
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#D1FAE5")))
        elif estado == "PARCIAL":
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEF3C7")))
        elif estado == "PENDIENTE":
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEE2E2")))

    # Fila de totales
    total_row_idx = len(rows) + 1
    style_cmds.extend([
        ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), colors.HexColor("#D9E1F2")),
        ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
        ("ALIGN", (0, total_row_idx), (5, total_row_idx), "RIGHT"),
    ])

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    # Footer
    elements.append(Spacer(1, 5*mm))
    footer_style = ParagraphStyle(
        "FooterStyle", parent=styles["Normal"],
        fontSize=7, textColor=colors.HexColor("#94a3b8"),
        alignment=1,  # center
    )
    elements.append(Paragraph(
        f"Generado por KYC DataHub el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        footer_style
    ))

    doc.build(elements)
    pdf_file.seek(0)

    modulo_safe = enc["modulo"].replace(" ", "_").replace(":", "").replace("/", "_")[:40]
    fname = f"lista_habilitados_{modulo_safe}.pdf"

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@router.get(
    "/reportes/excel",
    summary="Generar Reporte Excel de Pagos"
)
async def generar_reporte_excel_pagos(
    *,
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    curso_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por curso"),
    estudiante_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por estudiante (F-COBRANZA-003)"),
    estado: Optional[str] = Query(None, description="Filtrar por estado del pago"),
    current_user: User = Depends(require_staff)
):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from models.enrollment import Enrollment

    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para generar reportes")

    fecha_desde, fecha_hasta, fecha_desde_dt, fecha_hasta_dt = _parse_rango_fechas(fecha_desde, fecha_hasta)

    concepto_regex = None
    if current_user.rol == "cpd":
        concepto_regex = {"concepto": {"$regex": r"^matr[ií]cula$", "$options": "i"}}

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo exporta pagos de esos cursos.
    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    criteria = payment_service._construir_filtro_reporte_caja(
        fecha_desde_dt, fecha_hasta_dt, curso_id=curso_id, estudiante_id=estudiante_id, estado=estado, cursos_permitidos=cursos_permitidos
    )
    if concepto_regex:
        criteria.update(concepto_regex)
    
    payments = await Payment.find(criteria).sort("-fecha_comprobante").to_list()
    
    student_ids = list({p.estudiante_id for p in payments if p.estudiante_id})
    enrollment_ids = list({p.inscripcion_id for p in payments if p.inscripcion_id})
    curso_ids = list({p.curso_id for p in payments if p.curso_id})
    
    students_task = Student.find({"_id": {"$in": [str(s) for s in student_ids]}}).to_list()
    enrollments_task = Enrollment.find({"_id": {"$in": [str(e) for e in enrollment_ids]}}).to_list()
    courses_task = Course.find({"_id": {"$in": [str(c) for c in curso_ids]}}).to_list()
    
    students, enrollments, courses = await asyncio.gather(students_task, enrollments_task, courses_task)
    
    students_map = {s.id: s for s in students}
    enrollments_map = {e.id: e for e in enrollments}
    courses_map = {c.id: c for c in courses}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Caja"
    
    headers = ["Nombre del Estudiante", "C.I.", "Curso", "Método", "Fecha Comprobante", "Fecha Registro", "Moneda", "Monto", "Concepto", "Total Cuotas", "Nº Transacción", "Estado", "Motivo Reversión"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.auto_filter.ref = ws.dimensions
    
    from core.timezone_utils import to_bolivia_time
    for payment in payments:
        student = students_map.get(payment.estudiante_id)
        nombre_estudiante = student.nombre if student and student.nombre else "Sin nombre"
        # F-COBRANZA-042 (2026-07-22): Carnet de Identidad del estudiante.
        # Prioriza `carnet_identidad`, fallback a `registro` (compatibilidad con
        # estudiantes pre-F-027 que no tienen CI separado del registro).
        estudiante_ci = ""
        if student:
            estudiante_ci = (getattr(student, "carnet_identidad", None) or "").strip() or \
                            (getattr(student, "registro", None) or "").strip()

        course = courses_map.get(payment.curso_id)
        # F-COBRANZA-022 (2026-07-22): Joel pidio usar el codigo del programa
        # (DIPL-IA-2026) en vez del nombre largo en el XLSX, para que el reporte
        # sea mas compacto y matchee con el codigo que se ve en el sistema.
        nombre_curso = course.codigo if course and course.codigo else (course.nombre_programa if course else "Sin curso")

        total_cuotas = 0
        enrollment = enrollments_map.get(payment.inscripcion_id)
        if enrollment:
            total_cuotas = enrollment.cantidad_cuotas

        fecha_comprobante_bolivia = to_bolivia_time(payment.fecha_comprobante) if payment.fecha_comprobante else "Sin registrar"
        fecha_registro_bolivia = to_bolivia_time(payment.fecha_subida)

        # F-COBRANZA-005 (2026-07-21): los pagos anulados se exportan con monto
        # negativo en la columna "Monto", de modo que la SUMA al pie del Excel
        # (o la fórmula SUM del usuario) coincida con el extracto bancario
        # sin necesidad de restar manualmente.
        # F-048 (2026-07-22, audio Sandra): los RECHAZADOS también deben ser
        # negativos. Regla de Kevin: Débitos = anulados/rechazados, Créditos = aprobados.
        # Caso Luis Valdez (CI 5384101): pago 288 Bs RECHAZADO aparecía como +288
        # en el XLSX, confundiendo a cobranza. "Aparece como rechazado pero no
        # esta su contraparte" (audio 18:51).
        monto_exportar = payment.cantidad_pago
        if payment.estado_pago in (EstadoPago.ANULADO, EstadoPago.RECHAZADO) and monto_exportar > 0:
            monto_exportar = -monto_exportar

        row = [
            nombre_estudiante,
            estudiante_ci,
            nombre_curso,
            payment.metodo_pago,
            fecha_comprobante_bolivia,
            fecha_registro_bolivia,
            "Bs",
            monto_exportar,
            payment.concepto or "",
            total_cuotas,
            payment.numero_transaccion or "Caja / S/N",
            payment.estado_pago.value if payment.estado_pago else "",
            payment.motivo_reversion or ""
        ]
        ws.append(row)

    column_widths = [30, 12, 18, 15, 20, 20, 10, 15, 20, 15, 25, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"reporte_caja_{fecha_desde}_{fecha_hasta}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========================================================================
# F-COBRANZA-043 (2026-07-22): Reporte de Caja - Export PDF
# ========================================================================
# Kevin: "se deberia pooder tener esa opcion de descargar como pdf en el mismo
# modelo que creaste me gusto ... obviamente debe ser los mismos datos qu el
# excel que exportas mas lo delc arte que te dije ahorita" (las 4 tarjetas
# KPI: Cantidad, Total Aprobado, Total Pendiente, Total Anulado).
#
# Genera el PDF en el backend con reportlab para mantener consistencia con
# el XLSX que ya se genera acá. Landscape A4 con:
#   1. Encabezado: titulo + rango de fechas + filtros aplicados
#   2. Bloque de 4 tarjetas KPI (Cantidad, Aprobado, Pendiente, Anulado)
#   3. Tabla de pagos (mismas columnas que el XLSX)
#   4. Pie de pagina con totales
@router.get(
    "/reportes/caja/pdf",
    summary="Generar Reporte PDF de Caja (mismos datos que XLSX + tarjetas KPI)",
)
async def generar_reporte_pdf_caja(
    *,
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    curso_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por curso"),
    estudiante_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por estudiante (F-COBRANZA-003)"),
    estado: Optional[str] = Query(None, description="Filtrar por estado del pago"),
    current_user: User = Depends(require_staff)
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para generar reportes")

    fecha_desde_str, fecha_hasta_str, fecha_desde_dt, fecha_hasta_dt = _parse_rango_fechas(fecha_desde, fecha_hasta)

    concepto_regex = None
    if current_user.rol == "cpd":
        concepto_regex = {"concepto": {"$regex": r"^matr[ií]cula$", "$options": "i"}}

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    # Traer TODOS los pagos del rango (sin paginar) para el PDF completo.
    criteria = payment_service._construir_filtro_reporte_caja(
        fecha_desde_dt, fecha_hasta_dt, curso_id=curso_id, estudiante_id=estudiante_id, estado=estado, cursos_permitidos=cursos_permitidos
    )
    if concepto_regex:
        criteria.update(concepto_regex)

    # Sin limite: PDF debe incluir todos. Si en el futuro hay miles, agregar
    # parametro de paginacion o limite.
    from models.payment import Payment
    cursor = Payment.find(criteria).sort("-fecha_subida")
    payments_all = await cursor.to_list()

    # Enriquecer (nombre, C.I., curso, etc.) igual que el XLSX
    enriched = await payment_service.enrich_payments_with_details_bulk(payments_all)

    # Calcular resumen (4 tarjetas KPI)
    # enriched viene como list[dict] (de model_dump en enrich_payments_with_details_bulk),
    # así que usamos .get() directamente.
    def _g(p, key, default=None):
        return p.get(key, default) if isinstance(p, dict) else getattr(p, key, default)

    cantidad_pagos = len(enriched)
    total_aprobado = sum(float(_g(p, "cantidad_pago", 0)) for p in enriched if _g(p, "estado_pago") == "aprobado")
    total_pendiente = sum(float(_g(p, "cantidad_pago", 0)) for p in enriched if _g(p, "estado_pago") == "pendiente")
    total_anulado = sum(abs(float(_g(p, "cantidad_pago", 0))) for p in enriched if _g(p, "estado_pago") in ("anulado", "rechazado"))

    # Construir el PDF
    pdf_file = BytesIO()
    doc = SimpleDocTemplate(
        pdf_file,
        pagesize=landscape(A4),
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=10*mm, bottomMargin=10*mm,
        title="Reporte de Caja - KYC DataHub",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#0c4a6e"), spaceAfter=2*mm)
    subtitle_style = ParagraphStyle("SubtitleStyle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#475569"), spaceAfter=4*mm)

    elements = []
    elements.append(Paragraph("Reporte de Caja", title_style))
    periodo_texto = f"Período: {fecha_desde_str or 'inicio'} → {fecha_hasta_str or 'hoy'}"
    if curso_id:
        from models.course import Course
        c = await Course.get(curso_id)
        if c:
            periodo_texto += f"  |  Curso: {c.codigo or c.nombre_programa}"
    if estudiante_id:
        from models.student import Student
        s = await Student.get(estudiante_id)
        if s:
            periodo_texto += f"  |  Estudiante: {s.nombre}"
    if estado:
        periodo_texto += f"  |  Estado: {estado}"
    elements.append(Paragraph(periodo_texto, subtitle_style))

    # 4 tarjetas KPI (mismo modelo visual que el dashboard)
    kpi_data = [
        ["CANTIDAD DE PAGOS", "TOTAL APROBADO", "TOTAL PENDIENTE", "TOTAL ANULADO"],
        [str(cantidad_pagos), f"Bs. {total_aprobado:,.2f}", f"Bs. {total_pendiente:,.2f}", f"Bs. {total_anulado:,.2f}"],
    ]
    kpi_table = Table(kpi_data, colWidths=[70*mm]*4, rowHeights=[9*mm, 16*mm])
    kpi_table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Values
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (0, 1), 28),  # Cantidad
        ("FONTSIZE", (1, 1), (-1, 1), 18),  # Montos
        ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#15803d")),  # Aprobado verde
        ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#ca8a04")),  # Pendiente amarillo
        ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#b91c1c")),  # Anulado rojo
        ("ALIGN", (0, 1), (-1, 1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Borders
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 5*mm))

    # Tabla de pagos (mismas columnas que el XLSX, sin "Motivo Reversión" para
    # ahorrar espacio; solo se muestra si el filtro de estado es anulado/rechazado)
    headers = ["Nombre del Estudiante", "C.I.", "Curso", "Método", "Fecha", "Monto", "Concepto", "Nº Transacción", "Estado"]
    rows = [headers]
    from core.timezone_utils import to_bolivia_time

    # F-068 (2026-07-22, Kevin): el PDF mostraba "Sin nombre", "—", "Sin curso"
    # porque `enrich_payments_with_details_bulk` retorna `nombre_estudiante` y
    # `estudiante_ci` PLANOS (no un dict anidado con `student` y `course`).
    # Hay que construir maps propios como hace el XLSX.
    from models.student import Student
    from models.course import Course
    student_ids_pdf = list({p.get("estudiante_id") for p in enriched if p.get("estudiante_id")})
    course_ids_pdf = list({p.get("curso_id") for p in enriched if p.get("curso_id")})
    students_pdf = await Student.find({"_id": {"$in": [str(s) for s in student_ids_pdf]}}).to_list() if student_ids_pdf else []
    courses_pdf = await Course.find({"_id": {"$in": [str(c) for c in course_ids_pdf]}}).to_list() if course_ids_pdf else []
    students_map_pdf = {s.id: s for s in students_pdf}
    courses_map_pdf = {c.id: c for c in courses_pdf}

    for p in enriched:
        def _g2(key, default=None):
            return p.get(key, default) if isinstance(p, dict) else getattr(p, key, default)
        # F-068: leer de maps propios (no de `student`/`course` que no existen)
        student_obj = students_map_pdf.get(_g2("estudiante_id"))
        course_obj = courses_map_pdf.get(_g2("curso_id"))
        nombre = (student_obj.nombre if student_obj and getattr(student_obj, "nombre", None) else None) or _g2("nombre_estudiante") or "Sin nombre"
        ci = (
            (getattr(student_obj, "carnet_identidad", None) if student_obj else None)
            or (getattr(student_obj, "registro", None) if student_obj else None)
            or _g2("estudiante_ci")
            or ""
        ).strip()
        nombre_curso = ""
        if course_obj:
            nombre_curso = getattr(course_obj, "codigo", None) or getattr(course_obj, "nombre_programa", None) or ""
        if not nombre_curso:
            nombre_curso = "Sin curso"
        # to_bolivia_time retorna STRING ya formateado (ej "22/07/2026 14:30").
        # Si retorna None (no hay fecha), mostrar "Sin fecha".
        fecha = _g2("fecha_subida")
        if fecha:
            fecha_str = to_bolivia_time(fecha) if hasattr(fecha, 'isoformat') else str(fecha)
        else:
            fecha_str = "Sin fecha"
        monto = float(_g2("cantidad_pago", 0))
        estado_pago = _g2("estado_pago", "")
        # F-068: si es enum (no string), extraer .value
        if hasattr(estado_pago, "value"):
            estado_pago_str = estado_pago.value
        else:
            estado_pago_str = str(estado_pago) if estado_pago else ""
        # Anulados: mostrar como negativo (mismo criterio que el XLSX)
        if estado_pago_str == "anulado" and monto > 0:
            monto = -monto
        rows.append([
            Paragraph(str(nombre)[:40], styles["BodyText"]),
            ci or "—",
            nombre_curso,
            _g2("metodo_pago") or "",
            fecha_str,
            f"{monto:,.2f}",
            Paragraph(str(_g2("concepto") or "")[:50], styles["BodyText"]),
            str(_g2("numero_transaccion") or "Caja / S/N")[:18],
            estado_pago_str,
        ])

    data_table = Table(rows, repeatRows=1, colWidths=[
        50*mm, 18*mm, 25*mm, 20*mm, 22*mm, 18*mm, 55*mm, 25*mm, 18*mm,
    ])
    data_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0c4a6e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
    ]))
    elements.append(data_table)

    # Pie de pagina (lo agrega automaticamente reportlab al render)

    doc.build(elements)
    pdf_file.seek(0)

    filename = f"reporte_caja_{fecha_desde_str or 'inicio'}_{fecha_hasta_str or 'hoy'}.pdf"
    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========================================================================
# F-COBRANZA-023 (2026-07-22): Reporte de Caja - Formato Extracto Bancario
# ========================================================================
# Joel pidio que el reporte de caja tenga formato estilo estado de cuenta
# bancaria (como el Banco Bisa que mando como ejemplo). Estructura:
#   - Encabezado: Banco, Cuenta, Periodo, Saldo Inicial
#   - Tabla: Fecha | Comprobante | Concepto | DEBITOS | CREDITOS | Saldo
#   - Totales: Total Operaciones, Total Debitos, Total Creditos, Saldo Final
#
# Reglas contables (lo que Joel definio):
#   - CREDITOS: pagos APROBADOS (incluye los que abandonan/congelan, porque
#     su dinero SI entro al sistema). NO se filtran por estado del enrollment.
#   - DEBITOS: pagos ANULADOS o RECHAZADOS (salen del sistema).
#   - PENDIENTES: NO se muestran (estan en limbo, no son ingreso ni egreso).
#   - Saldo Final = Saldo Inicial + Total Creditos - Total Debitos
#
# El endpoint NO modifica nada, solo lee y genera XLSX con el formato pedido.

@router.get(
    "/reportes/caja/extracto-bancario",
    summary="Reporte de Caja Formato Extracto Bancario (Debitos/Creditos)",
)
async def get_extracto_bancario(
    *,
    fecha_desde: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    curso_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por curso"),
    estudiante_id: Optional[PydanticObjectId] = Query(None, description="Filtrar por estudiante"),
    current_user: User = Depends(require_staff)
):
    """
    F-COBRANZA-023: Genera un XLSX con formato extracto bancario (estilo
    Banco Bisa que Joel paso como ejemplo) para el reporte de caja.

    Reglas contables:
    - CREDITOS = pagos aprobados (incluye abandonos/congelados, su dinero si entro)
    - DEBITOS  = pagos anulados o rechazados (salen del sistema)
    - PENDIENTES NO se muestran (estan en limbo)
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from models.enrollment import Enrollment
    from core.timezone_utils import to_bolivia_time, format_fecha

    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver reportes de caja")

    fecha_desde_str, fecha_hasta_str, fecha_desde_dt, fecha_hasta_dt = _parse_rango_fechas(fecha_desde, fecha_hasta)

    concepto_regex = None
    if current_user.rol == "cpd":
        concepto_regex = {"concepto": {"$regex": r"^matr[ií]cula$", "$options": "i"}}

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    criteria = payment_service._construir_filtro_reporte_caja(
        fecha_desde_dt, fecha_hasta_dt,
        curso_id=curso_id, estudiante_id=estudiante_id,
        estado=None,  # No filtrar por estado: queremos todos para mostrar debitos y creditos
        cursos_permitidos=cursos_permitidos
    )
    if concepto_regex:
        criteria.update(concepto_regex)

    # Solo nos interesan aprobado, anulado, rechazado (NO pendientes)
    criteria["estado_pago"] = {"$in": ["aprobado", "anulado", "rechazado"]}

    # Ordenar por fecha de comprobante (ascendente para el saldo acumulado)
    payments = await Payment.find(criteria).sort("+fecha_comprobante").to_list()

    # Cargar info relacionada
    student_ids = list({p.estudiante_id for p in payments if p.estudiante_id})
    enrollment_ids = list({p.inscripcion_id for p in payments if p.inscripcion_id})
    curso_ids = list({p.curso_id for p in payments if p.curso_id})

    students_task = Student.find({"_id": {"$in": [str(s) for s in student_ids]}}).to_list()
    enrollments_task = Enrollment.find({"_id": {"$in": [str(e) for e in enrollment_ids]}}).to_list()
    courses_task = Course.find({"_id": {"$in": [str(c) for c in curso_ids]}}).to_list()

    students, enrollments, courses = await asyncio.gather(students_task, enrollments_task, courses_task)
    students_map = {s.id: s for s in students}
    courses_map = {c.id: c for c in courses}

    # Calcular resumen previo (todos los aprobados sin filtro de fecha, para
    # tener el "Saldo Inicial" = todo lo cobrado ANTES del rango)
    criterios_saldo_inicial = payment_service._construir_filtro_reporte_caja(
        None, fecha_desde_dt,
        curso_id=curso_id, estudiante_id=estudiante_id,
        estado=None, cursos_permitidos=cursos_permitidos
    )
    if concepto_regex:
        criterios_saldo_inicial.update(concepto_regex)
    criterios_saldo_inicial["estado_pago"] = "aprobado"
    pagos_previos_aprobados = await Payment.find(criterios_saldo_inicial).to_list()
    saldo_inicial = sum(p.cantidad_pago for p in pagos_previos_aprobados)

    # Calcular debitos previos (anulados antes del rango)
    criterios_deb_prev = payment_service._construir_filtro_reporte_caja(
        None, fecha_desde_dt,
        curso_id=curso_id, estudiante_id=estudiante_id,
        estado=None, cursos_permitidos=cursos_permitidos
    )
    if concepto_regex:
        criterios_deb_prev.update(concepto_regex)
    criterios_deb_prev["estado_pago"] = {"$in": ["anulado", "rechazado"]}
    pagos_previos_deb = await Payment.find(criterios_deb_prev).to_list()
    debitos_previos = sum(p.cantidad_pago for p in pagos_previos_deb)
    saldo_inicial -= debitos_previos  # ajustar por débitos previos

    # Generar XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracto Bancario"

    # ======= ESTILOS =======
    title_font = Font(bold=True, size=14, color="1F4E78")
    subtitle_font = Font(bold=True, size=11, color="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    debito_font = Font(color="C00000", bold=True)  # rojo
    credito_font = Font(color="00B050", bold=True)  # verde
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(border_style="thin", color="A0A0A0")
    cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ======= ENCABEZADO (estilo estado de cuenta) =======
    ws.merge_cells("A1:F1")
    ws["A1"] = "ESTADO DE CUENTA - SISTEMA KyC DataHub"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "Banco UAGRM - Postgrado"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Info de cuenta
    info_rows = [
        ("Cuenta:", "POSTGRADO-UAGRM"),
        ("Titular:", "UAGRM - Direccion de Posgrado"),
        ("Moneda:", "Bs (Bolivianos)"),
        ("Período:", f"{fecha_desde_str or '(inicio)'} al {fecha_hasta_str or '(hoy)'}"),
        ("Generado:", format_fecha(datetime.now(), "%Y-%m-%d %H:%M", fallback="")),
    ]
    for i, (label, value) in enumerate(info_rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2, value=value)

    # Saldo inicial
    row_saldo = 4 + len(info_rows)
    ws.cell(row=row_saldo, column=1, value="SALDO INICIAL:").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row_saldo, start_column=2, end_row=row_saldo, end_column=6)
    saldo_cell = ws.cell(row=row_saldo, column=2, value=saldo_inicial)
    saldo_cell.font = Font(bold=True, size=11)
    saldo_cell.number_format = '#,##0.00 "Bs"'

    # ======= TABLA DE MOVIMIENTOS =======
    header_row = row_saldo + 2
    headers = ["Fecha", "Comprobante", "Concepto / Descripción", "Débitos", "Créditos", "Saldo"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    # Cuerpo: una fila por pago
    saldo = saldo_inicial
    total_creditos = 0.0
    total_debitos = 0.0
    row = header_row + 1
    for p in payments:
        student = students_map.get(p.estudiante_id)
        nombre = student.nombre if student and student.nombre else "Sin nombre"
        course = courses_map.get(p.curso_id)
        codigo_curso = course.codigo if course and course.codigo else (course.nombre_programa if course else "")
        fecha = format_fecha(p.fecha_comprobante, "%Y-%m-%d", fallback="Sin fecha")
        comprobante = p.numero_transaccion or p.id or "S/N"
        # Construir concepto
        if p.estado_pago == EstadoPago.ANULADO:
            tipo_mov = "ANULACIÓN"
        elif p.estado_pago == EstadoPago.RECHAZADO:
            tipo_mov = "RECHAZO"
        else:
            tipo_mov = "PAGO"
        # Concepto: tipo + concepto del pago + nombre estudiante + código curso
        concepto_str = f"{tipo_mov} {p.concepto or 'Pago'} - {nombre}"
        if codigo_curso:
            concepto_str += f" ({codigo_curso})"
        if p.detalle:
            concepto_str += f" | {p.detalle}"
        # Débitos / Créditos
        if p.estado_pago in (EstadoPago.ANULADO, EstadoPago.RECHAZADO):
            debito = p.cantidad_pago
            credito = 0.0
            total_debitos += debito
            saldo -= debito
        else:  # aprobado
            debito = 0.0
            credito = p.cantidad_pago
            total_creditos += credito
            saldo += credito

        # Escribir fila
        ws.cell(row=row, column=1, value=fecha).border = cell_border
        ws.cell(row=row, column=2, value=str(comprobante)[:30]).border = cell_border
        ws.cell(row=row, column=3, value=concepto_str[:120]).border = cell_border
        c4 = ws.cell(row=row, column=4, value=debito if debito > 0 else None)
        c4.border = cell_border
        c4.number_format = '#,##0.00'
        if debito > 0:
            c4.font = debito_font
        c5 = ws.cell(row=row, column=5, value=credito if credito > 0 else None)
        c5.border = cell_border
        c5.number_format = '#,##0.00'
        if credito > 0:
            c5.font = credito_font
        c6 = ws.cell(row=row, column=6, value=saldo)
        c6.border = cell_border
        c6.number_format = '#,##0.00 "Bs"'
        c6.font = Font(bold=True)

        row += 1

    # Fila vacía
    row += 1

    # ======= TOTALES (estilo estado de cuenta) =======
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL OPERACIONES:").font = total_font
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.cell(row=total_row, column=4, value=len(payments)).font = total_font
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).fill = total_fill

    row += 1
    ws.cell(row=row, column=1, value="TOTAL DÉBITOS:").font = total_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=4, value=total_debitos)
    c.font = Font(bold=True, color="C00000")
    c.number_format = '#,##0.00 "Bs"'
    c.fill = total_fill

    row += 1
    ws.cell(row=row, column=1, value="TOTAL CRÉDITOS:").font = total_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=5, value=total_creditos)
    c.font = Font(bold=True, color="00B050")
    c.number_format = '#,##0.00 "Bs"'
    c.fill = total_fill

    row += 1
    ws.cell(row=row, column=1, value="SALDO FINAL:").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=6, value=saldo)
    c.font = Font(bold=True, size=12)
    c.number_format = '#,##0.00 "Bs"'
    c.fill = total_fill

    # Anchos de columna
    column_widths = [12, 18, 60, 14, 14, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Pie de pagina
    row += 3
    ws.cell(row=row, column=1, value=(
        "NOTA: Los CREDITOS incluyen todos los pagos aprobados (incluso si el "
        "estudiante abandona o congela despues, su dinero SI entro al sistema). "
        "Los DEBITOS son pagos anulados o rechazados (salen del sistema). "
        "Los pagos PENDIENTES no se muestran (estan en revision)."
    )).font = Font(italic=True, size=9, color="606060")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 35

    # Guardar
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"extracto_bancario_{fecha_desde_str or 'inicio'}_{fecha_hasta_str or 'hoy'}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class CajaDirectoRequest(BaseModel):
    estudiante_id: PydanticObjectId
    inscripcion_id: PydanticObjectId
    cantidad_pago: float = Field(..., gt=0, description="Monto cobrado en Caja (Bs)")
    concepto: Optional[str] = None
    numero_cuota: Optional[int] = None
    remitente: Optional[str] = None
    cuenta_destino: Optional[str] = None


@router.post(
    "/caja-directo",
    response_model=PaymentResponse,
    summary="Registrar Cobro Directo en Caja (requiere comprobante)"
)
async def registrar_cobro_caja_directo(
    *,
    # F-COBRANZA-026 (2026-07-22): comprobante obligatorio también para caja-directo
    file: UploadFile = File(..., description="Comprobante obligatorio (foto del recibo/factura)"),
    payload: str = Form(..., description="CajaDirectoRequest serializado como JSON string"),
    current_user: User = Depends(require_cobranza)
) -> Any:
    """
    Registrar un cobro físico directo en Caja para cualquier estudiante.
    Se crea directamente como APROBADO sin requerir la intervención o credenciales del estudiante.

    F-COBRANZA-026: ahora requiere comprobante obligatorio (foto del recibo/factura).
    """
    import json
    from core.cloudinary_utils import upload_image, upload_pdf

    # Parsear payload JSON
    try:
        payload_dict = json.loads(payload)
        payload_obj = CajaDirectoRequest(**payload_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Payload inválido: {e}")

    # F-COBRANZA-026: validar comprobante
    if not file:
        raise HTTPException(status_code=400, detail="El comprobante es obligatorio (foto del recibo/factura)")

    # Subir comprobante a Cloudinary
    try:
        from models.student import Student
        estudiante = await Student.get(payload_obj.estudiante_id)
        folder = f"payments/{payload_obj.estudiante_id}"
        public_id = f"caja_{payload_obj.inscripcion_id}_{int(datetime.now().timestamp())}"

        image_types = ["image/jpeg", "image/jpg", "image/png", "image/webp"]
        pdf_type = "application/pdf"

        if file.content_type in image_types:
            comprobante_url = await upload_image(file, folder, public_id)
        elif file.content_type == pdf_type:
            comprobante_url = await upload_pdf(file, folder, public_id)
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Formato no permitido: {file.content_type}. Use imagen o PDF"
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al subir comprobante: {e}")

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados no puede cobrar en caja
    # para inscripciones fuera de sus cursos.
    filtro_rol = filtro_cursos_por_rol(current_user)
    if filtro_rol:
        from services import enrollment_service
        target_enrollment = await enrollment_service.get_enrollment(payload_obj.inscripcion_id)
        if not target_enrollment or target_enrollment.curso_id not in filtro_rol["curso_id"]["$in"]:
            raise HTTPException(status_code=403, detail="No tienes asignado el curso de esta inscripción")

    try:
        payment = await payment_service.create_caja_directo_payment(
            estudiante_id=payload_obj.estudiante_id,
            inscripcion_id=payload_obj.inscripcion_id,
            cantidad_pago=payload_obj.cantidad_pago,
            admin_username=current_user.nombre_visible,  # ISSUE-R-PERFIL-GENERICO
            concepto=payload_obj.concepto,
            numero_cuota=payload_obj.numero_cuota,
            remitente=payload_obj.remitente,
            cuenta_destino=payload_obj.cuenta_destino,
            comprobante_url=comprobante_url,  # F-COBRANZA-026
        )
        return await payment_service.enrich_payment_with_details(payment)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al registrar cobro directo: {str(e)}")


# ========================================================================
# F-COBRANZA-016 (2026-07-21): Exportar lista de pagos a XLSX (no CSV)
# ========================================================================
# Joel pidió: "se mejoren todas las exportaciones y sean tablas, no CSV".
# Este endpoint devuelve un Excel formateado con todos los pagos que
# matcheen los filtros. Reemplaza al `downloadCSV()` que el frontend
# generaba manualmente con un Blob.

@router.get(
    "/export/excel",
    summary="Exportar lista de pagos a XLSX (reemplaza al CSV)",
)
async def export_payments_excel(
    *,
    q: Optional[str] = Query(None, description="Búsqueda por transacción o comprobante"),
    estado: Optional[str] = Query(None, description="Filtrar por estado"),
    curso_id: Optional[PydanticObjectId] = Query(None),
    estudiante_id: Optional[PydanticObjectId] = Query(None),
    tipo_concepto: Optional[str] = Query(None),
    current_user: User | Student = Depends(get_current_user)
):
    """
    F-COBRANZA-016: exporta TODOS los pagos que matcheen los filtros como
    un archivo .xlsx con formato de tabla (no CSV). Mismas reglas de RBAC
    que GET /payments/.

    Columnas:
      ID, Fecha Comprobante, Fecha Subida, Estudiante, Registro, Curso,
      Concepto, Módulo, Monto (Bs), Método, Banco, Remitente,
      Nº Transacción, Estado, Comprobante (URL)
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from models.enrollment import Enrollment
    from core.timezone_utils import format_fecha

    if isinstance(current_user, User):
        if current_user.rol not in ["superadmin", "admin", "mae", "cpd", "cobranza"]:
            raise HTTPException(status_code=403, detail="No autorizado para exportar pagos")

        filtro_rol = filtro_cursos_por_rol(current_user)
        cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

        # Traemos hasta 5,000 pagos (suficiente para producción; si crece, paginar).
        payments, _ = await payment_service.get_all_payments(
            page=1, per_page=5000, q=q, estado=estado, curso_id=curso_id,
            estudiante_id=estudiante_id, cursos_permitidos=cursos_permitidos,
            tipo_concepto=tipo_concepto,
        )

        # Filtrar RBAC adicional (CPD solo ve matrículas, etc.) — replica
        # la lógica de list_payments.
        filtered = []
        for p in payments:
            concepto_lower = (p.concepto or "").lower().strip()
            is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
            if current_user.rol == "cpd" and not is_matricula:
                continue
            filtered.append(p)
        payments = filtered
    else:
        # Estudiante solo puede exportar SUS pagos
        payments = await payment_service.get_payments_by_student(current_user.id)
        if estado and estado != "Todos los estados":
            payments = [p for p in payments if p.estado_pago.value == estado]
        if tipo_concepto:
            if tipo_concepto == "matricula":
                payments = [p for p in payments if "matricula" in (p.concepto or "").lower() or "matrícula" in (p.concepto or "").lower()]
            elif tipo_concepto == "colegiatura":
                payments = [p for p in payments if "matricula" not in (p.concepto or "").lower() and "matrícula" not in (p.concepto or "").lower()]

    # Enriquecer con estudiante + curso para el Excel
    enriched = await payment_service.enrich_payments_with_details_bulk(payments)
    student_ids = list({p.estudiante_id for p in payments if p.estudiante_id})
    enrollment_ids = list({p.inscripcion_id for p in payments if p.inscripcion_id})
    curso_ids = list({p.curso_id for p in payments if p.curso_id})

    students_task = Student.find({"_id": {"$in": [str(s) for s in student_ids]}}).to_list()
    enrollments_task = Enrollment.find({"_id": {"$in": [str(e) for e in enrollment_ids]}}).to_list()
    courses_task = Course.find({"_id": {"$in": [str(c) for c in curso_ids]}}).to_list()
    students, enrollments, courses = await asyncio.gather(students_task, enrollments_task, courses_task)
    students_map = {s.id: s for s in students}
    courses_map = {c.id: c for c in courses}

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    headers = [
        "ID", "Fecha Comprobante", "Fecha Subida", "Estudiante", "C.I.",
        "Curso", "Concepto", "Detalle (Desglose)", "Nº Módulo", "Tipo Movimiento",
        "Débito (Bs)", "Crédito (Bs)", "Monto (Bs)", "Método", "Banco",
        "Remitente", "Nº Transacción", "Estado", "Comprobante URL",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for p in enriched:
        student = students_map.get(p.get("estudiante_id"))
        student_name = student.nombre if student and student.nombre else "Sin nombre"
        # F-COBRANZA-036 (2026-07-22): Sandra pidio columna C.I. en el reporte
        # de caja. Usamos el campo `estudiante_ci` que enrich_payments_with_details_bulk
        # ya lleno (prioriza carnet_identidad; si no hay, cae al registro).
        student_ci = p.get("estudiante_ci") or ""
        course = courses_map.get(p.get("curso_id"))
        # F-COBRANZA-022 (2026-07-22): Joel pidio usar el codigo del programa
        # (DIPL-IA-2026) en vez del nombre largo en el XLSX, para que el reporte
        # sea mas compacto y matchee con el codigo que se ve en el sistema.
        course_name = (course.codigo if course and course.codigo else (course.nombre_programa if course and course.nombre_programa else "Sin curso"))
        modulo = p.get("numero_cuota") or "N/A"

        row = [
            str(p.get("_id", "")),
            format_fecha(p.get("fecha_comprobante"), "%Y-%m-%d", fallback="Sin registrar"),
            format_fecha(p.get("fecha_subida"), "%Y-%m-%d %H:%M", fallback=""),
            student_name,
            student_ci,
            course_name,
            p.get("concepto") or "",
            p.get("detalle") or "",  # F-COBRANZA-020: desglose separado
            modulo,
            # F-COBRANZA-037 (2026-07-22): Sandra pidio columnas Debito/Credito
            # para ver la diferencia entre pagos y anulaciones. Los rechazos
            # tambien van a Debito. Los pagos aprobados van a Credito.
            p.get("tipo_movimiento") or "PAGO",
            float(p.get("debito") or 0),
            float(p.get("credito") or 0),
            p.get("cantidad_pago", 0),
            p.get("metodo_pago") or "Transferencia",
            p.get("banco") or "Caja UAGRM",
            p.get("remitente") or "",
            p.get("numero_transaccion") or "S/N",
            # F-COBRANZA-016 fix (2026-07-22): el campo se llama `estado` en
            # el dict enriquecido (devuelto por enrich_payments_with_details_bulk
            # que ya lo convierte a .value del enum). Antes usaba `estado_pago`
            # que devolvía el enum crudo y se mostraba como "EstadoPago.APROBADO"
            # en el XLSX. Bug detectado por Joel al abrir el Excel descargado.
            p.get("estado") or "",
            p.get("comprobante_url") or "",
        ]
        ws.append(row)

    # Auto-width básico
    column_widths = [28, 14, 18, 30, 12, 28, 22, 10, 12, 14, 18, 22, 18, 14, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"pagos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================================
# F-074 (2026-07-23): VISTA MATRICIAL DE PAGOS
# ============================================================================
# Endpoints para vista alternativa de Gestión de Pagos (estilo Excel de Sandra).
# Filas = estudiantes, columnas = MATRÍCULA | MODULO 1..N | TOTAL INGRESOS | POR COBRAR.
# Importante: rutas estáticas (matriz, resumen-modulos) declaradas ANTES de /{payment_id}
# para evitar el bug FastAPI de matchear la palabra como ObjectId (F-070-FIX-2).


@router.get(
    "/matriz",
    summary="F-074: Matriz estudiante-vs-módulos (vista alternativa Gestión de Pagos)",
)
async def get_matriz_pagos_endpoint(
    modulo_index: Optional[int] = Query(
        None,
        ge=0,
        description="Si viene, devuelve solo esa columna de módulo (0=matrícula virtual, 1=Módulo 1, etc.)",
    ),
    current_user: User = Depends(require_staff),
) -> Any:
    """
    Devuelve la matriz de pagos estilo Excel de Sandra: cada estudiante con
    el detalle de cuánto pagó en matrícula y en cada módulo. Respeta la
    segmentación por curso del rol (Cobranza con cursos_asignados solo ve
    su alcance).
    """
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver la matriz de pagos")

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    return await payment_service.get_matriz_pagos(
        cursos_permitidos=cursos_permitidos,
        modulo_index=modulo_index,
    )


@router.get(
    "/matriz/por-pago",
    summary="F-087: Vista 'Por Pago' de Gestión de Pagos (1 fila por pago individual)",
)
async def get_matriz_por_pago_endpoint(
    curso_id: Optional[str] = Query(
        None,
        description="Filtrar por curso. None = todos los cursos del alcance del usuario",
    ),
    modulo_index: Optional[int] = Query(
        None,
        ge=0,
        description="Filtrar por módulo (0=matrícula, 1..N=módulos)",
    ),
    estado_pago: Optional[str] = Query(
        None,
        description="Filtrar por estado: aprobado | pendiente | rechazado | anulado",
    ),
    subido_por: Optional[str] = Query(
        None,
        description='Filtrar por quién subió: "estudiante" | "encargado"',
    ),
    page: int = Query(1, ge=1, description="Número de página"),
    per_page: int = Query(50, ge=1, le=500, description="Resultados por página (max 500)"),
    current_user: User = Depends(require_staff),
) -> Any:
    """
    F-087 (2026-07-28): 3ra vista de Gestión de Pagos. A diferencia de la
    matriz tradicional (que agrupa por estudiante y suma por módulo), esta
    vista muestra UNA FILA POR CADA PAGO INDIVIDUAL, para que Kevin/Sandra
    puedan auditar el origen de cada Bs.

    Cada fila incluye: estudiante, CI, curso, módulo (parseado del concepto),
    monto, fecha_subida, fecha_comprobante, número de transacción, comprobante
    URL, quién subió (estudiante/encargado/null), método de pago, banco,
    remitente, y estado.

    La columna modulo_index intenta ser precisa:
    - Si concepto = "Matrícula" → modulo_index = 0
    - Si concepto = "Pago Módulo 1" → modulo_index = 1
    - Si concepto = "Pago Módulos 1, 2" → modulo_index = 1 y se crea una segunda
      fila con modulo_index = 2 (split prorrateado del monto, igual que la
      matriz tradicional).

    Nota sobre sub-pagos: el split es una vista, no un cambio en BD. El pago
    real sigue siendo 1 documento. Si en el futuro se quiere granularidad
    estricta, se debe modelar a nivel de sub-pagos en el schema.

    Reglas:
    - Permiso: `puede_ver_economico` (mismo que la matriz).
    - Respeta segmentación de cursos por rol (cobranza con cursos_asignados).
    - Ordena por fecha_subida DESC (lo más reciente primero).
    - Paginación: page/per_page. Default 50 por página.
    """
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver pagos por-pago")

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    from beanie import PydanticObjectId as _POI
    curso_oid = _POI(curso_id) if curso_id else None

    return await payment_service.get_matriz_por_pago(
        cursos_permitidos=cursos_permitidos,
        curso_id=curso_oid,
        modulo_index=modulo_index,
        estado_pago=estado_pago,
        subido_por=subido_por,
        page=page,
        per_page=per_page,
    )


@router.get(
    "/resumen-modulos",
    summary="F-074: Resumen por módulo (KPI cards vista Matriz)",
)
async def get_resumen_modulos_endpoint(
    current_user: User = Depends(require_staff),
) -> Any:
    """
    Resumen agregado por módulo (cantidad de pagos, monto total, monto
    pendiente, estudiantes cursando). Excluye suspendidos (regla F-073).
    """
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver el resumen por módulos")

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    return await payment_service.get_resumen_modulos(
        cursos_permitidos=cursos_permitidos,
    )


# =============================================================================
# F-088 (2026-07-29): Vista "Deudores" unificada para Cobranza
# =============================================================================
# Reunión 2026-07-29 con Lic. Sandra Zabala: pidió una vista a "un solo golpe
# visual" donde pueda ver, para un curso, qué estudiantes deben qué módulos.
# Hoy tiene que descargar módulo por módulo en Excel y filtrar manualmente los
# que no pagaron (lo cual es lento y propenso a errores).
#
# Aquí: estudiantes como filas, módulos como columnas, con un check visual
# (verde = pagado, rojo = debe, gris = no_le_toca). Filtro "solo deudores"
# para enfocarse solo en los que deben algo. Botón "Exportar a Excel" que
# genera el mismo layout para enviar por WhatsApp / imprimir.
#
# Permisos: solo personal económico (puede_ver_economico).
# =============================================================================
@router.get(
    "/deudores",
    summary="F-088: Vista 'Deudores' unificada (estudiantes × módulos, con filtro solo deudores)",
)
async def get_deudores_endpoint(
    curso_id: str = Query(..., description="ID del curso (obligatorio)"),
    solo_deudores: bool = Query(
        True,
        description="Si True, solo retorna estudiantes con deuda_total > 0. Si False, retorna todos los inscritos.",
    ),
    current_user: User = Depends(require_staff),
) -> Any:
    """
    F-088 (2026-07-29): vista "Deudores" unificada para cobranza.

    Devuelve una matriz de estudiantes vs módulos del curso, con:
    - Estado por celda (pagado / parcial / debe / no_le_toca)
    - Datos de contacto (celular, email) para enviar WhatsApp directo
    - Total de deuda por fila y resumen por columna
    - Filtro `solo_deudores` para enfocarse en los que deben algo

    Permisos: mismo que la matriz (puede_ver_economico). Respeta
    segmentación de cursos por rol (cobranza con cursos_asignados solo ve
    sus cursos asignados).
    """
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para ver deudores")

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    from beanie import PydanticObjectId as _POI
    try:
        curso_oid = _POI(curso_id)
    except Exception:
        raise HTTPException(status_code=400, detail="curso_id inválido")

    try:
        return await payment_service.get_matriz_deudores(
            curso_id=curso_oid,
            cursos_permitidos=cursos_permitidos,
            solo_deudores=solo_deudores,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get(
    "/deudores/export-excel",
    summary="F-088: Exportar vista 'Deudores' a XLSX (mismo layout que la vista)",
)
async def export_deudores_excel(
    curso_id: str = Query(..., description="ID del curso (obligatorio)"),
    solo_deudores: bool = Query(True, description="Si True, solo exporta deudores"),
    current_user: User = Depends(require_staff),
) -> Any:
    """
    F-088 (2026-07-29): exporta la vista deudores a XLSX.

    Layout del Excel:
    - Header: nombre del curso + fecha de generación
    - Columnas: Estudiante | CI | Celular | Email | Matrícula | Módulo 1 | M2 | ... | M N | Deuda Total
    - Filas: estudiantes
    - Celdas:
      - Matrícula / Módulos: "Bs. X / Bs. Y" (pagado / costo) + check visual con color
      - Verde si pagado completo, rojo si debe, gris si no_le_toca
    - Fila TOTAL al final con cuánto se debe por columna
    """
    if not puede_ver_economico(current_user):
        raise HTTPException(status_code=403, detail="No autorizado para exportar deudores")

    filtro_rol = filtro_cursos_por_rol(current_user)
    cursos_permitidos = filtro_rol["curso_id"]["$in"] if filtro_rol else None

    from beanie import PydanticObjectId as _POI
    try:
        curso_oid = _POI(curso_id)
    except Exception:
        raise HTTPException(status_code=400, detail="curso_id inválido")

    try:
        data = await payment_service.get_matriz_deudores(
            curso_id=curso_oid,
            cursos_permitidos=cursos_permitidos,
            solo_deudores=solo_deudores,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Generar XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Deudores"

    # Estilos
    font_header = Font(bold=True, color="FFFFFF", size=12)
    font_subheader = Font(bold=True, color="FFFFFF", size=10)
    font_cell = Font(size=10)
    font_total = Font(bold=True, size=10)
    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_pagado = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # verde
    fill_debe = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # rojo
    fill_no_le_toca = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # gris
    fill_total = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber
    border_thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Header del reporte
    curso_nombre = data["curso"]["nombre"]
    curso_codigo = data["curso"]["codigo"] or ""
    fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M")
    total_estudiantes = data["resumen"]["total_estudiantes"]
    total_deudores = data["resumen"]["total_deudores"]
    deuda_total = data["resumen"]["deuda_total_curso"]

    ws["A1"] = f"REPORTE DE DEUDORES — {curso_nombre}"
    ws["A2"] = f"Código: {curso_codigo}  |  Generado: {fecha_gen}"
    ws["A3"] = f"Estudiantes: {total_estudiantes}  |  Deudores: {total_deudores}  |  Deuda total: Bs. {deuda_total:,.2f}"
    if solo_deudores:
        ws["A4"] = "Filtro aplicado: SOLO DEUDORES"
    else:
        ws["A4"] = "Filtro aplicado: TODOS LOS INSCRITOS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"].font = Font(italic=True, size=10)
    ws["A4"].font = Font(italic=True, size=10, color="6B7280")

    # Header de la tabla (fila 6)
    header_row = 6
    cols = ["#", "Estudiante", "CI", "Celular", "Email", "Registro", "Matrícula"]
    # F-XXX (2026-07-29): columnas como "Módulo 1, 2..." en vez del nombre
    # largo del módulo. El nombre real queda en el atributo "title" del
    # header cuando se abre en Excel, o se puede consultar en la columna
    # auxiliar "Módulo N (nombre)".
    for idx_mod, _m in enumerate(data["curso"]["modulos"], start=1):
        cols.append(f"Módulo {idx_mod}")
    cols.append("Deuda Total")
    cols.append("Módulos que debe")

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = font_subheader
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Filas de estudiantes
    for i, est in enumerate(data["estudiantes"], start=1):
        row = header_row + i
        # # (correlativo)
        c = ws.cell(row=row, column=1, value=i)
        c.alignment = align_center
        c.border = border_thin
        c.font = font_cell
        # Estudiante
        c = ws.cell(row=row, column=2, value=est["nombre"])
        c.alignment = align_left
        c.border = border_thin
        c.font = font_cell
        # CI
        c = ws.cell(row=row, column=3, value=est["ci"])
        c.alignment = align_center
        c.border = border_thin
        c.font = font_cell
        # Celular
        c = ws.cell(row=row, column=4, value=est["celular"])
        c.alignment = align_center
        c.border = border_thin
        c.font = font_cell
        # Email
        c = ws.cell(row=row, column=5, value=est["email"])
        c.alignment = align_left
        c.border = border_thin
        c.font = font_cell
        # Registro
        c = ws.cell(row=row, column=6, value=est["registro"])
        c.alignment = align_center
        c.border = border_thin
        c.font = font_cell
        # Matrícula
        mat = est["matricula"]
        mat_label = "—" if mat["estado"] == "no_le_toca" else f"Bs. {mat['pagado']:,.2f} / Bs. {mat['costo']:,.2f}"
        c = ws.cell(row=row, column=7, value=mat_label)
        c.alignment = align_center
        c.border = border_thin
        c.font = font_cell
        if mat["estado"] == "pagado":
            c.fill = fill_pagado
        elif mat["estado"] == "debe":
            c.fill = fill_debe
        else:
            c.fill = fill_no_le_toca
        # Módulos
        for j, mod in enumerate(est["modulos"], start=8):
            mod_label = "—" if mod["estado"] == "no_le_toca" else f"Bs. {mod['pagado']:,.2f} / Bs. {mod['costo']:,.2f}"
            c = ws.cell(row=row, column=j, value=mod_label)
            c.alignment = align_center
            c.border = border_thin
            c.font = font_cell
            if mod["estado"] == "pagado":
                c.fill = fill_pagado
            elif mod["estado"] == "debe":
                c.fill = fill_debe
            else:
                c.fill = fill_no_le_toca
        # Deuda total
        col_deuda = 8 + len(est["modulos"])
        deuda_label = f"Bs. {est['deuda_total']:,.2f}" if est["deuda_total"] > 0.01 else "—"
        c = ws.cell(row=row, column=col_deuda, value=deuda_label)
        c.alignment = align_right
        c.border = border_thin
        c.font = font_total
        if est["deuda_total"] > 0.01:
            c.fill = fill_debe
        # Módulos pendientes
        col_mod_pend = col_deuda + 1
        if est["modulos_pendientes"]:
            mod_pend_label = ", ".join(f"M{m}" for m in est["modulos_pendientes"])
        else:
            mod_pend_label = "—"
        c = ws.cell(row=row, column=col_mod_pend, value=mod_pend_label)
        c.alignment = align_center
        c.border = border_thin
        c.font = font_cell

    # Fila TOTAL
    total_row = header_row + len(data["estudiantes"]) + 1
    if data["estudiantes"]:
        # Celda vacía para las primeras 6 columnas
        for col_idx in range(1, 7):
            c = ws.cell(row=total_row, column=col_idx, value="")
            c.fill = fill_total
            c.border = border_thin
        # Label "TOTAL DEUDA"
        c = ws.cell(row=total_row, column=7, value="TOTAL DEUDA →")
        c.font = Font(bold=True, size=10)
        c.alignment = align_right
        c.fill = fill_total
        c.border = border_thin
        # Suma por columna de módulos
        for j, mod in enumerate(data["curso"]["modulos"], start=8):
            col_total = sum(
                est["modulos"][j - 8]["pendiente"]
                for est in data["estudiantes"]
                if j - 8 < len(est["modulos"])
            )
            c = ws.cell(row=total_row, column=j, value=f"Bs. {col_total:,.2f}" if col_total > 0.01 else "—")
            c.font = font_total
            c.alignment = align_center
            c.fill = fill_total
            c.border = border_thin
        # Deuda total
        col_deuda = 8 + len(data["curso"]["modulos"])
        c = ws.cell(row=total_row, column=col_deuda, value=f"Bs. {deuda_total:,.2f}")
        c.font = Font(bold=True, size=11, color="DC2626")
        c.alignment = align_right
        c.fill = fill_total
        c.border = border_thin
        # Última celda vacía
        c = ws.cell(row=total_row, column=col_deuda + 1, value="")
        c.fill = fill_total
        c.border = border_thin

    # Anchos de columna
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22  # Matrícula
    for j in range(8, 8 + len(data["curso"]["modulos"])):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.column_dimensions[get_column_letter(8 + len(data["curso"]["modulos"]))].width = 16
    ws.column_dimensions[get_column_letter(9 + len(data["curso"]["modulos"]))].width = 22

    # Altura del header
    ws.row_dimensions[header_row].height = 30

    # Generar el archivo en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Nombre del archivo
    safe_nombre = (curso_codigo or "curso").replace("/", "_").replace(" ", "_")
    filename = f"deudores_{safe_nombre}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/me",
    response_model=List[PaymentResponse],
    summary="Ver Mis Pagos (Estudiante autenticado)"
)
async def get_my_payments(
    current_user: Student = Depends(get_current_user)
) -> Any:
    """
    FIX-ERRORES-500: lista los pagos del estudiante autenticado.
    Importante: este endpoint debe declararse ANTES de /{id} para que
    no se matchee con id="me" (que rompe PydanticObjectId).
    """
    from beanie import PydanticObjectId
    payments = await Payment.find(
        Payment.estudiante_id == PydanticObjectId(current_user.id)
    ).sort("-created_at").to_list()
    return payments


@router.get(
    "/{id}",
    response_model=PaymentResponse,
    summary="Ver Pago"
)
async def get_payment(
    *,
    id: PydanticObjectId,
    current_user: User | Student = Depends(get_current_user)
) -> Any:
    payment = await payment_service.get_payment(id)
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    if isinstance(current_user, Student):
        if payment.estudiante_id != current_user.id:
            raise HTTPException(
                status_code=403,
                detail="No tienes permiso para ver este pago"
            )
            
    if isinstance(current_user, User):
        # AUDITORÍA (CRÍTICO #1): mismo guard general que en list_payments.
        if current_user.rol not in ["superadmin", "admin", "mae", "cpd", "cobranza"]:
            raise HTTPException(status_code=403, detail="No autorizado para ver este pago")

        # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados no puede ver pagos de otros cursos.
        filtro_rol = filtro_cursos_por_rol(current_user)
        if filtro_rol and payment.curso_id not in filtro_rol["curso_id"]["$in"]:
            raise HTTPException(status_code=403, detail="No tienes asignado el curso de este pago")

        concepto_lower = (payment.concepto or "").lower().strip()
        is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
        
        if current_user.rol == "cpd" and not is_matricula:
            raise HTTPException(status_code=403, detail="El rol CPD solo tiene acceso a pagos de concepto Matrícula")
    
    return await payment_service.enrich_payment_with_details(payment)


@router.put(
    "/{id}/aprobar",
    response_model=PaymentResponse,
    summary="Aprobar Pago"
)
async def aprobar_pago(
    *,
    id: PydanticObjectId,
    current_user: User = Depends(require_staff)
) -> Any:
    if current_user.rol not in ["superadmin", "admin", "cpd", "cobranza"]:
        raise HTTPException(status_code=403, detail="Su rol no tiene permisos para aprobar pagos")
        
    payment = await payment_service.get_payment(id)
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados no puede aprobar pagos de otros cursos.
    filtro_rol = filtro_cursos_por_rol(current_user)
    if filtro_rol and payment.curso_id not in filtro_rol["curso_id"]["$in"]:
        raise HTTPException(status_code=403, detail="No tienes asignado el curso de este pago")
        
    concepto_lower = (payment.concepto or "").lower().strip()
    is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
    
    if current_user.rol == "cpd" and not is_matricula:
        raise HTTPException(status_code=403, detail="El rol CPD solo puede aprobar pagos con concepto de Matrícula.")
        
    try:
        payment = await payment_service.aprobar_pago(
            payment_id=id,
            # ISSUE-R-PERFIL-GENERICO: nombre_visible en vez de username, para
            # que Cobranza (rol rotativo) quede identificado por función.
            admin_username=current_user.nombre_visible
        )
        return await payment_service.enrich_payment_with_details(payment)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put(
    "/{id}/rechazar",
    response_model=PaymentResponse,
    summary="Rechazar Pago"
)
async def rechazar_pago(
    *,
    id: PydanticObjectId,
    rejection: PaymentRejection,
    current_user: User = Depends(require_staff)
) -> Any:
    if current_user.rol not in ["superadmin", "admin", "cpd", "cobranza"]:
        raise HTTPException(status_code=403, detail="Su rol no tiene permisos para rechazar pagos")
        
    payment = await payment_service.get_payment(id)
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados no puede rechazar pagos de otros cursos.
    filtro_rol = filtro_cursos_por_rol(current_user)
    if filtro_rol and payment.curso_id not in filtro_rol["curso_id"]["$in"]:
        raise HTTPException(status_code=403, detail="No tienes asignado el curso de este pago")
        
    concepto_lower = (payment.concepto or "").lower().strip()
    is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
    
    if current_user.rol == "cpd" and not is_matricula:
        raise HTTPException(status_code=403, detail="El rol CPD solo puede rechazar pagos con concepto de Matrícula.")
        
    try:
        payment = await payment_service.rechazar_pago(
            payment_id=id,
            admin_username=current_user.nombre_visible,  # ISSUE-R-PERFIL-GENERICO
            motivo=rejection.motivo
        )
        return await payment_service.enrich_payment_with_details(payment)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put(
    "/{id}/anular",
    response_model=PaymentResponse,
    summary="Anular Pago (Rollback Financiero)"
)
async def anular_pago(
    *,
    id: PydanticObjectId,
    reversion: PaymentReversion,
    current_user: User = Depends(require_staff)
) -> Any:
    """
    ISSUE-P-CANALES: Anula un pago previamente aprobado y restaura las deudas del estudiante.
    Solo disponible para Cobranzas, Admin y SuperAdmin. El CPD no maneja flujos de caja.
    """
    if current_user.rol not in ["superadmin", "admin", "cobranza"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="Solo el personal financiero (Cobranzas/Administrador) puede realizar reversiones de caja."
        )

    # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados no puede anular pagos de otros cursos.
    filtro_rol = filtro_cursos_por_rol(current_user)
    if filtro_rol:
        target_payment = await payment_service.get_payment(id)
        if not target_payment or target_payment.curso_id not in filtro_rol["curso_id"]["$in"]:
            raise HTTPException(status_code=403, detail="No tienes asignado el curso de este pago")
        
    try:
        payment = await payment_service.anular_pago(
            payment_id=id,
            admin_username=current_user.nombre_visible,  # ISSUE-R-PERFIL-GENERICO
            motivo=reversion.motivo
        )
        return await payment_service.enrich_payment_with_details(payment)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post(
    "/{id}/upload-by-encargado",
    response_model=PaymentResponse,
    summary="Subir Comprobante de Pago (Encargado de Programa)"
)
async def upload_comprobante_by_encargado(
    *,
    id: PydanticObjectId,
    file: UploadFile = File(..., description="Comprobante de pago (imagen o PDF)"),
    numero_transaccion: Optional[str] = Form(None, description="Número de transacción (opcional)"),
    remitente: Optional[str] = Form(None, description="Remitente del pago (opcional)"),
    fecha_comprobante: Optional[str] = Form(None, description="Fecha del comprobante (YYYY-MM-DD, opcional)"),
    current_user: User = Depends(require_staff)
) -> Any:
    """
    F-COBRANZA-011 (2026-07-21): el personal de COBRANZA puede subir el
    comprobante de pago del estudiante cuando este no puede hacerlo por
    sí mismo (problemas técnicos, falta de acceso, etc.).

    Roles permitidos: SUPERADMIN, ADMIN, COBRANZA.
    Roles NO permitidos: CPD, COORDINADOR, ENCARGADO_CURSO, DOCENTE, ESTUDIANTE.

    Decisión de Joel (2026-07-21 20:30): "debería subirlo el de cobranzas, y
    que esté en el modal de gestión de pagos [...] no esté en el del encargado
    porque sería confuncion por ahora". Encargado de programa NO sube: lo hace
    cobranza. La UI expone este endpoint solo en /app/payments con el botón
    "Subir comprobante del estudiante".

    Diferencias vs `create_payment`:
    - El pago YA EXISTE (creado por el estudiante con o sin comprobante).
    - Solo se actualiza el comprobante_url y datos opcionales.
    - Se registra en auditoría con `subido_por=cobranza_id`.
    - El estudiante ve en su perfil quién subió el comprobante.
    - Al subir, el pago ya estaba APROBADO (F-COBRANZA-004), así que el saldo
      del enrollment NO se vuelve a tocar.
    """
    # 1. Validar rol: SOLO personal financiero (cobranza) y administrativos.
    roles_permitidos = ["superadmin", "admin", "cobranza"]
    if current_user.rol not in roles_permitidos:
        raise HTTPException(
            status_code=403,
            detail=f"Su rol ({current_user.rol}) no puede subir comprobantes en nombre de estudiantes. Solo cobranza, admin y superadmin están autorizados."
        )

    # 2. Obtener el pago
    payment = await payment_service.get_payment(id)
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    # 3. ISSUE-P-SEGMENTACION: encargado con cursos_asignados solo puede
    #    subir comprobantes de SUS cursos asignados.
    filtro_rol = filtro_cursos_por_rol(current_user)
    if filtro_rol and payment.curso_id not in filtro_rol["curso_id"]["$in"]:
        raise HTTPException(
            status_code=403,
            detail="No tienes asignado el curso de este pago"
        )

    # 4. Validar que el pago no esté anulado
    if payment.estado_pago == EstadoPago.ANULADO:
        raise HTTPException(
            status_code=400,
            detail="No se puede subir comprobante a un pago anulado."
        )

    # 5. Subir el archivo a Cloudinary
    from core.cloudinary_utils import upload_image, upload_pdf
    folder = f"payments/{payment.estudiante_id}"
    safe_transaction = (numero_transaccion or f"encargado_{current_user.id}").replace(' ', '_').replace('/', '_')
    public_id = f"voucher_{safe_transaction}"

    image_types = ["image/jpeg", "image/jpg", "image/png", "image/webp"]
    pdf_type = "application/pdf"

    try:
        if file.content_type in image_types:
            comprobante_url = await upload_image(file, folder, public_id)
        elif file.content_type == pdf_type:
            from core.cloudinary_utils import upload_pdf
            comprobante_url = await upload_pdf(file, folder, public_id)
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Tipo de archivo no soportado: {file.content_type}. Use JPEG, PNG, WEBP o PDF."
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al subir el archivo: {str(e)}"
        )

    # 6. Actualizar el pago
    from datetime import datetime
    from core.timezone_utils import utcnow_naive
    try:
        update_dict = {
            "comprobante_url": comprobante_url,
            "updated_at": utcnow_naive()
        }
        if numero_transaccion:
            update_dict["numero_transaccion"] = numero_transaccion
        if remitente:
            update_dict["remitente"] = remitente
        if fecha_comprobante:
            try:
                update_dict["fecha_comprobante"] = datetime.strptime(fecha_comprobante, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="fecha_comprobante debe tener formato YYYY-MM-DD"
                )

        await Payment.find_one({"_id": id}).update({"$set": update_dict})

        # F-087 (2026-07-28): marca el pago como subido por encargado, ya sea
        # que el comprobante sea nuevo o que se esté reemplazando uno anterior.
        # La acción la hizo el encargado, no el estudiante, así que este campo
        # refleja al responsable más reciente de la subida.
        await Payment.find_one({"_id": id}).update({"$set": {"subido_por": "encargado"}})

        # Re-leer el pago actualizado
        payment = await payment_service.get_payment(id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar el pago: {str(e)}")

    # 7. Audit log
    await payment_service._registrar_auditoria_financiera(
        accion="UPLOAD COMPROBANTE BY ENCARGADO",
        payment_id=payment.id,
        estudiante_id=payment.estudiante_id,
        monto=payment.cantidad_pago,
        admin_username=current_user.nombre_visible,
        detalles=f"Comprobante subido por {current_user.nombre_visible} (rol={current_user.rol}) en nombre del estudiante. URL={comprobante_url[:80]}..."
    )

    # 8. Notificar al estudiante
    try:
        from services.notification_service import create_notification
        await create_notification(
            destinatario_id=payment.estudiante_id,
            tipo_destinatario="student",
            titulo="Comprobante subido por tu encargado",
            mensaje=f"El encargado {current_user.nombre_visible} subió el comprobante de tu pago de Bs. {payment.cantidad_pago} por el concepto '{payment.concepto}'.",
            tipo_alerta="info",
            ruta="/app/payments",
            referencia_tipo="payment",
            referencia_id=payment.id
        )
    except Exception as e:
        print(f"Error al enviar notificación de comprobante subido: {str(e)}")

    return await payment_service.enrich_payment_with_details(payment)


@router.delete(
    "/{id}",
    summary="Eliminar Pago (Borrado Definitivo)"
)
async def eliminar_pago(
    *,
    id: PydanticObjectId,
    current_user: User = Depends(require_superadmin)
) -> Any:
    """
    Elimina físicamente un pago de la base de datos. Operación destructiva y
    financiera, restringida a SUPERADMIN (mismo criterio que eliminar usuarios
    o cursos). Pensado para limpiar pagos de prueba o registros erróneos que no
    deben computar en la contabilidad.

    Tras el borrado se recalcula el saldo/estado de la inscripción desde los
    pagos APROBADOS restantes, para que los totales económicos queden
    consistentes.
    """
    payment = await payment_service.get_payment(id)
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    try:
        resultado = await payment_service.eliminar_pago(
            payment_id=id,
            admin_username=current_user.nombre_visible  # ISSUE-R-PERFIL-GENERICO
        )
        return resultado
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/enrollment/{enrollment_id}", response_model=List[PaymentResponse])
async def get_payments_by_enrollment(
    *,
    enrollment_id: PydanticObjectId,
    current_user: User | Student = Depends(get_current_user)
) -> Any:
    if isinstance(current_user, Student):
        from services import enrollment_service
        enrollment = await enrollment_service.get_enrollment(enrollment_id)
        if not enrollment:
            raise HTTPException(status_code=404, detail="Inscripción no encontrada")
        if enrollment.estudiante_id != current_user.id:
            raise HTTPException(status_code=403, detail="No tienes permiso")
            
    if isinstance(current_user, User):
        if current_user.rol not in ["superadmin", "admin", "mae", "cpd", "cobranza"]:
            raise HTTPException(status_code=403, detail="No autorizado para ver los pagos de esta inscripción")

        # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo ve inscripciones de esos cursos.
        filtro_rol = filtro_cursos_por_rol(current_user)
        if filtro_rol:
            from services import enrollment_service
            target_enrollment = await enrollment_service.get_enrollment(enrollment_id)
            if not target_enrollment or target_enrollment.curso_id not in filtro_rol["curso_id"]["$in"]:
                raise HTTPException(status_code=403, detail="No tienes asignado el curso de esta inscripción")
    
    payments = await payment_service.get_payments_by_enrollment(enrollment_id)
    
    filtered_payments = []
    for p in payments:
        if isinstance(current_user, User):
            concepto_lower = (p.concepto or "").lower().strip()
            is_matricula = "matricula" in concepto_lower or "matrícula" in concepto_lower
            if current_user.rol == "cpd" and not is_matricula:
                continue
        filtered_payments.append(p)
        
    return await payment_service.enrich_payments_with_details_bulk(filtered_payments)


@router.get("/enrollment/{enrollment_id}/resumen")
async def get_resumen_pagos(
    *,
    enrollment_id: PydanticObjectId,
    current_user: User | Student = Depends(get_current_user)
) -> Any:
    if isinstance(current_user, Student):
        from services import enrollment_service
        enrollment = await enrollment_service.get_enrollment(enrollment_id)
        if not enrollment:
            raise HTTPException(status_code=404, detail="Inscripción no encontrada")
        if enrollment.estudiante_id != current_user.id:
            raise HTTPException(status_code=403, detail="No tienes permiso")
            
    if isinstance(current_user, User):
        # AUDITORÍA (CRÍTICO #1): mismo guard general, además de la restricción específica de CPD.
        if current_user.rol not in ["superadmin", "admin", "mae", "cpd", "cobranza"]:
            raise HTTPException(status_code=403, detail="No autorizado para ver el resumen de pagos")

        if current_user.rol == "cpd":
            raise HTTPException(
                status_code=403,
                detail="El rol CPD tiene estrictamente prohibido visualizar flujos de caja y estados financieros"
            )

        # ISSUE-P-SEGMENTACION: Cobranza con cursos_asignados solo ve inscripciones de esos cursos.
        filtro_rol = filtro_cursos_por_rol(current_user)
        if filtro_rol:
            from services import enrollment_service
            target_enrollment = await enrollment_service.get_enrollment(enrollment_id)
            if not target_enrollment or target_enrollment.curso_id not in filtro_rol["curso_id"]["$in"]:
                raise HTTPException(status_code=403, detail="No tienes asignado el curso de esta inscripción")
    
    resumen = await payment_service.get_resumen_pagos_enrollment(enrollment_id)
    return resumen

