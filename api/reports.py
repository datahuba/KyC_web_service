"""
API de Reportes Financieros
===========================

F-CUENTAS-POR-COBRAR (2026-07-29): reporte de CxC real vs estimada para
informes financieros del staff. Devuelve totales globales + desglose por
curso y por enrollment. Soporta export XLSX para auditoría.

Endpoints:
- GET /reports/cuentas-por-cobrar              → JSON con desglose completo
- GET /reports/cuentas-por-cobrar/xlsx        → XLSX descargable
- GET /reports/cuentas-por-cobrar/resumen     → totales solamente (para dashboard)

RBAC: todos requieren require_staff (cualquier rol administrativo).
Para usuarios segmentados (encargado_curso con cursos_asignados), se
filtra automáticamente por los cursos asignados.
"""

from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from api.dependencies import require_staff
from models.user import User
from services import cuentas_por_cobrar_service, cuentas_historicas_service
from services.cuentas_por_cobrar_service import CxCResumen
from pydantic import BaseModel, Field

router = APIRouter()


# ========================================================================
# SCHEMAS DE SALIDA (Pydantic, F-CUENTAS-POR-COBRAR)
# ========================================================================

class ModuloCxCOut(BaseModel):
    nombre: str
    modulo_index: int
    costo: float
    monto_pagado: float
    saldo_pendiente: float
    iniciado_en: Optional[datetime] = None
    cuenta_cxc_real: bool


class EnrollmentCxCOut(BaseModel):
    enrollment_id: str
    estudiante_id: str
    estudiante_nombre: str
    estudiante_registro: Optional[str] = None
    curso_id: str
    curso_nombre: str
    estado: str
    total_a_pagar: float
    total_pagado: float
    recaudacion_efectiva: float = 0.0
    total_devengado: float = 0.0
    cxc_devengada: float = 0.0
    proyeccion_futura: float = 0.0
    saldo_estimado: float
    saldo_a_la_fecha: float
    modulos: List[ModuloCxCOut]


class CursoCxCOut(BaseModel):
    curso_id: str
    curso_nombre: str
    curso_codigo: Optional[str] = None
    cantidad_estudiantes: int
    recaudacion_efectiva: float = 0.0
    total_devengado: float = 0.0
    cxc_devengada: float = 0.0
    proyeccion_futura: float = 0.0
    total_estimado: float
    total_a_la_fecha: float


class CxCResumenOut(BaseModel):
    recaudacion_efectiva: float = 0.0
    total_devengado: float = 0.0
    cxc_devengada: float = 0.0
    proyeccion_futura: float = 0.0
    total_estimado: float
    total_a_la_fecha: float
    total_modulos_iniciados: int
    total_modulos_no_iniciados: int
    cantidad_enrollments: int
    cantidad_cursos: int
    por_curso: List[CursoCxCOut]
    detalle: List[EnrollmentCxCOut]
    generado_en: datetime


class CxCResumenReducidoOut(BaseModel):
    """Solo totales, para tarjetas del dashboard."""
    recaudacion_efectiva: float = 0.0
    total_devengado: float = 0.0
    cxc_devengada: float = 0.0
    proyeccion_futura: float = 0.0
    total_estimado: float
    total_a_la_fecha: float
    diferencia: float = Field(..., description="estimado - cxc_devengada")
    total_modulos_iniciados: int
    total_modulos_no_iniciados: int
    cantidad_enrollments: int
    cantidad_cursos: int
    generado_en: datetime


def _to_out(resumen: CxCResumen) -> CxCResumenOut:
    """Convierte CxCResumen (servicio) a CxCResumenOut (Pydantic)."""
    return CxCResumenOut(
        recaudacion_efectiva=resumen.recaudacion_efectiva,
        total_devengado=resumen.total_devengado,
        cxc_devengada=resumen.cxc_devengada,
        proyeccion_futura=resumen.proyeccion_futura,
        total_estimado=resumen.total_estimado,
        total_a_la_fecha=resumen.total_a_la_fecha,
        total_modulos_iniciados=resumen.total_modulos_iniciados,
        total_modulos_no_iniciados=resumen.total_modulos_no_iniciados,
        cantidad_enrollments=resumen.cantidad_enrollments,
        cantidad_cursos=resumen.cantidad_cursos,
        por_curso=[
            CursoCxCOut(**acc.__dict__) if hasattr(acc, "__dict__") else acc
            for acc in resumen.por_curso
        ],
        detalle=[
            EnrollmentCxCOut(
                enrollment_id=d.enrollment_id,
                estudiante_id=d.estudiante_id,
                estudiante_nombre=d.estudiante_nombre,
                estudiante_registro=d.estudiante_registro,
                curso_id=d.curso_id,
                curso_nombre=d.curso_nombre,
                estado=d.estado,
                total_a_pagar=d.total_a_pagar,
                total_pagado=d.total_pagado,
                recaudacion_efectiva=d.recaudacion_efectiva,
                total_devengado=d.total_devengado,
                cxc_devengada=d.cxc_devengada,
                proyeccion_futura=d.proyeccion_futura,
                saldo_estimado=d.saldo_estimado,
                saldo_a_la_fecha=d.saldo_a_la_fecha,
                modulos=[
                    ModuloCxCOut(
                        nombre=m.nombre,
                        modulo_index=m.modulo_index,
                        costo=m.costo,
                        monto_pagado=m.monto_pagado,
                        saldo_pendiente=m.saldo_pendiente,
                        iniciado_en=m.iniciado_en,
                        cuenta_cxc_real=m.cuenta_cxc_real,
                    ) for m in d.modulos
                ],
            ) for d in resumen.detalle
        ],
        generado_en=resumen.generado_en,
    )


# ========================================================================
# ENDPOINTS
# ========================================================================

@router.get(
    "/cuentas-por-cobrar",
    response_model=CxCResumenOut,
    summary="[Staff] Reporte de CxC real vs estimada (con desglose)",
)
async def get_cuentas_por_cobrar(
    curso_id: Optional[str] = Query(None, description="Filtrar por curso específico"),
    current_user: User = Depends(require_staff),
) -> CxCResumenOut:
    resumen = await cuentas_por_cobrar_service.generar_resumen_cxc(
        current_user=current_user,
        curso_id=curso_id,
    )
    return _to_out(resumen)


@router.get(
    "/cuentas-por-cobrar/resumen",
    response_model=CxCResumenReducidoOut,
    summary="[Staff] Resumen rápido (solo totales, para dashboard)",
)
async def get_cuentas_por_cobrar_resumen(
    current_user: User = Depends(require_staff),
) -> CxCResumenReducidoOut:
    resumen = await cuentas_por_cobrar_service.generar_resumen_cxc(
        current_user=current_user,
    )
    return CxCResumenReducidoOut(
        recaudacion_efectiva=resumen.recaudacion_efectiva,
        total_devengado=resumen.total_devengado,
        cxc_devengada=resumen.cxc_devengada,
        proyeccion_futura=resumen.proyeccion_futura,
        total_estimado=resumen.total_estimado,
        total_a_la_fecha=resumen.total_a_la_fecha,
        diferencia=round(resumen.total_estimado - resumen.cxc_devengada, 2),
        total_modulos_iniciados=resumen.total_modulos_iniciados,
        total_modulos_no_iniciados=resumen.total_modulos_no_iniciados,
        cantidad_enrollments=resumen.cantidad_enrollments,
        cantidad_cursos=resumen.cantidad_cursos,
        generado_en=resumen.generado_en,
    )


@router.get(
    "/cuentas-por-cobrar/xlsx",
    summary="[Staff] Exportar CxC a XLSX",
    response_class=StreamingResponse,
)
async def export_cuentas_por_cobrar_xlsx(
    curso_id: Optional[str] = Query(None),
    current_user: User = Depends(require_staff),
) -> StreamingResponse:
    """
    F-CUENTAS-POR-COBRAR: exporta el reporte de CxC a un XLSX con dos hojas:
    - 'Resumen por curso' (totales por programa)
    - 'Detalle por enrollment' (línea por estudiante/módulo)

    Filtros: igual que GET /cuentas-por-cobrar.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    resumen = await cuentas_por_cobrar_service.generar_resumen_cxc(
        current_user=current_user,
        curso_id=curso_id,
    )

    wb = Workbook()
    # Hoja 1: Resumen por curso
    ws1 = wb.active
    ws1.title = "Resumen por curso"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8A1F2F", end_color="8A1F2F", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    headers_1 = ["Curso", "Código", "Estudiantes", "CxC Estimada (Bs)", "CxC a la Fecha (Bs)"]
    for col, h in enumerate(headers_1, start=1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for r, acc in enumerate(resumen.por_curso, start=2):
        ws1.cell(row=r, column=1, value=acc.curso_nombre)
        ws1.cell(row=r, column=2, value=acc.curso_codigo or "")
        ws1.cell(row=r, column=3, value=acc.cantidad_estudiantes)
        ws1.cell(row=r, column=4, value=float(acc.total_estimado))
        ws1.cell(row=r, column=5, value=float(acc.total_a_la_fecha))

    # Fila total
    r = len(resumen.por_curso) + 2
    ws1.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=r, column=3, value=resumen.cantidad_enrollments).font = Font(bold=True)
    ws1.cell(row=r, column=4, value=float(resumen.total_estimado)).font = Font(bold=True)
    ws1.cell(row=r, column=5, value=float(resumen.total_a_la_fecha)).font = Font(bold=True)

    # Anchos de columna
    ws1.column_dimensions['A'].width = 50
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 22

    # Formato moneda para CxC
    for row in ws1.iter_rows(min_row=2, min_col=4, max_col=5, max_row=r):
        for c in row:
            c.number_format = '#,##0.00'

    # Hoja 2: Detalle por enrollment
    ws2 = wb.create_sheet("Detalle por enrollment")
    headers_2 = [
        "Estudiante", "Registro", "Curso", "Estado", "Módulo",
        "Costo (Bs)", "Pagado (Bs)", "Saldo (Bs)", "Iniciado en", "CxC Real?"
    ]
    for col, h in enumerate(headers_2, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    r = 2
    for d in resumen.detalle:
        for m in d.modulos:
            ws2.cell(row=r, column=1, value=d.estudiante_nombre)
            ws2.cell(row=r, column=2, value=d.estudiante_registro or "")
            ws2.cell(row=r, column=3, value=d.curso_nombre)
            ws2.cell(row=r, column=4, value=d.estado)
            ws2.cell(row=r, column=5, value=m.nombre)
            ws2.cell(row=r, column=6, value=float(m.costo))
            ws2.cell(row=r, column=7, value=float(m.monto_pagado))
            ws2.cell(row=r, column=8, value=float(m.saldo_pendiente))
            ws2.cell(row=r, column=9, value=m.iniciado_en)
            ws2.cell(row=r, column=10, value="Sí" if m.cuenta_cxc_real else "No")
            r += 1

    # Anchos columna
    widths = [30, 12, 40, 15, 30, 12, 12, 12, 20, 10]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # Formato moneda
    for row in ws2.iter_rows(min_row=2, min_col=6, max_col=8, max_row=r - 1):
        for c in row:
            c.number_format = '#,##0.00'

    # Formato fecha
    for row in ws2.iter_rows(min_row=2, min_col=9, max_col=9, max_row=r - 1):
        for c in row:
            c.number_format = 'dd/mm/yyyy HH:mm'

    # Generar bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"cuentas_por_cobrar_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(buf.getbuffer().nbytes),
        },
    )


# ============================================================================
# CUENTAS HISTORICAS (F-CUENTAS-HISTORICAS, 2026-08-16)
# ============================================================================
# Los programas historicos salieron del Dashboard y de Cuentas por Cobrar
# (ver F-HISTORICOS-FUERA-DEL-DASHBOARD). Estos endpoints son su contraparte:
# el expediente economico de esos programas, aparte de la cartera corriente.


class HistEnrollmentOut(BaseModel):
    enrollment_id: str
    estudiante_id: str
    estudiante_nombre: str
    estudiante_registro: Optional[str] = None
    curso_id: str
    curso_nombre: str
    estado: str
    total_a_pagar: float
    total_pagado: float
    saldo: float


class HistCursoOut(BaseModel):
    curso_id: str
    curso_nombre: str
    curso_codigo: Optional[str] = None
    fecha_inicio: Optional[datetime] = None
    fecha_fin: Optional[datetime] = None
    cantidad_estudiantes: int
    total_esperado: float
    total_cobrado: float
    saldo_pendiente: float
    avance_pct: float


class HistResumenOut(BaseModel):
    total_programas: int
    total_estudiantes: int
    total_esperado: float
    total_cobrado: float
    saldo_pendiente: float
    avance_pct: float
    por_curso: List[HistCursoOut]
    detalle: List[HistEnrollmentOut]
    generado_en: datetime


def _hist_to_out(r) -> HistResumenOut:
    return HistResumenOut(
        total_programas=r.total_programas,
        total_estudiantes=r.total_estudiantes,
        total_esperado=r.total_esperado,
        total_cobrado=r.total_cobrado,
        saldo_pendiente=r.saldo_pendiente,
        avance_pct=r.avance_pct,
        por_curso=[HistCursoOut(**vars(c)) for c in r.por_curso],
        detalle=[HistEnrollmentOut(**vars(d)) for d in r.detalle],
        generado_en=r.generado_en,
    )


@router.get(
    "/cuentas-historicas",
    response_model=HistResumenOut,
    summary="[Staff] Resumen economico de los programas historicos",
)
async def get_cuentas_historicas(
    curso_id: Optional[str] = Query(None, description="Filtrar por un programa historico"),
    current_user: User = Depends(require_staff),
) -> HistResumenOut:
    """
    Expediente economico de los programas marcados `es_historico=True`.

    Son programas de carga retroactiva que ya NO cuentan como cartera
    corriente (quedan fuera del Dashboard y de Cuentas por Cobrar), pero
    cuyos datos hay que conservar y poder consultar.

    A diferencia de CxC, aca SI se cuentan las inscripciones COMPLETADO: en
    un historico ese es el estado esperado, y excluirlas dejaria el informe
    practicamente vacio. Se siguen excluyendo CANCELADO y RETIRADO.
    """
    resumen = await cuentas_historicas_service.generar_resumen_historico(
        current_user=current_user,
        curso_id=curso_id,
    )
    return _hist_to_out(resumen)


@router.get(
    "/cuentas-historicas/xlsx",
    summary="[Staff] Exportar Cuentas Historicas a XLSX",
    response_class=StreamingResponse,
)
async def export_cuentas_historicas_xlsx(
    curso_id: Optional[str] = Query(None),
    current_user: User = Depends(require_staff),
) -> StreamingResponse:
    """Mismo reporte en Excel, con dos hojas: resumen por programa y detalle."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    r = await cuentas_historicas_service.generar_resumen_historico(
        current_user=current_user, curso_id=curso_id
    )

    wb = Workbook()
    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="8A1F2F")  # guindo UAGRM
    centro = Alignment(horizontal="center")

    def _encabezar(ws, columnas):
        ws.append(columnas)
        for celda in ws[1]:
            celda.font = encabezado
            celda.fill = fondo
            celda.alignment = centro

    ws1 = wb.active
    ws1.title = "Resumen por programa"
    _encabezar(ws1, [
        "Codigo", "Programa", "Estudiantes",
        "Total esperado (Bs)", "Cobrado (Bs)", "Saldo (Bs)", "Avance %",
    ])
    for c in r.por_curso:
        ws1.append([
            c.curso_codigo or "", c.curso_nombre, c.cantidad_estudiantes,
            c.total_esperado, c.total_cobrado, c.saldo_pendiente, c.avance_pct,
        ])
    ws1.append([])
    ws1.append([
        "TOTAL", "", r.total_estudiantes,
        r.total_esperado, r.total_cobrado, r.saldo_pendiente, r.avance_pct,
    ])
    for celda in ws1[ws1.max_row]:
        celda.font = Font(bold=True)
    for col, ancho in zip("ABCDEFG", (18, 52, 12, 20, 16, 16, 10)):
        ws1.column_dimensions[col].width = ancho

    ws2 = wb.create_sheet("Detalle por estudiante")
    _encabezar(ws2, [
        "Programa", "Estudiante", "Registro", "Estado",
        "Total a pagar (Bs)", "Pagado (Bs)", "Saldo (Bs)",
    ])
    for d in r.detalle:
        ws2.append([
            d.curso_nombre, d.estudiante_nombre, d.estudiante_registro or "",
            d.estado, d.total_a_pagar, d.total_pagado, d.saldo,
        ])
    for col, ancho in zip("ABCDEFG", (52, 34, 14, 14, 18, 16, 16)):
        ws2.column_dimensions[col].width = ancho

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "cuentas-historicas-%s.xlsx" % datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(buf.getbuffer().nbytes),
        },
    )
