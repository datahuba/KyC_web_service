"""
Servicio de Estudiantes
=======================

Lógica de negocio para estudiantes (Funciones).
"""

import csv
import re
import unicodedata
import openpyxl
from io import BytesIO, StringIO
from typing import List, Optional, Union
from models.student import Student
from models.enums import EstadoTitulo
from schemas.student import StudentCreate, StudentUpdateSelf, StudentUpdateAdmin
from beanie import PydanticObjectId
from beanie.operators import Or, RegEx, In


def _escape_regex(s: str) -> str:
    """
    Escapa caracteres especiales de regex MongoDB para que un string
    de usuario sea tratado como literal (no como patron).

    Sin escape, un usuario que busca 'A.B+' recibira matches de
    cualquier 'A' + cualquier caracter + 'B+', lo cual no es lo que
    esperan. Con escape, busca literalmente 'A.B+'.
    """
    return s.replace("\\", "\\\\").replace(".", "\\.").replace("*", "\\*") \
            .replace("+", "\\+").replace("?", "\\?").replace("(", "\\(") \
            .replace(")", "\\)").replace("[", "\\[").replace("]", "\\]") \
            .replace("{", "\\{").replace("}", "\\}").replace("|", "\\|") \
            .replace("^", "\\^").replace("$", "\\$")


async def get_students(
    page: int = 1,
    per_page: int = 10,
    q: Optional[str] = None,
    activo: Optional[bool] = None,
    estado_titulo: Optional[EstadoTitulo] = None,
    curso_id: Optional[PydanticObjectId] = None,
    # F-2026-08-12-EC-CURSOS-FILTRO (Kevin 2026-08-12 post-reunion UAGRM):
    # si el usuario es ENCARGADO_CURSO / COORDINADOR / COBRANZA-segmentado,
    # el endpoint le pasa su lista de cursos_asignados y el service filtra
    # a estudiantes que estan en al menos uno de esos cursos. Si la lista
    # está vacía, retorna lista vacía (no se muestran todos los estudiantes).
    cursos_asignados: Optional[list] = None,
) -> tuple[List[Student], int]:
    """
    Obtener lista de estudiantes con filtros avanzados y paginación
    """
    query = Student.find()

    if q:
        # F-FIX-FILTROS-STUDENTS (2026-08-10, Kevin): antes el filtro `q`
        # estaba mal armado. Comparaba el campo con un dict literal
        # `{"$regex": q, "$options": "i"}`, lo cual MongoDB no interpretaria
        # como regex sino como un valor escalar (un dict), y siempre
        # retornaba 0 matches. Resultado: ?search=X, ?carnet=X, etc
        # siempre devolvian los primeros N estudiantes (pagina 1) sin
        # aplicar el filtro. Ahora se usa el operador RegEx() de Beanie,
        # que construye correctamente la query Mongo `{campo: {$regex: ..., $options: 'i'}}`.
        # Ademas, para carnet y registro se hace match exacto (case-insensitive
        # pero exacto) ya que son campos unicos. Para nombre y email se usa
        # regex parcial.
        q_escaped = _escape_regex(q)
        # Para carnet/registro: match exacto (case-insensitive)
        query = query.find(
            Or(
                RegEx(Student.nombre, q_escaped, "i"),
                RegEx(Student.email, q_escaped, "i"),
                RegEx(Student.carnet, "^" + q_escaped + "$", "i"),
                RegEx(Student.registro, "^" + q_escaped + "$", "i")
            )
        )

    if activo is not None:
        query = query.find(Student.activo == activo)

    if estado_titulo:
        if estado_titulo == EstadoTitulo.SIN_TITULO:
            query = query.find(
                Or(
                    Student.titulo.estado == EstadoTitulo.SIN_TITULO,
                    Student.titulo == None
                )
            )
        else:
            query = query.find(Student.titulo.estado == estado_titulo)

    # F-2026-08-12-EC-CURSOS-FILTRO: si llega curso_id especifico Y
    # cursos_asignados, intersectamos (AND logico). El EC puede pedir un
    # curso especifico dentro de los que le pertenecen, pero NO un curso
    # fuera de su lista.
    if curso_id is not None and cursos_asignados is not None:
        if curso_id not in cursos_asignados:
            # El EC pidio un curso que no le pertenece → lista vacia
            return [], 0
        cursos_asignados = [curso_id]
    elif curso_id is not None:
        cursos_asignados = [curso_id]

    if cursos_asignados is not None:
        # Filtrar por intersección: estudiantes que tienen al menos uno
        # de los cursos_asignados en su lista_cursos_ids. Si cursos_asignados
        # es lista vacia, retorna 0 matches (no estudiantes sin cursos).
        if not cursos_asignados:
            return [], 0
        query = query.find({"lista_cursos_ids": {"$in": cursos_asignados}})

    total_count = await query.count()
    skip = (page - 1) * per_page

    students = await query.sort("-created_at").skip(skip).limit(per_page).to_list()

    return students, total_count


async def get_student(id: PydanticObjectId) -> Optional[Student]:
    """Obtener estudiante por ID"""
    return await Student.get(id)


async def accept_terms(student: Student) -> Student:
    """
    ISSUE-Q-PRE: Registra la aceptación del reglamento de Posgrado.

    Idempotente: si ya había aceptado antes, no pisa la fecha original
    de la primera aceptación (se conserva como evidencia histórica).
    """
    from core.timezone_utils import utcnow_naive

    if not student.terminos_aceptados:
        student.terminos_aceptados = True
        student.fecha_aceptacion_terminos = utcnow_naive()
        await student.save()

    return student


async def create_student(student_in: StudentCreate) -> Student:
    """
    Crear nuevo estudiante
    
    Si se provee password, se usa; sino, se hashea el carnet (fallback).
    Si se provee course_id, se inscribe automáticamente (y se validan los datos primero).
    """
    from core.security import get_password_hash
    from models.course import Course
    from schemas.enrollment import EnrollmentCreate
    from services import enrollment_service
    
    # 1. Validaciones robustas de Unicidad (Registro, Carnet, Correo) en base de datos
    check_conditions = []
    if student_in.registro:
        check_conditions.append(Student.registro == student_in.registro)
    if student_in.carnet:
        check_conditions.append(Student.carnet == student_in.carnet)
    if student_in.email and student_in.email.strip():
        check_conditions.append(Student.email == student_in.email.strip().lower())
        
    if check_conditions:
        existing = await Student.find_one(Or(*check_conditions))
        if existing:
            if student_in.registro and existing.registro == student_in.registro:
                raise ValueError(f"Ya existe un estudiante registrado con el Registro Académico: '{student_in.registro}'.")
            if student_in.carnet and existing.carnet == student_in.carnet:
                raise ValueError(f"Ya existe un estudiante registrado con el Carnet de Identidad (C.I.): '{student_in.carnet}'.")
            if student_in.email and existing.email and existing.email.lower() == student_in.email.strip().lower():
                raise ValueError(f"Ya existe un estudiante registrado con el Correo Electrónico: '{student_in.email}'.")

    student_data = student_in.model_dump(exclude_unset=True)
    
    # Extraer campos opcionales sin romper el resto de la lógica
    course_id = student_data.pop("course_id", None)
    password_input = student_data.pop("password", None)
    
    # Normalizar correo electrónico si se ha proporcionado
    if "email" in student_data and student_data["email"]:
        student_data["email"] = student_data["email"].strip().lower()
    
    # 2. Validar existencia del curso ANTES de crear al estudiante (Ahorro de BD)
    course_obj = None
    if course_id:
        course_obj = await Course.get(course_id)
        if not course_obj:
            raise ValueError("Curso no encontrado")
        # F-HISTORICO-AUTOSERVICIO-EXCEL-FIX (2026-08-04): los cursos historicos
        # aceptan carga de estudiantes aunque activo=False (cursos cerrados para
        # carga retroactiva). Mismo criterio que el endpoint de auto-enroll
        # (linea 835). Si no es historico y esta inactivo, rechazar.
        if not course_obj.activo and not course_obj.es_historico:
            raise ValueError("El curso seleccionado está inactivo")

    # 3. Lógica Inteligente de Contraseña
    # ISSUE-Q-PASSWORD-UNIFICADA (2026-07-08): la contraseña inicial de
    # estudiantes ahora usa la misma convención institucional 'Uagrm.<CI>'
    # que ya se usaba para docentes/staff (GAP-1), unificando el criterio en
    # toda la plataforma. Antes era el carnet crudo sin prefijo.
    if password_input:
        student_data["password"] = get_password_hash(password_input)
    else:
        student_data["password"] = get_password_hash(f"Uagrm.{student_data['carnet']}")
        
    # 4. Persistir Estudiante
    student = Student(**student_data)
    await student.insert()
    
    # 5. Puente de Inscripción Integrado
    if course_obj:
        try:
            await enrollment_service.create_enrollment(
                enrollment_in=EnrollmentCreate(
                    estudiante_id=student.id,
                    curso_id=course_obj.id,
                    descuento_id=None,
                    descuento_personalizado=None
                ),
                admin_username="system_student_create"
            )
        except Exception as e:
            # Rollback compensatorio si la inscripción falla por error interno
            await student.delete()
            raise ValueError(f"Error en la auto-inscripción: {str(e)}")
            
    return student


async def update_student(
    student: Student,
    student_in: Union[StudentUpdateSelf, StudentUpdateAdmin]
) -> Student:
    """Actualizar estudiante existente"""
    from core.security import get_password_hash
    
    update_data = student_in.model_dump(exclude_unset=True)
    
    # 1. Validaciones robustas de Unicidad en Modificación (Excluyendo al propio estudiante)
    check_conditions = []
    new_registro = update_data.get("registro")
    new_carnet = update_data.get("carnet")
    new_email = update_data.get("email")
    
    if new_registro and new_registro != student.registro:
        check_conditions.append(Student.registro == new_registro)
    if new_carnet and new_carnet != student.carnet:
        check_conditions.append(Student.carnet == new_carnet)
    if new_email and new_email.strip() and (not student.email or new_email.strip().lower() != student.email.lower()):
        check_conditions.append(Student.email == new_email.strip().lower())
        
    if check_conditions:
        existing = await Student.find_one(
            Or(*check_conditions),
            Student.id != student.id
        )
        if existing:
            if new_registro and existing.registro == new_registro:
                raise ValueError(f"El Registro Académico '{new_registro}' ya está siendo usado por otro estudiante.")
            if new_carnet and existing.carnet == new_carnet:
                raise ValueError(f"El Carnet de Identidad '{new_carnet}' ya está registrado en otra cuenta.")
            if new_email and existing.email and existing.email.lower() == new_email.strip().lower():
                raise ValueError(f"El Correo Electrónico '{new_email}' ya está registrado en otra cuenta.")

    if "password" in update_data and update_data["password"]:
        update_data["password"] = get_password_hash(update_data["password"])
        
    if "email" in update_data and update_data["email"]:
        update_data["email"] = update_data["email"].strip().lower()
        # ISSUE-A-VERIFICACION: si el correo realmente cambió, la verificación
        # anterior (si existía) ya no aplica al nuevo correo.
        if update_data["email"] != (student.email or "").strip().lower():
            update_data["email_verificado"] = False
            update_data["fecha_verificacion_email"] = None
    
    for field, value in update_data.items():
        setattr(student, field, value)
    
    await student.save()
    return student


async def delete_student(id: PydanticObjectId) -> Student:
    """Eliminar estudiante"""
    student = await Student.get(id)
    if student:
        await student.delete()
    return student


# ============================================================================
# LOGICA DE IMPORTACIÓN MASIVA DESDE EXCEL OPTIMIZADA DE ALTA VELOCIDAD (Bulk Write)
# ============================================================================

def _normalize_header(value) -> str:
    """Normaliza un encabezado: minúsculas, sin acentos y sin espacios sobrantes."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s.strip()


def _clean_text(value) -> Optional[str]:
    """
    Convierte una celda a texto limpio. Quita el sufijo '.0' de números que Excel/Forms
    leyó como float (ej. carnet 2969698.0 -> '2969698'). Devuelve None si queda vacío.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    if not s:
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s if s else None


def _split_carnet(value) -> tuple[Optional[str], Optional[str]]:
    """
    Separa un carnet de identidad en (numero_limpio, complemento).

    Ej: '2726683 - 1J' -> ('2726683', '1J'), '1313665-1D' -> ('1313665', '1D'),
    '2969698' -> ('2969698', None).

    El complemento (ej. '1D'/'1J'/'1O') es un dato oficial DISTINTO de
    `extension` (que es el lugar de expedición del carnet, ej. 'SC'/'LPZ') --
    se guarda aparte en `Student.complemento_carnet` en vez de descartarse,
    y el número limpio (sin complemento) es el que se usa para las
    validaciones de unicidad/duplicados entre archivos y con la BD.
    """
    cleaned = _clean_text(value)
    if not cleaned:
        return None, None
    partes = re.split(r"\s*-\s*", cleaned, maxsplit=1)
    numero = partes[0].strip() if partes and partes[0].strip() else None
    complemento = partes[1].strip() if len(partes) > 1 and partes[1].strip() else None
    return numero, complemento


def _clean_carnet(value) -> Optional[str]:
    """Compatibilidad: devuelve solo el número limpio del carnet (sin complemento)."""
    numero, _ = _split_carnet(value)
    return numero


def _parse_fecha_nacimiento(value):
    """
    Parsea la fecha de nacimiento de una celda del Excel, detectando
    automáticamente si el formato es DÍA/MES/AÑO o MES/DÍA/AÑO (la etiqueta
    de la cabecera del archivo NO es confiable: puede decir "mm/dd/aaaa" y
    en realidad venir en día/mes/año, o viceversa, según quién armó la
    plantilla).

    Heurística de detección (sobre los dos primeros números separados por
    '/' o '-'):
    - Si el PRIMER número es > 12, no puede ser un mes -> es DÍA/MES/AÑO.
    - Si el SEGUNDO número es > 12, no puede ser un mes -> es MES/DÍA/AÑO
      (el primer número es el mes).
    - Si ambos son <= 12 (ambiguo, ej. '4/5/1990'), se asume DÍA/MES/AÑO por
      ser la convención predominante en Bolivia/Latinoamérica.

    Si la celda ya es un datetime real (Excel la guardó como fecha nativa),
    se usa tal cual sin reinterpretar -- solo se aplica esta heurística a
    celdas de texto.

    Devuelve un datetime o None si no se pudo parsear.
    """
    from datetime import datetime as _dt
    from core.timezone_utils import utcnow_naive

    if value is None:
        return None
    if isinstance(value, _dt):
        return value
    s = str(value).strip()
    if not s:
        return None

    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$", s)
    if not m:
        return None

    primero, segundo, anio_str = m.groups()
    primero_n, segundo_n = int(primero), int(segundo)
    anio = int(anio_str)
    if anio < 100:
        # Años de 2 dígitos: asumir 1900s si > 30 (heurística estándar), sino 2000s.
        anio += 1900 if anio > 30 else 2000

    if primero_n > 12:
        dia, mes = primero_n, segundo_n
    elif segundo_n > 12:
        mes, dia = primero_n, segundo_n
    else:
        # Ambiguo: convención día/mes/año (Bolivia/Latinoamérica) por defecto.
        dia, mes = primero_n, segundo_n

    try:
        return _dt(anio, mes, dia)
    except ValueError:
        return None


def _detectar_leyenda_colores(sheet) -> dict:
    """
    Detecta una leyenda de colores al final de la hoja (ISSUE-Q-LEYENDA-COLORES,
    2026-07-08): algunos archivos reales marcan filas con un color de fondo
    para indicar una condición especial (ej. "Descuento Facultad", "Pendiente
    de pago"), y más abajo en la misma hoja incluyen una leyenda: una celda
    con ESE MISMO color de fondo junto a una celda de texto que explica qué
    significa. Esto es importante porque indica a qué estudiantes aplica un
    descuento y a cuáles no -- información que de otro modo se perdería
    silenciosamente al importar (el importador no lee colores, solo valores).

    Retorna un dict {color_rgb: texto_significado}. Vacío si no se encuentra
    ninguna leyenda reconocible.
    """
    leyenda = {}
    try:
        for row in range(1, sheet.max_row + 1):
            for col in range(1, min(sheet.max_column, 6) + 1):
                cell = sheet.cell(row=row, column=col)
                # Patrón real de una muestra de leyenda: la celda coloreada
                # está VACÍA (el color es solo la "muestra"/swatch); si la
                # celda coloreada tiene contenido, es una fila de datos
                # normal marcada con ese color, no la leyenda.
                if cell.value is not None:
                    continue
                fill = cell.fill
                if not fill or fill.patternType != "solid":
                    continue
                try:
                    color = fill.fgColor.rgb
                except Exception:
                    continue
                if not isinstance(color, str) or color in ("00000000", "FFFFFFFF", None):
                    continue
                if color in leyenda:
                    continue
                # Buscar el texto explicativo en otra celda de la misma fila
                # que NO tenga relleno de color (para no confundir con otra
                # celda coloreada que sea dato, no leyenda).
                for otro_col in range(1, sheet.max_column + 1):
                    if otro_col == col:
                        continue
                    otra_cell = sheet.cell(row=row, column=otro_col)
                    otro_fill = otra_cell.fill
                    tiene_color = bool(
                        otro_fill and otro_fill.patternType == "solid"
                        and isinstance(getattr(otro_fill.fgColor, "rgb", None), str)
                        and otro_fill.fgColor.rgb not in ("00000000", "FFFFFFFF")
                    )
                    if tiene_color:
                        continue
                    texto_val = otra_cell.value
                    if texto_val and isinstance(texto_val, str) and texto_val.strip():
                        leyenda[color] = texto_val.strip()
                        break
    except Exception:
        # La detección de leyenda es un extra informativo, nunca debe romper
        # la importación real de estudiantes si falla por algún motivo.
        return {}
    return leyenda


def _parse_amount(value) -> float:
    """Convierte una celda monetaria a float. Devuelve 0.0 si no es un número válido."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "")
    if not s:
        return 0.0
    s = s.replace(",", "")  # separador de miles
    try:
        return float(s)
    except ValueError:
        return 0.0


async def import_students_from_excel(
    file_content: bytes,
    curso_id: Optional[PydanticObjectId] = None,
    filename: Optional[str] = None
) -> dict:
    """
    Importar estudiantes de forma masiva desde un archivo de Excel (.xlsx).

    AUTO-INSCRIPCIÓN OPCIONAL:
    Si se proporciona `curso_id`, todos los estudiantes recién creados en esta importación
    serán inscritos automáticamente a ese curso/diplomado reutilizando la lógica financiera
    oficial de `enrollment_service.create_enrollment` (cálculo de precios, módulos y descuentos).
    Los errores de inscripción se reportan por estudiante y no interrumpen el resto del lote.
    """
    from core.security import get_password_hash
    
    is_csv = bool(filename and filename.lower().strip().endswith('.csv'))
    try:
        if is_csv:
            # CSV: decodificar respetando el BOM (utf-8-sig) con fallback a latin-1,
            # autodetectar el delimitador (coma o punto y coma) y volcar las filas
            # en una hoja de openpyxl para reutilizar la misma lógica de Excel.
            try:
                text = file_content.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = file_content.decode('latin-1')
            sample = text[:2048]
            delimiter = ';' if sample.count(';') > sample.count(',') else ','
            reader = csv.reader(StringIO(text), delimiter=delimiter)
            wb = openpyxl.Workbook()
            sheet = wb.active
            for row in reader:
                sheet.append(row)
        else:
            # Cargar libro en memoria de forma optimizada
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            sheet = wb.active
        if not sheet or sheet.max_row < 1:
            raise ValueError("El archivo no contiene datos.")
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"No se pudo parsear el archivo: {str(e)}")
        
    # 1. ESCANEO DINÁMICO DE CABECERAS (FILA 1) — normalizado sin acentos + detección financiera
    header_row = [_normalize_header(sheet.cell(row=1, column=col_idx).value)
                  for col_idx in range(1, sheet.max_column + 1)]

    col_nombre = col_apellido = col_registro = col_carnet = 0
    col_extension = col_email = col_celular = col_domicilio = 0
    col_matricula = 0
    col_fecha_nacimiento = 0
    col_tipo_sangre = 0
    col_matricula_comprobante = 0  # columna con el LINK del comprobante de pago de matrícula (Google Forms)
    columnas_modulos = []  # lista de (col_idx, concepto) para pagos de módulos/cuotas

    for idx, header in enumerate(header_row, start=1):
        if not header:
            continue
        # Detectar la columna con el comprobante (link) de pago de matrícula ANTES de excluir adjuntos.
        # Ej. Google Forms: "Adjuntar Comprobante de Pago de la Matrícula (IMAGEN O PDF) Bs. 300".
        if "comprobante" in header and "matricula" in header and col_matricula_comprobante == 0:
            col_matricula_comprobante = idx
            continue
        # Ignorar columnas calculadas y de adjuntos/enlaces (no son datos ni montos individuales).
        # Esto evita que columnas de Google Forms como "Adjuntar ... Carnet de Identidad" (una URL)
        # se confundan con el carnet.
        if any(tok in header for tok in ("total", "saldo", "cobrar", "adjuntar", "comprobante")) or header.startswith("http"):
            continue
        # --- Campos del estudiante (gana la PRIMERA coincidencia; nombre antes que apellido) ---
        if "nombre" in header and col_nombre == 0:
            col_nombre = idx
        elif "apellido" in header and col_apellido == 0:
            col_apellido = idx
        elif ("registro" in header or "matricula academica" in header) and col_registro == 0:
            col_registro = idx
        elif (header == "ci" or "carnet" in header or "documento de identidad" in header) and col_carnet == 0:
            col_carnet = idx
        elif "extension" in header and col_extension == 0:
            col_extension = idx
        elif ("correo" in header or "email" in header or "mail" in header) and col_email == 0:
            col_email = idx
        elif ("celular" in header or "telefono" in header or "telf" in header) and col_celular == 0:
            col_celular = idx
        elif ("domicilio" in header or "direccion" in header) and col_domicilio == 0:
            col_domicilio = idx
        elif ("nacimiento" in header or "fecha de nacimiento" in header) and col_fecha_nacimiento == 0:
            col_fecha_nacimiento = idx
        elif ("sangre" in header or "sanguineo" in header or "sanguinea" in header) and col_tipo_sangre == 0:
            col_tipo_sangre = idx
        # --- Columnas financieras (para migrar pagos si se selecciona un curso) ---
        elif "matricula" in header and col_matricula == 0:
            col_matricula = idx
        else:
            m = re.match(r"^m\s*(\d+)$", header) or re.match(r"^(?:modulo|cuota)\s*(\d+)$", header)
            if m:
                columnas_modulos.append((idx, f"Módulo {m.group(1)}"))

    if col_nombre == 0:
        raise ValueError("No se encontró la columna de 'Nombre' en la fila de cabecera del archivo.")
    if col_carnet == 0:
        raise ValueError("No se encontró la columna de 'CI' o 'Carnet' en la fila de cabecera del archivo.")

    # ISSUE-Q-LEYENDA-COLORES: detectar si el archivo trae una leyenda de
    # colores (ej. "amarillo = Descuento Facultad") para reportar qué filas
    # coinciden con cada color -- informativo, NO crea descuentos automáticamente.
    leyenda_colores = _detectar_leyenda_colores(sheet)
    marcados_por_color: dict = {}  # {significado_leyenda: [nombres de estudiantes]}

    errors = []
    candidates = []
    
    registros_en_archivo = set()
    carnets_en_archivo = set()
    emails_en_archivo = set()
    empty_row_streak = 0
    
    # 2. ESCANEAR FILAS Y VALIDAR EN MEMORIA (FILTRANDO VACÍOS Y DUPLICADOS INTERNOS)
    for row_idx in range(2, sheet.max_row + 1):
        try:
            nombres_val = sheet.cell(row=row_idx, column=col_nombre).value if col_nombre > 0 else None
            apellidos_val = sheet.cell(row=row_idx, column=col_apellido).value if col_apellido > 0 else None

            nombres_str = str(nombres_val).strip() if nombres_val is not None else ""
            apellidos_str = str(apellidos_val).strip() if apellidos_val is not None else ""
            # Combinar Nombre(s) + Apellido(s) cuando el archivo los trae separados (ej. Google Forms)
            nombre_str = f"{nombres_str} {apellidos_str}".strip() if (col_apellido > 0 and apellidos_str) else nombres_str

            # Carnet limpio: sin sufijo .0 de floats y con el complemento tras
            # el guion separado aparte (ej. '2726683 - 1J' -> carnet='2726683',
            # complemento_carnet='1J'), para que el mismo CI se detecte como
            # duplicado sin importar si el archivo lo trae con o sin
            # complemento, sin perder el dato oficial completo.
            carnet, complemento_carnet = _split_carnet(
                sheet.cell(row=row_idx, column=col_carnet).value if col_carnet > 0 else None
            )
            carnet_str = carnet if carnet else ""

            if not nombre_str and not carnet_str:
                empty_row_streak += 1
                if empty_row_streak >= 5:
                    break 
                continue
            else:
                empty_row_streak = 0
                
            email = sheet.cell(row=row_idx, column=col_email).value if col_email > 0 else None
            domicilio_val = sheet.cell(row=row_idx, column=col_domicilio).value if col_domicilio > 0 else None

            nombre = nombre_str if nombre_str else None
            registro = _clean_text(sheet.cell(row=row_idx, column=col_registro).value if col_registro > 0 else None)
            extension = _clean_text(sheet.cell(row=row_idx, column=col_extension).value if col_extension > 0 else None)
            email = str(email).strip().lower() if email is not None else None
            celular = _clean_text(sheet.cell(row=row_idx, column=col_celular).value if col_celular > 0 else None)
            domicilio = str(domicilio_val).strip() if domicilio_val is not None else None
            if domicilio == "":
                domicilio = None

            fecha_nacimiento = _parse_fecha_nacimiento(
                sheet.cell(row=row_idx, column=col_fecha_nacimiento).value if col_fecha_nacimiento > 0 else None
            )

            tipo_sangre_raw = _clean_text(
                sheet.cell(row=row_idx, column=col_tipo_sangre).value if col_tipo_sangre > 0 else None
            )
            tipo_sangre = None
            if tipo_sangre_raw:
                candidato = tipo_sangre_raw.strip().upper().replace(" ", "")
                # Normaliza variantes comunes ('0+' con cero en vez de letra O, etc.)
                candidato = candidato.replace("0", "O")
                if candidato in ("A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"):
                    tipo_sangre = candidato

            # Recolectar pagos de la fila desde las columnas financieras detectadas
            pagos_fila = []
            if col_matricula > 0:
                monto_mat = _parse_amount(sheet.cell(row=row_idx, column=col_matricula).value)
                if monto_mat > 0:
                    pagos_fila.append(("Matrícula", round(monto_mat, 2)))
            for (cidx, concepto_mod) in columnas_modulos:
                monto_mod = _parse_amount(sheet.cell(row=row_idx, column=cidx).value)
                if monto_mod > 0:
                    pagos_fila.append((concepto_mod, round(monto_mod, 2)))

            # Capturar el LINK del comprobante de matrícula (si existe y es una URL)
            matricula_comprobante_url = None
            if col_matricula_comprobante > 0:
                comp_val = sheet.cell(row=row_idx, column=col_matricula_comprobante).value
                comp_str = str(comp_val).strip() if comp_val is not None else ""
                if comp_str and comp_str.lower().startswith("http"):
                    matricula_comprobante_url = comp_str
                    
            # ISSUE-Q-LEYENDA-COLORES: si esta fila (columna Nombre o Apellido)
            # tiene el color de fondo de alguna entrada de la leyenda, registrar
            # a qué significado corresponde (ej. "Descuento Facultad").
            if leyenda_colores and nombre_str:
                for col_check in (col_nombre, col_apellido):
                    if col_check <= 0:
                        continue
                    fill_check = sheet.cell(row=row_idx, column=col_check).fill
                    color_check = None
                    if fill_check and fill_check.patternType == "solid":
                        try:
                            color_check = fill_check.fgColor.rgb
                        except Exception:
                            color_check = None
                    if color_check and color_check in leyenda_colores:
                        significado = leyenda_colores[color_check]
                        marcados_por_color.setdefault(significado, []).append(nombre_str)
                        break

            if not nombre:
                errors.append(f"Fila {row_idx}: El nombre completo es obligatorio.")
                continue
                
            if not carnet:
                errors.append(f"Fila {row_idx}: El carnet (CI) de '{nombre}' es obligatorio.")
                continue
                
            if not registro:
                # Si el archivo no trae columna de registro académico, se usa
                # el carnet (ya limpio de complemento) como usuario -- así
                # el usuario y la contraseña inicial son ambos el CI, mismo
                # formato usado en toda la plataforma para estudiantes.
                registro = carnet
                
            # Controlar duplicados en el mismo archivo para no meter llaves repetidas a BD
            if registro in registros_en_archivo:
                errors.append(f"Fila {row_idx}: El Registro Académico o correo '{registro}' de '{nombre}' está duplicado dentro de este archivo Excel.")
                continue
            
            if carnet in carnets_en_archivo:
                errors.append(f"Fila {row_idx}: El Carnet de Identidad (CI) '{carnet}' de '{nombre}' está duplicado dentro de este archivo Excel.")
                continue
                
            if email:
                if email in emails_en_archivo:
                    errors.append(f"Fila {row_idx}: El Correo Electrónico '{email}' de '{nombre}' está duplicado dentro de este archivo Excel.")
                    continue
                # ISSUE-EXCEL-EMAIL-VALID (2026-08-03, Kevin): pre-validar el
                # formato del email en el Excel ANTES de intentar crear el
                # Student (que tiene EmailStr de Pydantic y fallaba al
                # insertar el primero inválido, sin listar los demás).
                # Detectamos: espacios, formato mal escrito, etc.
                if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
                    errors.append(
                        f"Fila {row_idx}: El Correo Electrónico '{email}' de "
                        f"'{nombre}' no tiene un formato válido. "
                        f"Verifica que no tenga espacios ni caracteres raros."
                    )
                    continue
                emails_en_archivo.add(email)
                
            registros_en_archivo.add(registro)
            carnets_en_archivo.add(carnet)
            
            candidates.append({
                "row_idx": row_idx,
                "registro": registro,
                "nombre": nombre,
                "email": email,
                "carnet": carnet,
                "complemento_carnet": complemento_carnet,
                "extension": extension,
                "celular": celular,
                "domicilio": domicilio,
                "fecha_nacimiento": fecha_nacimiento,
                "tipo_sangre": tipo_sangre,
                "pagos": pagos_fila, # Montos a migrar (si el archivo trae columnas financieras)
                "matricula_comprobante_url": matricula_comprobante_url # Link del voucher de matrícula
            })
        except Exception as e:
            errors.append(f"Fila {row_idx}: Error al procesar datos de la fila: {str(e)}")
            
    # 3. VERIFICAR DUPLICADOS EN BASE DE DATOS (1 SOLA CONSULTA DE RED)
    # F-EXCEL-IMPORT-EXISTING (2026-08-03, Kevin): si el estudiante YA EXISTE
    # en la BD, NO lo rechazamos — lo inscribimos al curso seleccionado (si lo
    # hay). El usuario puede subir un Excel con estudiantes que ya están
    # registrados para inscribirlos en masa a un nuevo programa.
    existing_registros = set()
    existing_carnets = set()
    existing_emails = set()
    existing_students_by_carnet: dict = {}  # carnet -> Student object (para inscribir)
    existing_students_by_registro: dict = {}

    if candidates:
        all_registros_excel = [c["registro"] for c in candidates]
        all_carnets_excel = [c["carnet"] for c in candidates]
        all_emails_excel = [c["email"] for c in candidates if c["email"]]

        db_query = {
            "$or": [
                {"registro": {"$in": all_registros_excel}},
                {"carnet": {"$in": all_carnets_excel}}
            ]
        }
        if all_emails_excel:
            db_query["$or"].append({"email": {"$in": all_emails_excel}})

        existing_students_db = await Student.find(db_query).to_list()

        for s in existing_students_db:
            if s.registro:
                existing_registros.add(s.registro)
                existing_students_by_registro[s.registro] = s
            if s.carnet:
                existing_carnets.add(s.carnet)
                existing_students_by_carnet[s.carnet] = s
            if s.email:
                existing_emails.add(s.email.lower())

    # 4. PREPARAR OBJETOS DE INSERTIÓN Y HASHEAR CONTRASEÑAS (PROCESADOR CPU CONTINUO)
    # Y también identificar estudiantes existentes para inscribir al curso.
    students_to_insert = []
    financials_to_insert = []  # pagos por estudiante, alineado 1:1 con students_to_insert
    existing_to_enroll: list = []  # tuplas (c, student_obj) para inscribir al curso
    for c in candidates:
        # Buscar si el estudiante ya existe
        existing_student = None
        if c["carnet"] in existing_carnets:
            existing_student = existing_students_by_carnet.get(c["carnet"])
        elif c["registro"] in existing_registros:
            existing_student = existing_students_by_registro.get(c["registro"])

        if existing_student:
            # Estudiante YA EXISTE en BD — lo inscribimos al curso
            # (NO es un error, es flujo normal de re-inscripción).
            existing_to_enroll.append((c, existing_student))
            continue

        # ISSUE-Q-PASSWORD-UNIFICADA: contraseña inicial 'Uagrm.<CI>' (misma
        # convención institucional que docentes/staff, GAP-1).
        hashed_password = get_password_hash(f"Uagrm.{c['carnet']}")

        students_to_insert.append(
            Student(
                registro=c["registro"],
                password=hashed_password,
                nombre=c["nombre"],
                email=c["email"],
                carnet=c["carnet"],
                complemento_carnet=c["complemento_carnet"],
                extension=c["extension"],
                celular=c["celular"],
                domicilio=c["domicilio"],
                fecha_nacimiento=c["fecha_nacimiento"],
                tipo_sangre=c["tipo_sangre"],
                activo=True,
                lista_cursos_ids=[]
            )
        )
        financials_to_insert.append({
            "pagos": c["pagos"],
            "matricula_comprobante_url": c["matricula_comprobante_url"]
        })
        
    # 5. INSERCIÓN MASIVA DE ALTO RENDIMIENTO (1 SOLA ESCRITURA DE RED)
    success_count = 0
    inserted_ids: List[PydanticObjectId] = []
    # ISSUE-EXCEL-EMAIL-VALID (2026-08-03, Kevin): si hay errores de validación
    # (emails mal escritos, duplicados, formato inválido), SE IMPORTAN LAS
    # FILAS VALIDAS y se reportan las filas malas en `errors[]`. El usuario
    # puede corregir el Excel y re-subir las filas malas por separado.
    # Esto evita que 1 fila con email mal escrito impida importar el resto.
    if students_to_insert:
        insert_result = await Student.insert_many(students_to_insert)
        # inserted_ids conserva el mismo orden que students_to_insert
        inserted_ids = list(insert_result.inserted_ids)
        success_count = len(inserted_ids)

    # 6. AUTO-INSCRIPCIÓN OPCIONAL A UN CURSO + MIGRACIÓN DE PAGOS
    enrolled_count = 0
    migrated_payments_count = 0  # estudiantes a los que se les migró al menos un pago aprobado
    matricula_vouchers_count = 0  # comprobantes de matrícula (link) registrados como pago pendiente
    hay_columnas_financieras = bool(col_matricula or columnas_modulos or col_matricula_comprobante)

    # F-EXCEL-IMPORT-EXISTING (2026-08-03, Kevin): inscribir TANTO los estudiantes
    # nuevos COMO los que ya existen en BD (y no están inscritos en este curso).
    # Unificamos la lista de "estudiantes a inscribir" antes del loop paralelo.
    enrollable_inputs: list = []  # tuplas (student_id, student_obj, fin_dict)
    for i, sid in enumerate(inserted_ids):
        enrollable_inputs.append((sid, students_to_insert[i], financials_to_insert[i]))
    for c, existing_student in existing_to_enroll:
        # Para los existentes, también migramos los pagos del Excel (histórico)
        enrollable_inputs.append((existing_student.id, existing_student, {
            "pagos": c["pagos"],
            "matricula_comprobante_url": c["matricula_comprobante_url"]
        }))

    if curso_id and enrollable_inputs:
        import asyncio
        from core.timezone_utils import utcnow_naive
        from models.course import Course
        from models.payment import Payment
        from models.enums import EstadoPago
        from schemas.enrollment import EnrollmentCreate
        from services import enrollment_service

        course = await Course.get(curso_id)
        if not course:
            errors.append(
                f"Auto-inscripción cancelada: el curso seleccionado ({curso_id}) no existe. "
                f"Los {len(inserted_ids)} estudiantes se crearon correctamente pero no fueron inscritos."
            )
        elif not course.activo and not course.es_historico:
            # F-HISTORICO (2026-08-03, Kevin): los programas historicos aceptan
            # inscripciones aunque activo=False (cursos cerrados para carga retroactiva).
            errors.append(
                f"Auto-inscripción cancelada: el curso '{course.nombre_programa}' está inactivo y no acepta inscripciones. "
                f"Los {len(inserted_ids)} estudiantes se crearon correctamente pero no fueron inscritos."
            )
        else:
            # ============================================================
            # OPTIMIZACIÓN DE RENDIMIENTO (2026-07-09, ISSUE-Q-IMPORT-TIMEOUT)
            # ============================================================
            # DIAGNÓSTICO: cada estudiante en el bucle anterior (secuencial)
            # ejecutaba ~7-8 round-trips de red a MongoDB Atlas
            # (Student.get, Course.get, Enrollment.find_one, enrollment.insert,
            # course.save, student.save, Payment.insert_many,
            # Payment.find+enrollment.save de actualizar_saldo_enrollment),
            # uno detrás de otro. Con 53 estudiantes reales se midió ~2s por
            # estudiante (~106s totales), superando el timeout de 120s del
            # cliente aunque el servidor terminara la importación con éxito.
            #
            # FIX: los estudiantes recién insertados en este lote son
            # independientes entre sí (no hay una inscripción previa que
            # pueda chocar, ya se validó unicidad de carnet/registro/email
            # arriba), así que se procesan en paralelo controlado
            # (semáforo) en vez de secuencialmente. `course`/`student` ya
            # están en memoria (no hace falta volver a pedirlos por
            # estudiante) y las actualizaciones de `course.inscritos`
            # (que NO tiene optimistic locking) se acumulan en un set y se
            # guardan en un solo `course.save()` al final, evitando la
            # condición de carrera de mutar/guardar el mismo `Course`
            # desde tareas concurrentes.
            CONCURRENCIA_MAXIMA = 8
            semaphore = asyncio.Semaphore(CONCURRENCIA_MAXIMA)
            resultados_lock = asyncio.Lock()

            async def _inscribir_estudiante_importado(student_id, student_obj, fin: dict) -> None:
                nonlocal enrolled_count, migrated_payments_count, matricula_vouchers_count
                pagos = fin["pagos"]
                matricula_url = fin["matricula_comprobante_url"]
                async with semaphore:
                    try:
                        enrollment = await enrollment_service.create_enrollment(
                            enrollment_in=EnrollmentCreate(
                                estudiante_id=student_id,
                                curso_id=curso_id
                            ),
                            admin_username="import_masivo",
                            student=student_obj,
                            course=course,
                            skip_link_updates=True  # batcheado al final, ver abajo
                        )

                        payment_objs = []

                        # a) Montos numéricos = dinero histórico confirmado -> pagos APROBADOS
                        for (concepto, monto) in pagos:
                            payment_objs.append(Payment(
                                inscripcion_id=enrollment.id,
                                estudiante_id=student_id,
                                curso_id=curso_id,
                                metodo_pago="Migración",
                                concepto=concepto,
                                cantidad_pago=monto,
                                numero_cuota=None,
                                numero_transaccion=None,
                                comprobante_url=None,
                                remitente=None,
                                banco="Migración Excel",
                                monto_comprobante=monto,
                                fecha_comprobante=None,
                                cuenta_destino="Importación Masiva",
                                estado_pago=EstadoPago.APROBADO,
                                fecha_verificacion=utcnow_naive(),
                                verificado_por="import_masivo",
                            ))

                        # b) Link de comprobante de matrícula = voucher subido sin conciliar -> pago PENDIENTE
                        #    (solo si no vino ya un monto numérico de matrícula y el curso tiene matrícula > 0)
                        tiene_matricula_numerica = any(concepto == "Matrícula" for (concepto, _) in pagos)
                        registro_voucher = (
                            matricula_url
                            and not tiene_matricula_numerica
                            and enrollment.costo_matricula > 0
                        )
                        if registro_voucher:
                            payment_objs.append(Payment(
                                inscripcion_id=enrollment.id,
                                estudiante_id=student_id,
                                curso_id=curso_id,
                                metodo_pago="Transferencia",
                                concepto="Matrícula",
                                cantidad_pago=enrollment.costo_matricula,
                                numero_cuota=None,
                                numero_transaccion=None,
                                comprobante_url=matricula_url,
                                remitente=None,
                                banco=None,
                                monto_comprobante=enrollment.costo_matricula,
                                fecha_comprobante=None,
                                cuenta_destino=None,
                                estado_pago=EstadoPago.PENDIENTE,
                            ))

                        if payment_objs:
                            await Payment.insert_many(payment_objs)
                            # El motor oficial reconstruye matrícula, módulos, totales y estado con los APROBADOS.
                            # `enrollment` ya está en memoria, pasado para evitar el Enrollment.get() redundante.
                            await enrollment_service.actualizar_saldo_enrollment(
                                enrollment_id=enrollment.id,
                                monto_pago_aprobado=0.0,
                                enrollment=enrollment
                            )

                        async with resultados_lock:
                            enrolled_count += 1
                            course.inscritos.append(student_id)
                            if curso_id not in student_obj.lista_cursos_ids:
                                student_obj.lista_cursos_ids.append(curso_id)
                            if pagos:
                                migrated_payments_count += 1
                            if registro_voucher:
                                matricula_vouchers_count += 1
                    except ValueError as e:
                        async with resultados_lock:
                            errors.append(
                                f"Inscripción de '{student_obj.nombre}' (Registro {student_obj.registro}) fallida: {str(e)}"
                            )
                    except Exception as e:
                        async with resultados_lock:
                            errors.append(
                                f"Inscripción de '{student_obj.nombre}' (Registro {student_obj.registro}) fallida por error inesperado: {str(e)}"
                            )

            await asyncio.gather(*[
                _inscribir_estudiante_importado(student_id, student_obj, fin)
                for student_id, student_obj, fin in enrollable_inputs
            ])

            # Persistir en batch las referencias cruzadas acumuladas (1 sola
            # escritura de red para el Course, en vez de una por estudiante).
            # `student_obj.lista_cursos_ids` de cada estudiante recién
            # procesado (NUEVO o EXISTENTE) se persiste en batch.
            if course.inscritos:
                await course.save()

            # Actualizar lista_cursos_ids SOLO de los estudiantes NUEVOS insertados
            # (los existentes ya tienen su lista en la BD; modificarla indiscriminadamente
            # borraría cursos previamente inscritos).
            estudiantes_nuevos_con_curso = [
                student_id for student_id, student_obj in zip(inserted_ids, students_to_insert)
                if curso_id in student_obj.lista_cursos_ids
            ]
            if estudiantes_nuevos_con_curso:
                # update_many agregando el curso sin duplicar (no tocamos existentes
                # porque podrían tener otros cursos inscritos).
                from beanie.operators import AddToSet
                await Student.find(In(Student.id, estudiantes_nuevos_con_curso)).update(
                    AddToSet({"lista_cursos_ids": curso_id})
                )
    elif hay_columnas_financieras and not curso_id and success_count > 0:
        errors.append(
            "Se detectaron columnas de pago (Matrícula/Módulos) en el archivo, pero no se seleccionó un curso; "
            "los pagos NO fueron registrados. Para migrar los pagos, la importación debe hacerse seleccionando el curso."
        )

    return {
        "success_count": success_count,
        "enrolled_count": enrolled_count,
        "migrated_payments_count": migrated_payments_count,
        "matricula_vouchers_count": matricula_vouchers_count,
        "errors": errors,
        # ISSUE-Q-LEYENDA-COLORES: informativo -- el CPD debe revisar estos
        # grupos manualmente y decidir si crea/asigna el descuento correspondiente
        # (el importador NUNCA crea ni asigna descuentos automáticamente).
        "marcados_por_color": marcados_por_color
    }


async def get_student_financial_summary(student_id: PydanticObjectId) -> dict:
    """
    Obtener resumen financiero unificado de todas las inscripciones del estudiante (Ficha de Estado de Cuenta).
    Retorna: total_invertido, pagado, en_proceso, saldo_pendiente
    """
    from models.enrollment import Enrollment
    from models.payment import Payment
    from models.enums import EstadoPago

    # 1. Obtener todas las inscripciones del estudiante
    enrollments = await Enrollment.find(Enrollment.estudiante_id == student_id).to_list()

    total_invertido = sum(e.total_a_pagar for e in enrollments)
    pagado = sum(e.total_pagado for e in enrollments)
    saldo_pendiente = sum(e.saldo_pendiente for e in enrollments)

    # 2. Obtener la suma de pagos en estado PENDIENTE
    pending_payments = await Payment.find(
        Payment.estudiante_id == student_id,
        Payment.estado_pago == EstadoPago.PENDIENTE
    ).to_list()

    en_proceso = sum(p.cantidad_pago for p in pending_payments)

    return {
        "total_invertido": round(total_invertido, 2),
        "pagado": round(pagado, 2),
        "en_proceso": round(en_proceso, 2),
        "saldo_pendiente": round(saldo_pendiente, 2)
    }


