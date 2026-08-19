"""
Servicio de Inscripciones (Enrollments)
=======================================

Lógica de negocio para inscripciones.

Permisos:
---------
- CREAR inscripción: Solo ADMIN/SUPERADMIN
- ACTUALIZAR inscripción: Solo ADMIN/SUPERADMIN
- VER inscripciones: ADMIN (todas), STUDENT (solo las suyas)
"""

from typing import List, Optional, Any
from datetime import datetime


def _recalcular_estado_modulo(mod) -> None:
    """
    F-FIX-ESTADO-MODULOS-POST-DESCUENTO (2026-08-08, Kevin): cuando se cambia
    el costo de un modulo (mod.costo), hay que recalcular su estado
    financiero. Antes quedaba en "Parcial" aunque monto_pagado cubriera
    el nuevo costo con descuento (caso de becados).

    Reglas:
    - monto_pagado == 0 -> Pendiente
    - 0 < monto_pagado < costo - 0.01 -> Parcial
    - monto_pagado >= costo - 0.01 -> Pagado
    """
    costo = float(mod.costo or 0.0)
    pagado = float(mod.monto_pagado or 0.0)
    if pagado <= 0.005:
        mod.estado = "Pendiente"
    elif abs(pagado - costo) < 0.01 or pagado >= costo - 0.01:
        mod.estado = "Pagado"
    else:
        mod.estado = "Parcial"
from models.enrollment import Enrollment, ModuloEstado, CargoAdicionalItemSnapshot
from models.student import Student
from models.course import Course
from models.enums import EstadoInscripcion
from schemas.enrollment import EnrollmentCreate
from beanie import PydanticObjectId
from models.discount import Discount
from beanie.operators import In, Or
# F-046 FIX: helpers de tiempo estaban solo importados localmente en
# enrich_enrollment_dates; al refactorizar core/timezone_utils se rompió
# el uso en 9 funciones (subir_nota_borrador, actualizar_saldo_enrollment,
# cambiar_estado_inscripcion, eximir_matricula, rechazar_nota_borrador).
# El síntoma fue 500 al calificar módulo (Sandra, audio 22/7 19:12).
from core.timezone_utils import utcnow_naive

async def create_enrollment(
    enrollment_in: EnrollmentCreate,
    admin_username: str,
    student: Optional[Student] = None,
    course: Optional[Course] = None,
    skip_link_updates: bool = False
) -> Enrollment:
    """
    Crear una nueva inscripción (solo admins)

    OPTIMIZACIÓN DE IMPORTACIÓN MASIVA (2026-07-09, ISSUE-Q-IMPORT-TIMEOUT):
    `student`/`course` permiten pasar objetos ya obtenidos en memoria para
    evitar los round-trips `Student.get()`/`Course.get()` cuando el llamador
    ya los tiene (ej. import_students_from_excel, donde el mismo `course` se
    reutiliza para decenas de estudiantes en una sola importación).

    `skip_link_updates=True` omite los `course.save()`/`student.save()` que
    agregan la referencia cruzada (course.inscritos / student.lista_cursos_ids)
    -- estos dos documentos NO tienen optimistic locking (`use_revision`), por
    lo que mutarlos y guardarlos de forma concurrente (ej. en un
    `asyncio.gather` sobre varios estudiantes) puede perder escrituras
    (last-writer-wins). El llamador que use este flag es responsable de
    actualizar esas referencias por su cuenta, idealmente en un solo batch
    después de que todas las inscripciones individuales ya se crearon.
    """
    # 1. Obtener estudiante y curso (si no se proveyeron ya en memoria)
    if student is None:
        student = await Student.get(enrollment_in.estudiante_id)
        if not student:
            raise ValueError(f"Estudiante {enrollment_in.estudiante_id} no encontrado")

    if course is None:
        course = await Course.get(enrollment_in.curso_id)
        if not course:
            raise ValueError(f"Curso {enrollment_in.curso_id} no encontrado")

    # AUDITORÍA (MEDIO #7): create_enrollment_request (solicitud del propio
    # estudiante) sí valida curso.activo, pero esta vía directa (CPD/Encargado
    # de Curso) no lo hacía -- dos rutas de entrada con validación asimétrica,
    # permitiendo inscribir gente en cursos ya desactivados/cerrados.
    # F-HISTORICO (2026-08-03, Kevin): los programas historicos (es_historico=True)
    # SIEMPRE deben aceptar inscripciones, porque su proposito es cargar
    # retroactivamente estudiantes que cursaron en el pasado. El flag activo
    # puede estar en False (cerrado) pero es_historico=True significa que
    # es solo para carga historica, no operativo.
    if not course.activo and not course.es_historico:
        raise ValueError("Este curso no está activo y no acepta nuevas inscripciones")

    # 2. Validar que no esté ya inscrito
    existing = await Enrollment.find_one(
        Enrollment.estudiante_id == enrollment_in.estudiante_id,
        Enrollment.curso_id == enrollment_in.curso_id,
        Enrollment.estado != EstadoInscripcion.CANCELADO
    )
    if existing:
        raise ValueError(
            f"El estudiante ya está inscrito en este curso (Inscripción ID: {existing.id})"
        )
    
    # 3. ISSUE-P-PRECIO-UNICO (2026-07-08): el precio del programa es el
    # mismo para todos los estudiantes, sin importar procedencia/tipo.
    
    # 4. Obtener precios del curso (precio único)
    costo_total = course.get_costo_total()
    # F-2026-08-12-DESCUENTO-BECA (Kevin 2026-08-12, reunion UAGRM):
    # la matricula depende del tipo de estudiante (primer carrera vs
    # profesional con titulo). El override del curso o el default global
    # (200/500) se calcula en services/matricula_helper.py.
    from services.matricula_helper import get_matricula_for_student
    costo_matricula = get_matricula_for_student(course, student)

    # F-HISTORICO-IMPORT (2026-08-03, Kevin): los programas historicos pueden
    # tener cantidad_cuotas=0 (no se exige en historicos). Enrollment requiere
    # >= 1, asi que forzamos a 1 cuando es historico y no tiene cuotas.
    cantidad_cuotas_efectiva = course.cantidad_cuotas
    if course.es_historico and cantidad_cuotas_efectiva < 1:
        cantidad_cuotas_efectiva = 1
    
    # 5. Aplicar descuento del curso
    descuento_curso = 0.0
    descuento_curso_id = None

    if course.descuento_id:
        discount_obj = await Discount.get(course.descuento_id)
        if discount_obj and discount_obj.activo:
            descuento_curso = discount_obj.porcentaje
            descuento_curso_id = discount_obj.id
    elif course.descuento_curso:
        descuento_curso = course.descuento_curso

    # Total con SOLO descuento del curso (referencia para distribución de
    # módulos sin beca personal)
    total_con_descuento_curso = costo_total - (costo_total * descuento_curso / 100)

    # 6. Aplicar descuento del estudiante
    descuento_personal = 0.0
    descuento_estudiante_id = None

    if enrollment_in.descuento_id:
        discount_sel = await Discount.get(enrollment_in.descuento_id)
        if discount_sel and discount_sel.activo:
            descuento_personal = discount_sel.porcentaje
            descuento_estudiante_id = discount_sel.id
    elif enrollment_in.descuento_personalizado:
        descuento_personal = enrollment_in.descuento_personalizado

    # F-LOGICA-DESCUENTOS-MAX (2026-08-05, Kevin): "se queda con el descuento
    # de mayor porcentaje". Si el personal es menor, gana el curso y se
    # descarta el personal (el endpoint avisa al usuario).
    descuento_efectivo = max(descuento_curso, descuento_personal)

    colegiatura_final = costo_total - (costo_total * descuento_efectivo / 100)

    # ISSUE-P-CARGO-MULTIITEM: suma de todos los ítems de cargo adicional/
    # complementario al programa (ej. varios talleres incluidos). NINGÚN
    # ítem recibe descuentos de curso/estudiante -- se cobran íntegros.
    cargo_adicional = course.get_cargo_adicional_total()
    
    # MATEMÁTICA FINANCIERA:
    total_final = colegiatura_final + costo_matricula + cargo_adicional

    # ISSUE-P-RECALCULO-NOTA: snapshot de la nota mínima exigida por el descuento personal
    # (solo si viene de un Discount vinculado con condición académica; descuento_personalizado libre no aplica)
    nota_minima_snapshot = None
    if descuento_estudiante_id and discount_sel and discount_sel.nota_minima_requerida is not None:
        nota_minima_snapshot = discount_sel.nota_minima_requerida
    
    # 7. Copiar requisitos del curso
    requisitos_enrollment = [template.to_requisito() for template in course.requisitos]
    
    # 8. Clonación y distribución de módulos A PRUEBA DE BALAS
    modulos_enrollment = []
    if course.modulos:
        suma_costo_modulos = sum(mod.costo for mod in course.modulos)
        total_asignado = 0.0
        total_asignado_sin_beca = 0.0  # ISSUE-P-RECALCULO-NOTA
        
        for i, mod in enumerate(course.modulos):
            es_ultimo = i == len(course.modulos) - 1
            if es_ultimo:
                costo_final_mod = max(0.0, round(colegiatura_final - total_asignado, 2))
            else:
                if suma_costo_modulos > 0:
                    costo_final_mod = round((mod.costo / suma_costo_modulos) * colegiatura_final, 2)
                else:
                    costo_final_mod = round(colegiatura_final / len(course.modulos), 2)
                total_asignado += costo_final_mod

            # ISSUE-P-RECALCULO-NOTA: distribución paralela SIN el descuento personal,
            # usando la misma proporción, solo si el descuento tiene condición académica.
            costo_sin_beca_mod = None
            if nota_minima_snapshot is not None:
                if es_ultimo:
                    costo_sin_beca_mod = max(0.0, round(total_con_descuento_curso - total_asignado_sin_beca, 2))
                else:
                    if suma_costo_modulos > 0:
                        costo_sin_beca_mod = round((mod.costo / suma_costo_modulos) * total_con_descuento_curso, 2)
                    else:
                        costo_sin_beca_mod = round(total_con_descuento_curso / len(course.modulos), 2)
                    total_asignado_sin_beca += costo_sin_beca_mod
            
            modulos_enrollment.append(
                ModuloEstado(
                    nombre=mod.nombre,
                    costo=costo_final_mod,
                    costo_sin_beca_personal=costo_sin_beca_mod,
                    estado="Pendiente",
                    monto_pagado=0.0,
                    nota=None,
                    estado_academico="Cursando"
                )
            )
    
    # 9. Crear inscripción
    enrollment = Enrollment(
        estudiante_id=enrollment_in.estudiante_id,
        curso_id=enrollment_in.curso_id,
        costo_total=costo_total,
        costo_matricula=costo_matricula,
        cantidad_cuotas=cantidad_cuotas_efectiva,
        modulos=modulos_enrollment,
        
        descuento_curso_id=descuento_curso_id,
        descuento_curso_aplicado=descuento_curso,
        descuento_estudiante_id=descuento_estudiante_id,
        descuento_personalizado=descuento_personal,

        # ISSUE-P-CARGO-MULTIITEM: snapshot de la lista de ítems de cargo
        # adicional al momento de inscribirse (si el curso los tenía
        # definidos en ese momento).
        cargo_adicional_items=[
            CargoAdicionalItemSnapshot(nombre=item.nombre, costo=item.costo)
            for item in course.cargo_adicional_items
        ],
        
        total_a_pagar=round(total_final, 2),
        saldo_pendiente=round(total_final, 2),
        estado=EstadoInscripcion.PENDIENTE_PAGO,
        matricula_pagada=False,
        requisitos=requisitos_enrollment,
        nota_minima_beca=nota_minima_snapshot  # ISSUE-P-RECALCULO-NOTA
    )
    
    await enrollment.insert()
    
    # 10. Agregar estudiante a la lista / 11. Agregar curso al estudiante
    # (omitido si skip_link_updates=True -- ver docstring de esta función)
    if not skip_link_updates:
        if enrollment_in.estudiante_id not in course.inscritos:
            course.inscritos.append(enrollment_in.estudiante_id)
            await course.save()

        if enrollment_in.curso_id not in student.lista_cursos_ids:
            student.lista_cursos_ids.append(enrollment_in.curso_id)
            await student.save()
    
    return enrollment


async def enrich_enrollment_dates(enrollment: Enrollment) -> dict:
    """
    F-FIX-DESCONOCIDO-ENROLLMENTS (2026-08-09, Kevin): ahora joinea el nombre
    del estudiante y del curso ademas de las fechas. El bug era que el
    frontend mostraba "Desconocido" para IDs de estudiantes fuera de los
    primeros 100. Ahora el endpoint /enrollments/ devuelve los nombres
    directamente, sin necesidad de joinear en el cliente.

    Para listas grandes, usar `enrich_enrollments_batch()` que hace lookup
    con In() (1 query por coleccion) en vez de N queries individuales.
    """
    from core.timezone_utils import utcnow_naive, to_bolivia_time
    enrollment_dict = enrollment.model_dump()
    enrollment_dict["fecha_inscripcion"] = to_bolivia_time(enrollment.fecha_inscripcion)
    enrollment_dict["created_at"] = to_bolivia_time(enrollment.created_at)
    enrollment_dict["updated_at"] = to_bolivia_time(enrollment.updated_at)

    # F-LOGICA-DESCUENTOS-MAX (2026-08-05, Kevin): exponer al frontend el
    # descuento que REALMENTE se aplicó y de dónde viene, para que pueda
    # mostrar el mensaje "se aplicó el mayor (X%)" cuando corresponda.
    desc_curso = float(enrollment.descuento_curso_aplicado or 0)
    desc_personal = float(enrollment.descuento_personalizado or 0) if enrollment.descuento_personalizado is not None else 0.0
    descuento_efectivo = max(desc_curso, desc_personal)
    if descuento_efectivo > 0 and desc_personal > 0 and desc_curso >= desc_personal:
        # El personal fue menor, se aplicó el curso (se descarta el personal)
        origen = "curso"
        advertencia = (
            f"El descuento personal seleccionado ({desc_personal:.0f}%) es menor al "
            f"descuento del curso ({desc_curso:.0f}%). Se aplicó el descuento de mayor "
            f"porcentaje: el del curso ({desc_curso:.0f}%)."
        )
    elif descuento_efectivo > 0 and desc_personal > desc_curso:
        origen = "personal"
        advertencia = None
    elif descuento_efectivo > 0 and desc_curso > 0:
        origen = "curso"
        advertencia = None
    else:
        origen = "ninguno"
        advertencia = None
    enrollment_dict["descuento_efectivo"] = descuento_efectivo
    enrollment_dict["descuento_efectivo_origen"] = origen
    enrollment_dict["advertencia_descuento"] = advertencia

    # F-FIX-DESCONOCIDO-ENROLLMENTS: joinear nombre del estudiante
    # y del curso para que el frontend NO muestre "Desconocido".
    if enrollment.estudiante_id:
        try:
            from core.cache import get_students_bulk_cached
            students_map = await get_students_bulk_cached([enrollment.estudiante_id])
            student = students_map.get(enrollment.estudiante_id)
            if student:
                # student puede ser dict (de motor) o Beanie Student
                nombre = student.get("nombre") if isinstance(student, dict) else getattr(student, "nombre", None)
                registro = student.get("registro") if isinstance(student, dict) else getattr(student, "registro", None)
                ci = student.get("carnet") if isinstance(student, dict) else getattr(student, "carnet", None)
                enrollment_dict["estudiante_nombre"] = nombre
                enrollment_dict["estudiante_registro"] = registro
                enrollment_dict["estudiante_ci"] = ci
            else:
                enrollment_dict["estudiante_nombre"] = None
                enrollment_dict["estudiante_registro"] = None
                enrollment_dict["estudiante_ci"] = None
        except Exception:
            enrollment_dict["estudiante_nombre"] = None
            enrollment_dict["estudiante_registro"] = None
            enrollment_dict["estudiante_ci"] = None
    else:
        enrollment_dict["estudiante_nombre"] = None
        enrollment_dict["estudiante_registro"] = None
        enrollment_dict["estudiante_ci"] = None

    # Joinear nombre del curso
    if enrollment.curso_id:
        try:
            course = await Course.get(enrollment.curso_id)
            if course:
                enrollment_dict["curso_nombre"] = course.nombre_programa
                enrollment_dict["curso_codigo"] = course.codigo
            else:
                enrollment_dict["curso_nombre"] = None
                enrollment_dict["curso_codigo"] = None
        except Exception:
            enrollment_dict["curso_nombre"] = None
            enrollment_dict["curso_codigo"] = None
    else:
        enrollment_dict["curso_nombre"] = None
        enrollment_dict["curso_codigo"] = None

    return enrollment_dict


async def enrich_enrollments_batch(enrollments: List[Enrollment]) -> List[dict]:
    """
    F-FIX-DESCONOCIDO-ENROLLMENTS (2026-08-09, Kevin): versión optimizada para
    listas. Hace 2 queries batch (students + courses con In) en vez de N
    queries individuales de enrich_enrollment_dates.

    Uso: en /api/v1/enrollments/ (lista) y en cualquier endpoint que devuelva
    multiples enrollments. Misma estructura de response que enrich_enrollment_dates.
    """
    if not enrollments:
        return []

    from core.timezone_utils import to_bolivia_time
    from core.cache import get_students_bulk_cached

    # Recolectar IDs unicos
    estudiante_ids = list({e.estudiante_id for e in enrollments if e.estudiante_id})
    curso_ids = list({e.curso_id for e in enrollments if e.curso_id})

    # 1 query batch a students (con cache de F-CACHE-SHARED)
    students_map = {}
    if estudiante_ids:
        try:
            students_map = await get_students_bulk_cached(estudiante_ids)
        except Exception:
            students_map = {}

    # 1 query batch a courses
    courses_map: dict = {}
    if curso_ids:
        try:
            courses_list = await Course.find(In(Course.id, curso_ids)).to_list()
            for c in courses_list:
                # Map por str(id) y por id (ObjectId) para cubrir cualquier lookup
                key = str(c.id)
                courses_map[key] = c
                courses_map[c.id] = c
        except Exception:
            pass

    enriched = []
    for e in enrollments:
        d = e.model_dump()
        d["fecha_inscripcion"] = to_bolivia_time(e.fecha_inscripcion)
        d["created_at"] = to_bolivia_time(e.created_at)
        d["updated_at"] = to_bolivia_time(e.updated_at)

        # descuento MAX (mismo que enrich_enrollment_dates)
        desc_curso = float(e.descuento_curso_aplicado or 0)
        desc_personal = float(e.descuento_personalizado or 0) if e.descuento_personalizado is not None else 0.0
        descuento_efectivo = max(desc_curso, desc_personal)
        if descuento_efectivo > 0 and desc_personal > 0 and desc_curso >= desc_personal:
            origen = "curso"
            advertencia = (
                f"El descuento personal seleccionado ({desc_personal:.0f}%) es menor al "
                f"descuento del curso ({desc_curso:.0f}%). Se aplicó el descuento de mayor "
                f"porcentaje: el del curso ({desc_curso:.0f}%)."
            )
        elif descuento_efectivo > 0 and desc_personal > desc_curso:
            origen = "personal"
            advertencia = None
        elif descuento_efectivo > 0 and desc_curso > 0:
            origen = "curso"
            advertencia = None
        else:
            origen = "ninguno"
            advertencia = None
        d["descuento_efectivo"] = descuento_efectivo
        d["descuento_efectivo_origen"] = origen
        d["advertencia_descuento"] = advertencia

        # Joinear estudiante
        if e.estudiante_id:
            student = students_map.get(e.estudiante_id) or students_map.get(str(e.estudiante_id))
            if student:
                d["estudiante_nombre"] = student.get("nombre") if isinstance(student, dict) else getattr(student, "nombre", None)
                d["estudiante_registro"] = student.get("registro") if isinstance(student, dict) else getattr(student, "registro", None)
                d["estudiante_ci"] = student.get("carnet") if isinstance(student, dict) else getattr(student, "carnet", None)
            else:
                d["estudiante_nombre"] = None
                d["estudiante_registro"] = None
                d["estudiante_ci"] = None
        else:
            d["estudiante_nombre"] = None
            d["estudiante_registro"] = None
            d["estudiante_ci"] = None

        # Joinear curso
        if e.curso_id:
            course = courses_map.get(e.curso_id) or courses_map.get(str(e.curso_id))
            if course:
                d["curso_nombre"] = course.nombre_programa if not isinstance(course, dict) else course.get("nombre_programa")
                d["curso_codigo"] = course.codigo if not isinstance(course, dict) else course.get("codigo")
            else:
                d["curso_nombre"] = None
                d["curso_codigo"] = None
        else:
            d["curso_nombre"] = None
            d["curso_codigo"] = None

        enriched.append(d)

    return enriched


async def get_enrollment(id: PydanticObjectId) -> Optional[Enrollment]:
    """
    F-CACHE-SHARED (2026-08-08, Kevin): ahora usa el cache compartido
    en memoria (TTL 30s) para evitar round-trip a Mongo en cada llamada.

    Importante: el cache retorna un dict (de motor), NO un objeto Beanie.
    El codigo que llama a esta funcion debe estar preparado para recibir
    cualquiera de los dos. Beanie Enrollment tiene atributos (.id, .curso_id, etc.)
    que el dict no tiene directamente (usa keys ['_id'], ['curso_id']).

    Si necesitas acceso a campos especificos, usa .get('campo') con fallback
    o _to_beanie() para convertir el dict a objeto Beanie.

    Para mantener compatibilidad maxima, esta funcion:
    1. Si el cache esta deshabilitado, hace Enrollment.get(id) normal
    2. Si el cache retorna un dict, lo convierte a Enrollment con Pydantic
       (perdida de performance minima, ~5ms por conversion)
    3. Si no esta en cache, lo busca y guarda en cache
    """
    from core.cache import get_enrollment_cached, cache_enabled
    from models.enrollment import Enrollment as _Enrollment

    if not cache_enabled():
        return await Enrollment.get(id)

    cached_dict = await get_enrollment_cached(id)
    if cached_dict is None:
        return None

    # Convertir dict a Enrollment (Beanie) para mantener compatibilidad con
    # todos los callers que esperan un objeto Beanie (no dict).
    try:
        return _Enrollment(**{k: v for k, v in cached_dict.items() if k != "_found"})
    except Exception:
        # Si falla la conversion (schema cambio, campo nuevo requerido),
        # caer al Enrollment.get(id) directo para no romper.
        return await Enrollment.get(id)
    """
    F-CACHE-SHARED (2026-08-08, Kevin): ahora usa el cache compartido
    en memoria (TTL 30s) para evitar round-trip a Mongo en cada llamada.

    Importante: el cache retorna un dict (de motor), NO un objeto Beanie.
    El codigo que llama a esta funcion debe estar preparado para recibir
    cualquiera de los dos. Beanie Enrollment tiene atributos (.id, .curso_id, etc.)
    que el dict no tiene directamente (usa keys ['_id'], ['curso_id']).

    Si necesitas acceso a campos especificos, usa .get('campo') con fallback
    o _to_beanie() para convertir el dict a objeto Beanie.

    Para mantener compatibilidad maxima, esta funcion:
    1. Si el cache esta deshabilitado, hace Enrollment.get(id) normal
    2. Si el cache retorna un dict, lo convierte a Enrollment con Pydantic
       (perdida de performance minima, ~5ms por conversion)
    3. Si no esta en cache, lo busca y guarda en cache
    """
    if not _cache_enabled():
        return await Enrollment.get(id)

    from core.cache import get_enrollment_cached
    from models.enrollment import Enrollment as _Enrollment

    cached_dict = await get_enrollment_cached(id)
    if cached_dict is None:
        return None

    # Convertir dict a Enrollment (Beanie) para mantener compatibilidad con
    # todos los callers que esperan un objeto Beanie (no dict).
    try:
        return _Enrollment(**{k: v for k, v in cached_dict.items() if k != "_found"})
    except Exception:
        # Si falla la conversion (schema cambio, campo nuevo requerido),
        # caer al Enrollment.get(id) directo para no romper.
        return await Enrollment.get(id)


async def get_enrollments_by_student(student_id: PydanticObjectId) -> List[Enrollment]:
    return await Enrollment.find(Enrollment.estudiante_id == student_id).to_list()


async def get_enrollments_by_course(course_id: PydanticObjectId) -> List[Enrollment]:
    return await Enrollment.find(Enrollment.curso_id == course_id).to_list()


async def get_all_enrollments(
    page: int = 1,
    per_page: int = 10,
    q: Optional[str] = None,
    estado: Optional[EstadoInscripcion] = None,
    curso_id: Optional[PydanticObjectId] = None,
    estudiante_id: Optional[PydanticObjectId] = None,
    cursos_permitidos: Optional[List[PydanticObjectId]] = None,
    con_descuento: Optional[bool] = None,
    descuento_id: Optional[PydanticObjectId] = None,
    requiere_accion_documentos: Optional[bool] = None
) -> tuple[List[Enrollment], int]:
    """
    cursos_permitidos (ISSUE-R-ROLES): si se provee (no None), restringe los resultados
    únicamente a inscripciones de esos cursos. Se usa para segmentar el rol ENCARGADO_CURSO
    (y en el futuro COBRANZA, ISSUE-P-SEGMENTACION) a sus cursos asignados.

    con_descuento/descuento_id (fix 2026-07-09, reportado por el usuario: "no pude
    seleccionar los que están con descuento y cuál descuento"): permiten filtrar
    la tabla de inscripciones por si tienen algún descuento personal aplicado
    (True) o no (False), y opcionalmente por un Discount específico.
    """
    query = Enrollment.find()
    
    if estado:
        query = query.find(Enrollment.estado == estado)
    if curso_id:
        query = query.find(Enrollment.curso_id == curso_id)
    if estudiante_id:
        query = query.find(Enrollment.estudiante_id == estudiante_id)
    if cursos_permitidos is not None:
        query = query.find(In(Enrollment.curso_id, cursos_permitidos))
    if descuento_id:
        query = query.find(Enrollment.descuento_estudiante_id == descuento_id)
    elif con_descuento is True:
        query = query.find(
            Or(
                Enrollment.descuento_estudiante_id != None,
                Enrollment.descuento_personalizado > 0
            )
        )
    elif con_descuento is False:
        query = query.find(
            Enrollment.descuento_estudiante_id == None,
            Or(Enrollment.descuento_personalizado == None, Enrollment.descuento_personalizado <= 0)
        )
        
    if requiere_accion_documentos:
        # ISSUE-Q-DOCUMENTOS-KYC: Filtrar inscripciones que tengan algún documento pendiente de validación o subida
        query = query.find({"requisitos.estado": {"$in": ["pendiente", "en_proceso", "rechazado", "sin_subir"]}})
        
    if q:
        regex_pattern = {"$regex": q, "$options": "i"}
        students = await Student.find(Or(Student.nombre == regex_pattern, Student.carnet == regex_pattern)).to_list()
        student_ids = [s.id for s in students]
        
        courses = await Course.find(Course.nombre_programa == regex_pattern).to_list()
        course_ids = [c.id for c in courses]
        
        query = query.find(Or(In(Enrollment.estudiante_id, student_ids), In(Enrollment.curso_id, course_ids)))
    
    total_count = await query.count()
    skip = (page - 1) * per_page
    
    enrollments = await query.sort("-fecha_inscripcion").skip(skip).limit(per_page).to_list()
    return enrollments, total_count


async def update_enrollment_descuento(
    enrollment_id: PydanticObjectId,
    descuento_personalizado: Optional[float],
    admin_username: str,
    descuento_id: Optional[PydanticObjectId] = None
) -> Enrollment:
    """
    AUDITORÍA (CRÍTICO #3): antes solo recalculaba total_a_pagar/saldo_pendiente
    a nivel global, sin tocar el costo de cada módulo individual ni disparar
    actualizar_saldo_enrollment -- el kardex por módulo (usado para el
    prorrateo en cascada) quedaba desincronizado del nuevo total. Ahora se
    redistribuye el costo por módulo con la misma proporción que
    create_enrollment, y se llama a la cascada real al final para que
    monto_pagado/estado por módulo y el estado de la inscripción queden
    consistentes con los pagos históricos ya aprobados.

    BUG ENCONTRADO (2026-07-09, reportado por el usuario: "asigné la beca
    pero no veo el recálculo"): `EnrollmentUpdate.descuento_id` existía en el
    schema desde siempre, pero esta función SOLO leía `descuento_personalizado`
    (el porcentaje libre) -- si el CPD seleccionaba un `Discount` real del
    combo "Descuento Personal (Beca)", ese `descuento_id` se ignoraba en
    silencio y el porcentaje aplicado quedaba en 0 (o en el valor libre
    anterior), sin recalcular nada. Ahora, si se provee `descuento_id`, se
    resuelve el `Discount` real (igual que en `create_enrollment`) y se usa
    su porcentaje; `descuento_personalizado` sigue funcionando como
    porcentaje libre si no se selecciona un `Discount` concreto.
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError(f"Inscripción {enrollment_id} no encontrada")

    descuento_estudiante_id = enrollment.descuento_estudiante_id
    nota_minima_snapshot = enrollment.nota_minima_beca

    if descuento_id is not None:
        discount_sel = await Discount.get(descuento_id)
        if not discount_sel:
            raise ValueError(f"Descuento {descuento_id} no encontrado")
        if not discount_sel.activo:
            raise ValueError(f"El descuento '{discount_sel.nombre}' no está activo")
        descuento_personalizado = discount_sel.porcentaje
        descuento_estudiante_id = discount_sel.id
        nota_minima_snapshot = discount_sel.nota_minima_requerida
    elif descuento_personalizado is not None:
        # Porcentaje libre (sin vincular a un Discount concreto): se limpia
        # la referencia a un Discount anterior, si había, para no dejar un
        # descuento_estudiante_id apuntando a un porcentaje que ya no aplica.
        descuento_estudiante_id = None
        nota_minima_snapshot = None
    else:
        descuento_personalizado = enrollment.descuento_personalizado or 0.0

    total_con_descuento_curso = enrollment.costo_total - (enrollment.costo_total * enrollment.descuento_curso_aplicado / 100)
    # F-LOGICA-DESCUENTOS-MAX (2026-08-05, Kevin): "se queda con el descuento
    # de mayor porcentaje". Si el personal es menor, se descarta y se avisa.
    descuento_efectivo_rec = max(enrollment.descuento_curso_aplicado, descuento_personalizado)
    colegiatura_final = enrollment.costo_total - (enrollment.costo_total * descuento_efectivo_rec / 100)
    # ISSUE-P-CARGO-MULTIITEM: el cargo adicional (snapshot de ítems de esta
    # inscripción) se mantiene fuera del recálculo por descuento -- ningún
    # ítem recibe descuentos de curso/estudiante, se preserva íntegro.
    total_final = colegiatura_final + enrollment.costo_matricula + enrollment.get_cargo_adicional_total()

    # Redistribuir el costo de cada módulo con la misma proporción que tenía
    # el curso original (mismo patrón que create_enrollment), preservando
    # monto_pagado/nota/estado_academico de cada ModuloEstado. También se
    # recalcula costo_sin_beca_personal (ISSUE-P-RECALCULO-NOTA) usando la
    # misma proporción, para que la pérdida de beca por nota siga funcionando
    # correctamente después de reasignar el descuento desde este endpoint.
    if enrollment.modulos:
        suma_costo_modulos_actual = sum(m.costo for m in enrollment.modulos)
        total_asignado = 0.0
        total_asignado_sin_beca = 0.0
        for i, mod in enumerate(enrollment.modulos):
            es_ultimo = i == len(enrollment.modulos) - 1
            if suma_costo_modulos_actual > 0:
                proporcion = mod.costo / suma_costo_modulos_actual
            else:
                proporcion = 1 / len(enrollment.modulos)

            if es_ultimo:
                mod.costo = max(0.0, round(colegiatura_final - total_asignado, 2))
            else:
                mod.costo = round(proporcion * colegiatura_final, 2)
                total_asignado += mod.costo
            # F-FIX-ESTADO-MODULOS-POST-DESCUENTO (2026-08-08, Kevin): recalcular
            # estado del modulo despues de cambiar su costo. Sin esto, los
            # becados quedan en "Parcial" aunque monto_pagado cubra el nuevo
            # costo con descuento.
            _recalcular_estado_modulo(mod)

            if nota_minima_snapshot is not None:
                if es_ultimo:
                    mod.costo_sin_beca_personal = max(0.0, round(total_con_descuento_curso - total_asignado_sin_beca, 2))
                else:
                    mod.costo_sin_beca_personal = round(proporcion * total_con_descuento_curso, 2)
                    total_asignado_sin_beca += mod.costo_sin_beca_personal
            else:
                mod.costo_sin_beca_personal = None

    enrollment.descuento_personalizado = descuento_personalizado
    enrollment.descuento_estudiante_id = descuento_estudiante_id
    enrollment.nota_minima_beca = nota_minima_snapshot
    enrollment.total_a_pagar = round(total_final, 2)
    # Valor provisional para no violar el validador de saldo_pendiente antes
    # del primer save(); actualizar_saldo_enrollment recompone el definitivo
    # justo debajo a partir de los pagos históricos reales.
    enrollment.saldo_pendiente = round(max(0.0, total_final - enrollment.total_pagado), 2)
    enrollment.updated_at = utcnow_naive()
    
    await enrollment.save()

    await actualizar_saldo_enrollment(enrollment_id)
    return await Enrollment.get(enrollment_id)


async def cambiar_estado_enrollment(
    enrollment_id: PydanticObjectId,
    nuevo_estado: EstadoInscripcion,
    admin_username: str
) -> Enrollment:
    """
    AUDITORÍA (CRÍTICO #5): este endpoint genérico (PATCH /enrollments/{id})
    permitía fijar estado=SUSPENDIDO directamente, sin pasar por
    PassiveRequest ni congelado_service -- sin motivo_suspension, sin
    notificar al estudiante, sin cobrar la tasa de congelamiento. También
    permitía "sacar" una inscripción de SUSPENDIDO sin limpiar
    motivo_suspension/fecha_congelamiento/multa_reincorporacion_pendiente,
    dejando esos campos inconsistentes con el nuevo estado. Ambos casos
    ahora se bloquean explícitamente; deben usarse los endpoints dedicados
    (/passive-requests/, /enrollments/{id}/congelar,
    /enrollments/{id}/reactivar-congelado).

    F-083 (2026-07-28): también bloquea ir directo a RETIRADO. El retiro
    requiere registrar motivo_retiro, fecha_retiro, retirado_por, y debe
    usar el endpoint dedicado /enrollments/{id}/retirar.
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError(f"Inscripción {enrollment_id} no encontrada")

    if nuevo_estado == EstadoInscripcion.SUSPENDIDO:
        raise ValueError(
            "No se puede suspender una inscripción directamente. Usa el flujo de "
            "Solicitud de Pasivo o el de Congelamiento, que registran motivo y notifican al estudiante."
        )

    if nuevo_estado == EstadoInscripcion.RETIRADO:
        raise ValueError(
            "F-083: No se puede retirar una inscripción directamente. "
            "Usa el endpoint /enrollments/{id}/retirar que registra motivo_retiro y notifica al estudiante."
        )

    if enrollment.estado == EstadoInscripcion.SUSPENDIDO:
        raise ValueError(
            "Esta inscripción está suspendida (pasivo/congelado/abandono). "
            "Usa el endpoint de reactivación correspondiente en vez de cambiar el estado directamente."
        )

    if enrollment.estado == EstadoInscripcion.RETIRADO:
        raise ValueError(
            "F-083: Esta inscripción está RETIRADA. El retiro es definitivo: no se puede revertir a "
            "un estado anterior. Si el estudiante quiere volver, debe crear una nueva inscripción."
        )

    enrollment.estado = nuevo_estado
    enrollment.updated_at = utcnow_naive()

    await enrollment.save()
    return enrollment


# ========================================================================
# F-083 (2026-07-28): RETIRO VOLUNTARIO DE INSCRIPCIÓN
# ========================================================================
async def retirar_inscripcion(
    enrollment_id: PydanticObjectId,
    motivo_retiro: str,
    retirado_por: str,
    notificar_estudiante: bool = True
) -> Enrollment:
    """
    F-083 (2026-07-28): marca una inscripción como RETIRADO (abandono
    DEFINITIVO, no vuelve). Distinto de SUSPENDIDO+abandono (que es
    automático por inactividad y genera multa de reincorporación).

    Reglas de negocio (definidas con Lic. Sorich Cobranza y Lic. Sandra
    Zabala Cobranza, 2026-07-28 12:50 vía WhatsApp):
    - "retirados ya no vuelven, no son pasivos; pasivo tiene la opción
      de volver luego, y retirados ya no vuelven" (Lic. Sorich)
    - "esos ya no debería sumar sus pagos para cuentas por cobrar,
      solo queda lo que pagaron y se cierra" (Lic. Sorich)
    - El retiro es VOLUNTARIO (decisión del estudiante o del CPD/admin).
      Se diferencia del abandono automático (que es por inactividad y SÍ
      genera multa de reincorporación al volver).
    - El retiro NO es reversible. Si el estudiante se arrepiente, debe
      crear una nueva inscripción.
    - Lo que el estudiante YA pagó SÍ cuenta como ingreso (no se le
      descuenta). Lo que falta NO se cobra (no suma a "Por Cobrar").

    Args:
        enrollment_id: ID de la inscripción a retirar
        motivo_retiro: motivo del retiro (obligatorio, ej: 'cambio de
            ciudad', 'problemas económicos'). Se persiste en enrollment.
        retirado_por: username del usuario que ejecuta el retiro
            (admin/cpd/superadmin) o 'estudiante' si fue autoservicio.
        notificar_estudiante: si True (default), envía notification
            in-app al estudiante confirmando el retiro.

    Raises:
        ValueError: si la inscripción no existe, o si ya está en un
            estado terminal (COMPLETADO, CANCELADO, RETIRADO).

    Returns:
        Enrollment actualizado (estado=RETIRADO, motivo_retiro=...,
        fecha_retiro=now, retirado_por=...).
    """
    from core.timezone_utils import utcnow_naive

    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError(f"Inscripción {enrollment_id} no encontrada")

    # Validar estado actual: NO se puede retirar si ya es terminal
    estados_terminales = {
        EstadoInscripcion.COMPLETADO,
        EstadoInscripcion.CANCELADO,
        EstadoInscripcion.RETIRADO,
    }
    if enrollment.estado in estados_terminales:
        raise ValueError(
            f"F-083: No se puede retirar una inscripción en estado '{enrollment.estado.value}'. "
            f"El retiro es solo para inscripciones activas o suspendidas (pasivo/congelado)."
        )

    if not motivo_retiro or not motivo_retiro.strip():
        raise ValueError("F-083: El motivo del retiro es obligatorio.")

    # Marcar como RETIRADO (mantiene módulos, saldos, pagos históricos)
    estado_anterior = enrollment.estado
    enrollment.estado = EstadoInscripcion.RETIRADO
    enrollment.motivo_retiro = motivo_retiro.strip()
    enrollment.fecha_retiro = utcnow_naive()
    enrollment.retirado_por = retirado_por
    enrollment.updated_at = utcnow_naive()

    # Limpiar campos de suspensión si venía de SUSPENDIDO (porque ya no
    # está "suspendido", está "retirado" definitivo)
    if estado_anterior == EstadoInscripcion.SUSPENDIDO:
        enrollment.motivo_suspension = None
        enrollment.fecha_congelamiento = None
        enrollment.tasa_congelamiento_pagada = False
        enrollment.fecha_abandono = None
        enrollment.multa_reincorporacion_pendiente = False
        enrollment.mora_notificada = False

    await enrollment.save()

    # Notificar al estudiante
    if notificar_estudiante:
        try:
            from services.notification_service import create_notification
            await create_notification(
                destinatario_id=enrollment.estudiante_id,
                tipo_destinatario="student",
                titulo="Tu inscripción fue retirada",
                mensaje=(
                    f"Tu inscripción al curso fue marcada como RETIRADO. "
                    f"Motivo: {motivo_retiro.strip()}. "
                    f"Lo que ya pagaste queda registrado como ingreso a tu favor; "
                    f"no se te cobrará el saldo pendiente. "
                    f"Si tienes dudas, contacta al CPD."
                ),
                tipo_alerta="warning",
                ruta="/app/enrollments",
                referencia_tipo="enrollment",
                referencia_id=enrollment.id
            )
        except Exception as notif_err:
            # Si la notification falla, no rompemos el flujo principal
            import logging
            logging.getLogger("kyc.enrollment").warning(
                f"F-083: no se pudo notificar al estudiante {enrollment.estudiante_id}: {notif_err}"
            )

    return enrollment


# ========================================================================
# ISSUE-F-PRORRATEO: ALGORITMO FINANCIERO EN CASCADA (V2 - RECONSTRUCCIÓN ABSOLUTA)
# ========================================================================
async def actualizar_saldo_enrollment(
    enrollment_id: PydanticObjectId,
    monto_pago_aprobado: float = 0.0, # Se mantiene la firma por compatibilidad, pero ya no se usa ciegamente
    enrollment: Optional[Enrollment] = None,
    pagos_aprobados: Optional[list] = None
):
    """
    Reconstruye el saldo de la inscripción sumando todos los pagos históricos aprobados
    menos los anulados. Distribuye los fondos aprobados en cascada (Waterfall) reparando
    cualquier inconsistencia de base de datos.

    F-COBRANZA-014 (2026-07-21): ahora descuenta los pagos ANULADOS del `total_pagado`
    y `saldo_pendiente`. Antes, esos campos reflejaban solo la suma de aprobados,
    dejando un desfase de 3,534 BOB en producción que Joel detectó. La distribución
    de módulos (waterfall) sigue basándose en aprobados, ya que los módulos
    representan el estado ACADÉMICO histórico (un módulo que se pagó y luego se
    anuló queda en "Pagado" hasta que se reinscribe o se vuelve a pagar).

    F-085 (2026-07-28): REVERTIR la regla F-COBRANZA-014. La fórmula
    `total_pagado = aprobados - anulados` producía números NEGATIVOS en
    cualquier caso donde el monto del pago anulado era mayor que el resto
    aprobado (ej: Luis Fernando con matrícula 300 aprobado + Módulo 2940
    anulado → total_pagado = -2640). Esto rompía 5 endpoints financieros
    con ValidationError `total_pagado >= 0` y dejaba TODAS las páginas
    financieras en blanco.

    REGLA CORRECTA (idempotente, no produce negativos):
      `total_pagado = sum(pagos APROBADOS)`
    Al anular un pago aprobado, NO se resta del total — el pago simplemente
    deja de contar en aprobados. Esto es la misma regla que aplicamos en
    F-082 (Medardo) y F-084 (Anselmo) y es matemáticamente consistente.

    El "desfase" que F-COBRANZA-014 detectó en realidad era un BUG en la
    cascada (falla silenciosa de `actualizar_saldo_enrollment` por
    `RevisionIdWasChanged` que NO actualizaba `total_pagado` al aprobar).
    El fix correcto es el retry + notification (F-082), no la resta
    posterior de anulados.

    OPTIMIZACIÓN DE IMPORTACIÓN MASIVA (2026-07-09, ISSUE-Q-IMPORT-TIMEOUT):
    `enrollment`/`pagos_aprobados` permiten pasar el documento y los pagos ya
    obtenidos en memoria, evitando el `Enrollment.get()` + `Payment.find()`
    redundantes cuando el llamador (ej. import_students_from_excel) acaba de
    crear ambos en el mismo flujo.
    """
    from models.payment import Payment
    from models.enums import EstadoPago

    if enrollment is None:
        enrollment = await Enrollment.get(enrollment_id)
        if not enrollment:
            raise ValueError(f"Inscripción {enrollment_id} no encontrada")

    # 1. Recolectar la verdad absoluta: ¿Cuánto dinero REALMENTE tiene aprobado este alumno?
    if pagos_aprobados is None:
        pagos_aprobados = await Payment.find(
            Payment.inscripcion_id == enrollment_id,
            Payment.estado_pago == EstadoPago.APROBADO
        ).to_list()

    # F-085: NO recolectar anulados ni restarlos. La regla es:
    #   total_pagado = sum(aprobados)
    # Los anulados ya no cuentan en aprobados (cambiaron de estado),
    # restarlos de nuevo = doble resta = números negativos.

    dinero_aprobado_bruto = sum(p.cantidad_pago for p in pagos_aprobados)
    # F-085: total_pagado = sum(aprobados). Sin restar anulados.
    dinero_neto_pagado = round(dinero_aprobado_bruto, 2)
    tanque_de_agua = round(dinero_aprobado_bruto, 2)  # waterfall con aprobados

    # 2. Reiniciar los contadores de la inscripción a cero para reconstruirlos
    enrollment.matricula_pagada = False
    for mod in enrollment.modulos:
        mod.monto_pagado = 0.0
        mod.estado = "Pendiente"

    # 3. PASO 1: Cubrir la matrícula administrativa obligatoria
    if tanque_de_agua >= enrollment.costo_matricula:
        tanque_de_agua = round(tanque_de_agua - enrollment.costo_matricula, 2)
        enrollment.matricula_pagada = True

        # REGLA DE NEGOCIO UAGRM: Al pagar matrícula, el alumno pasa a ser "Activo"
        if enrollment.estado == EstadoInscripcion.PENDIENTE_PAGO:
            enrollment.estado = EstadoInscripcion.ACTIVO
    else:
        # El dinero no alcanzó ni para la matrícula
        enrollment.matricula_pagada = False
        tanque_de_agua = 0.0

    # 4. PASO 2: Cascada sobre los módulos académicos
    if tanque_de_agua > 0:
        for mod in enrollment.modulos:
            if tanque_de_agua <= 0.01:
                break # Se acabó el dinero

            costo_modulo = round(mod.costo, 2)

            if tanque_de_agua >= costo_modulo:
                # El dinero cubre este módulo completamente
                mod.monto_pagado = costo_modulo
                mod.estado = "Pagado"
                tanque_de_agua = round(tanque_de_agua - costo_modulo, 2)
            else:
                # El dinero restante se vierte en este módulo (Pago parcial)
                mod.monto_pagado = round(tanque_de_agua, 2)
                mod.estado = "Parcial"
                tanque_de_agua = 0.0

    # 5. Actualizar los totales globales (Total Pagado y Saldo Pendiente)
    # F-085 (2026-07-28): `total_pagado = sum(aprobados)`. NO restar anulados
    # (la resta introducida por F-COBRANZA-014 producía números negativos).
    enrollment.total_pagado = dinero_neto_pagado
    enrollment.saldo_pendiente = max(0.0, round(enrollment.total_a_pagar - dinero_neto_pagado, 2))
    
    # 6. Evolución a "Completado" si la deuda llegó a cero
    if enrollment.esta_completamente_pagado() and enrollment.matricula_pagada:
        enrollment.estado = EstadoInscripcion.COMPLETADO
    elif not enrollment.esta_completamente_pagado() and enrollment.matricula_pagada:
        # Prevención de retroceso en caso de reversión de pagos
        enrollment.estado = EstadoInscripcion.ACTIVO
    elif enrollment.matricula_exenta and enrollment.estado == EstadoInscripcion.PENDIENTE_PAGO:
        # ISSUE-M-EXENCION: MAE autorizó cursar sin haber pagado la matrícula.
        # NO se toca matricula_pagada (sigue reflejando la realidad financiera
        # para reportes/caja); solo se desbloquea el estado académico.
        enrollment.estado = EstadoInscripcion.ACTIVO
    
    enrollment.updated_at = utcnow_naive()
    await enrollment.save()


# ========================================================================
# ISSUE-M-EXENCION: BYPASS DE MATRÍCULA (SOLO MAE)
# ========================================================================
async def otorgar_matricula_exenta(enrollment_id: PydanticObjectId, otorgado_por: str) -> Enrollment:
    """
    Autoriza a un estudiante a cursar académicamente SIN haber pagado la
    matrícula institucional. NO condona la deuda: `saldo_pendiente` y
    `matricula_pagada` siguen reflejando la realidad financiera exacta
    (Cobranza sigue viendo y cobrando la deuda con normalidad).

    Solo desbloquea el estado académico (`estado` pasa a ACTIVO) para que
    el estudiante pueda acceder a módulos/aula virtual mientras se resuelve
    su situación de matrícula.
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError(f"Inscripción {enrollment_id} no encontrada")

    if enrollment.estado in (EstadoInscripcion.COMPLETADO, EstadoInscripcion.CANCELADO):
        raise ValueError(f"No se puede otorgar exención sobre una inscripción en estado '{enrollment.estado.value}'")

    enrollment.matricula_exenta = True
    enrollment.matricula_exenta_otorgada_por = otorgado_por
    enrollment.matricula_exenta_fecha = utcnow_naive()

    # Desbloqueo académico inmediato (no depende de que llegue un pago nuevo)
    if enrollment.estado == EstadoInscripcion.PENDIENTE_PAGO:
        enrollment.estado = EstadoInscripcion.ACTIVO

    enrollment.updated_at = utcnow_naive()
    await enrollment.save()
    return enrollment


async def revocar_matricula_exenta(enrollment_id: PydanticObjectId) -> Enrollment:
    """
    Revoca una exención de matrícula previamente otorgada. Si la matrícula
    real sigue sin pagarse, el estado académico vuelve a PENDIENTE_PAGO
    (se re-bloquea el acceso). Si mientras tanto ya se pagó la matrícula de
    forma real, el estado ACTIVO se mantiene por la vía normal.
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError(f"Inscripción {enrollment_id} no encontrada")

    if not enrollment.matricula_exenta:
        raise ValueError("Esta inscripción no tiene una exención de matrícula activa")

    enrollment.matricula_exenta = False

    if not enrollment.matricula_pagada and enrollment.estado == EstadoInscripcion.ACTIVO:
        enrollment.estado = EstadoInscripcion.PENDIENTE_PAGO

    enrollment.updated_at = utcnow_naive()
    await enrollment.save()
    return enrollment


# ========================================================================
# LÓGICA ACADÉMICA (ISSUE P)
# ========================================================================
async def actualizar_nota_modulo(
    enrollment_id: PydanticObjectId, 
    modulo_index: int, 
    nota: float, 
    evaluador_username: str
) -> Enrollment:
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError(f"Inscripción no encontrada")
        
    if modulo_index < 0 or modulo_index >= len(enrollment.modulos):
        raise ValueError(f"Índice de módulo {modulo_index} fuera de rango")
        
    modulo = enrollment.modulos[modulo_index]
    modulo.nota = round(nota, 2)
    
    if nota >= 64.0:
        modulo.estado_academico = "Aprobado"
    else:
        modulo.estado_academico = "Reprobado"

    # ========================================================================
    # ISSUE-P-RECALCULO-NOTA: pérdida de beca por módulo si no se mantiene la nota mínima
    # ========================================================================
    recalculo_necesario = False
    if (
        enrollment.nota_minima_beca is not None
        and modulo.costo_sin_beca_personal is not None
        and nota < enrollment.nota_minima_beca
        and modulo.costo != modulo.costo_sin_beca_personal
    ):
        modulo.costo = modulo.costo_sin_beca_personal
        recalculo_necesario = True

    notas_evaluadas = [m.nota for m in enrollment.modulos if m.nota is not None]
    if notas_evaluadas:
        promedio = sum(notas_evaluadas) / len(notas_evaluadas)
        enrollment.nota_final = round(promedio, 2)

    if recalculo_necesario:
        # El módulo afectado ya no vale lo mismo: recalculamos el total a pagar
        # (matrícula + costos de módulo actualizados). saldo_pendiente se actualiza
        # aquí también (aunque sea una aproximación) para no violar el validador del
        # modelo; el valor definitivo lo recompone actualizar_saldo_enrollment justo debajo.
        enrollment.total_a_pagar = round(
            enrollment.costo_matricula + sum(m.costo for m in enrollment.modulos), 2
        )
        enrollment.saldo_pendiente = round(max(0.0, enrollment.total_a_pagar - enrollment.total_pagado), 2)

    enrollment.updated_at = utcnow_naive()
    await enrollment.save()

    if recalculo_necesario:
        # Reconstruye monto_pagado/estado por módulo y saldo_pendiente global a partir
        # de los pagos históricos aprobados (misma cascada de ISSUE-F-PRORRATEO).
        # No modifica ni elimina ningún registro de Payment.
        await actualizar_saldo_enrollment(enrollment_id)
        enrollment = await Enrollment.get(enrollment_id)

    return enrollment


# ========================================================================
# F-NOTAS-MODULOS-EJECUTADOS (2026-08-18, decisión de Kevin)
# ========================================================================
# Un programa que arranca a mitad de camino (ej. entra en el módulo 5) tiene
# los módulos anteriores ya dictados, con nota. La carga inicial trae pagos
# por módulo pero NUNCA trajo notas — eso dejaba a esos estudiantes con el
# historial académico en blanco hasta que alguien las cargara a mano, módulo
# por módulo, desde la libreta.
#
# Kevin eligió: "un Excel aparte, solo de notas", para estudiantes que YA
# EXISTEN en el sistema (a diferencia de la carga inicial, que tambien crea
# estudiantes nuevos). Columnas: CI + "Nota Modulo N" por cada modulo con
# nota conocida.
import re as _re


def _detectar_columnas_notas(header_row: list) -> tuple:
    """
    Busca la columna de CI/carnet y las columnas "Nota Modulo N".

    Devuelve (col_carnet, [(col_idx, modulo_index_0based), ...]).
    col_carnet es 0 si no se encontro.
    """
    col_carnet = 0
    columnas_notas = []
    for idx, header in enumerate(header_row, start=1):
        if not header:
            continue
        if col_carnet == 0 and (
            header == "ci" or "carnet" in header or "cedula" in header or "identidad" in header
        ):
            col_carnet = idx
            continue
        # "nota modulo 1", "notamodulo1", "nota m1": exige la palabra "nota"
        # Y un numero, para no confundir con la columna "Modulo" que trae el
        # numero de modulo del estudiante (el mismo bug que hubo en el pago
        # fantasma de 1 Bs del Excel de carga inicial).
        if "nota" in header:
            m = _re.search(r"(\d+)", header)
            if m:
                columnas_notas.append((idx, int(m.group(1)) - 1))
    return col_carnet, columnas_notas


async def cargar_notas_modulos_excel(
    file_content: bytes,
    curso_id: PydanticObjectId,
    evaluador_username: str,
) -> dict:
    """
    Carga notas de modulos YA EJECUTADOS para estudiantes que ya estan
    inscritos en `curso_id`, desde un Excel de CI + "Nota Modulo N".

    A diferencia de la carga inicial, NO crea estudiantes ni inscripciones:
    si el CI no tiene inscripcion en este curso, esa fila se reporta como
    fallida y se sigue con las demas.

    Reutiliza `actualizar_nota_modulo` por cada (estudiante, modulo) en vez
    de escribir el campo directo, para no perderse el recalculo de perdida
    de beca por nota ni la actualizacion del promedio — la misma logica que
    ya usa el flujo manual desde la libreta.
    """
    import openpyxl
    from io import BytesIO
    from services.student_service import _normalize_header

    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
    except Exception as e:
        raise ValueError(f"No se pudo parsear el archivo: {e}")

    if not sheet or sheet.max_row < 2:
        raise ValueError("El archivo no contiene datos.")

    header_row = [
        _normalize_header(sheet.cell(row=1, column=c).value)
        for c in range(1, sheet.max_column + 1)
    ]
    col_carnet, columnas_notas = _detectar_columnas_notas(header_row)

    if col_carnet == 0:
        raise ValueError("No se encontro la columna de CI/Carnet en el Excel.")
    if not columnas_notas:
        raise ValueError(
            "No se encontraron columnas de notas. Deben llamarse 'Nota Modulo 1', "
            "'Nota Modulo 2', etc."
        )

    # Cargar TODOS los enrollments de este curso de una vez, indexados por
    # estudiante, para no hacer una query por fila.
    enrollments = await Enrollment.find(Enrollment.curso_id == curso_id).to_list()
    if not enrollments:
        raise ValueError("Este curso no tiene ningun estudiante inscrito todavia.")

    estudiante_ids = list({e.estudiante_id for e in enrollments})
    students = await Student.find(In(Student.id, estudiante_ids)).to_list()
    carnet_a_estudiante_id = {s.carnet: s.id for s in students if s.carnet}
    estudiante_id_a_enrollment = {e.estudiante_id: e for e in enrollments}

    actualizados = 0
    fallidos: List[dict] = []

    for row_idx in range(2, sheet.max_row + 1):
        carnet_raw = sheet.cell(row=row_idx, column=col_carnet).value
        carnet = str(carnet_raw).strip() if carnet_raw is not None else ""
        if not carnet:
            continue  # fila vacia, no es un error

        estudiante_id = carnet_a_estudiante_id.get(carnet)
        enrollment = estudiante_id_a_enrollment.get(estudiante_id) if estudiante_id else None
        if not enrollment:
            fallidos.append({
                "fila": row_idx,
                "carnet": carnet,
                "motivo": "No tiene inscripcion en este curso (¿esta el CI bien escrito?)",
            })
            continue

        for col_idx, modulo_idx in columnas_notas:
            valor = sheet.cell(row=row_idx, column=col_idx).value
            if valor is None or str(valor).strip() == "":
                continue
            try:
                nota = float(valor)
            except (TypeError, ValueError):
                fallidos.append({
                    "fila": row_idx, "carnet": carnet,
                    "motivo": f"Nota invalida en modulo {modulo_idx + 1}: '{valor}'",
                })
                continue
            if modulo_idx < 0 or modulo_idx >= len(enrollment.modulos):
                fallidos.append({
                    "fila": row_idx, "carnet": carnet,
                    "motivo": f"El curso no tiene modulo {modulo_idx + 1}",
                })
                continue
            try:
                enrollment = await actualizar_nota_modulo(
                    enrollment.id, modulo_idx, nota, evaluador_username
                )
                actualizados += 1
            except Exception as e:
                fallidos.append({
                    "fila": row_idx, "carnet": carnet,
                    "motivo": f"Modulo {modulo_idx + 1}: {str(e)}",
                })

    return {"actualizados": actualizados, "fallidos": fallidos}


# ========================================================================
# F-FIX-DESCUENTO-ITEM (2026-08-05, Kevin): helper para recalcular el
# total_a_pagar de un enrollment YA EXISTENTE al cual se le acaba de
# asignar un descuento (individual o del curso). Aplica la logica MAX
# (max entre descuento del curso y descuento personal) y redistribuye
# el costo entre los modulos proporcionalmente.
#
# Usado por /courses/{id}/initial-enrollments cuando el item trae
# descuento_id o descuento_personalizado y el estudiante ya estaba
# inscrito (rama "existing").
# ========================================================================
async def _recalcular_total_enrollment(enrollment: Enrollment, course: Course) -> None:
    """Recalcula total_a_pagar + redistribuye costo entre modulos con la logica MAX.

    F-FIX-DESCUENTO-TOTAL-PAGAR (2026-08-05, Kevin): el bug era que
    `total_a_pagar` se calculaba desde `enrollment.costo_total` (sin
    descuento) en vez de sumar los `enrollment.modulos[].costo` (que ya
    estan con el descuento aplicado a cada modulo). Resultado: para
    estudiantes con 50% de descuento, total_a_pagar quedaba en 3240
    (2940 modulos + 300 matricula, sin descuento) en vez de 1770
    (1470 modulos con 50% + 300 matricula). El Por Cobrar del sistema
    quedaba inflado en ~Bs 11,550 para DIPL-IA-2026 y ~Bs 23,790
    sumando todos los programas.

    Fix: total_a_pagar = SUM(modulos[].costo) + costo_matricula + cargo_adicional.
    Si los modulos no tienen costo seteado, los redistribuimos primero
    desde colegiatura_final (que SI incluye el descuento)."""
    desc_curso = float(enrollment.descuento_curso_aplicado or 0)
    desc_personal = float(enrollment.descuento_personalizado or 0) if enrollment.descuento_personalizado is not None else 0.0
    # F-LOGICA-DESCUENTOS-MAX: el estudiante se queda con el descuento
    # de mayor porcentaje. Si personal > curso, gana personal. Si no,
    # gana el curso.
    descuento_efectivo = max(desc_curso, desc_personal)
    colegiatura_final = enrollment.costo_total - (enrollment.costo_total * descuento_efectivo / 100)
    cargo_adicional = enrollment.get_cargo_adicional_total()
    # Redistribuir el costo entre los modulos proporcionalmente (solo si
    # los modulos no tienen ya el costo seteado, ej. cuando se acaba de
    # asignar el descuento y los modulos siguen con el costo sin descuento).
    suma_costo_modulos_actual = sum(m.costo or 0 for m in enrollment.modulos)
    # Si los modulos tienen costo 0 o todos iguales al costo original del
    # curso, redistribuir. Si ya tienen costos con descuento aplicados
    # (modulos[i].costo < course.modulos[i].costo), respetarlos.
    if enrollment.modulos and all(
        m.costo == (course.modulos[i].costo if i < len(course.modulos) else 0)
        for i, m in enumerate(enrollment.modulos)
    ):
        # Modulos sin descuento previo: redistribuir colegiatura_final
        suma_costo_modulos = sum(m.costo for m in course.modulos)
        total_asignado = 0.0
        for i, mod in enumerate(enrollment.modulos):
            es_ultimo = i == len(enrollment.modulos) - 1
            if es_ultimo:
                mod.costo = max(0.0, round(colegiatura_final - total_asignado, 2))
            else:
                if suma_costo_modulos > 0:
                    proporcion = course.modulos[i].costo / suma_costo_modulos
                else:
                    proporcion = 1.0 / len(enrollment.modulos)
                mod.costo = round(proporcion * colegiatura_final, 2)
                total_asignado += mod.costo
            # F-FIX-ESTADO-MODULOS-POST-DESCUENTO (2026-08-08, Kevin): recalcular
            # estado del modulo despues de cambiar su costo.
            _recalcular_estado_modulo(mod)

    # F-FIX-DESCUENTO-TOTAL-PAGAR: total_a_pagar se calcula desde la
    # SUMA de los costos actuales de los modulos (que ya tienen el descuento
    # aplicado) + matricula + cargo. Esto garantiza que si los modulos
    # fueron actualizados por separado, total_a_pagar refleja esa realidad.
    suma_modulos_actual = sum(m.costo or 0 for m in enrollment.modulos)
    total_final = suma_modulos_actual + enrollment.costo_matricula + cargo_adicional
    enrollment.total_a_pagar = round(total_final, 2)
    enrollment.saldo_pendiente = round(max(0.0, enrollment.total_a_pagar - enrollment.total_pagado), 2)


# ========================================================================
# ISSUE-Q-NOTA-BORRADOR: Notas de docente como borrador validado por CPD
# ========================================================================
async def subir_nota_borrador(
    enrollment_id: PydanticObjectId,
    modulo_index: int,
    nota_borrador: float
) -> Enrollment:
    """
    El docente propone una nota que queda como BORRADOR hasta que CPD la valide.
    No afecta nota_final ni la lógica de pérdida de beca (ISSUE-P-RECALCULO-NOTA).
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError("Inscripción no encontrada")
    if modulo_index < 0 or modulo_index >= len(enrollment.modulos):
        raise ValueError(f"Índice de módulo {modulo_index} fuera de rango")

    modulo = enrollment.modulos[modulo_index]
    modulo.nota_borrador = round(nota_borrador, 2)
    modulo.estado_validacion_nota = "pendiente_validacion"
    enrollment.updated_at = utcnow_naive()
    await enrollment.save()
    return enrollment


async def subir_notas_borrador_bulk(
    items: List[Any],
) -> tuple[int, int, List[dict]]:
    """
    Guarda notas borrador para múltiples inscripciones en una sola operación eficiente.
    Reutiliza una sola query MongoDB en lote y guarda concurrentemente.
    Retorna (exitosos, fallidos, resultados).
    """
    import asyncio
    from bson import ObjectId
    
    # Extraer IDs únicos para batch query
    def _to_obj_id(raw_id: Any) -> ObjectId:
        if isinstance(raw_id, ObjectId):
            return raw_id
        return ObjectId(str(raw_id))

    enrollment_ids = list(set(_to_obj_id(item.enrollment_id if hasattr(item, 'enrollment_id') else item['enrollment_id']) for item in items))
    
    enrollments_list = await Enrollment.find({"_id": {"$in": enrollment_ids}}).to_list()
    enrollments_dict = {e.id: e for e in enrollments_list}
    
    exitosos = 0
    fallidos = 0
    resultados = []
    to_save = []
    
    now = utcnow_naive()
    
    for item in items:
        raw_eid = item.enrollment_id if hasattr(item, 'enrollment_id') else item['enrollment_id']
        m_idx = item.modulo_index if hasattr(item, 'modulo_index') else item['modulo_index']
        nota = item.nota if hasattr(item, 'nota') else item['nota']
        
        eid = _to_obj_id(raw_eid)
        enrollment = enrollments_dict.get(eid)
        
        if not enrollment:
            fallidos += 1
            resultados.append({
                "enrollment_id": str(eid),
                "modulo_index": m_idx,
                "exito": False,
                "error": "Inscripción no encontrada"
            })
            continue
            
        if not enrollment.modulos or m_idx < 0 or m_idx >= len(enrollment.modulos):
            fallidos += 1
            resultados.append({
                "enrollment_id": str(eid),
                "modulo_index": m_idx,
                "exito": False,
                "error": f"Índice de módulo {m_idx} fuera de rango"
            })
            continue
            
        modulo = enrollment.modulos[m_idx]
        modulo.nota_borrador = round(float(nota), 2)
        modulo.estado_validacion_nota = "pendiente_validacion"
        enrollment.updated_at = now
        to_save.append(enrollment)
        
        exitosos += 1
        resultados.append({
            "enrollment_id": str(eid),
            "modulo_index": m_idx,
            "exito": True,
            "nota_guardada": round(float(nota), 2)
        })
        
    if to_save:
        unique_to_save = list({e.id: e for e in to_save}.values())
        await asyncio.gather(*(e.save() for e in unique_to_save))
        
    return exitosos, fallidos, resultados



async def validar_nota_borrador(
    enrollment_id: PydanticObjectId,
    modulo_index: int,
    evaluador_username: str
) -> Enrollment:
    """
    CPD/Admin/Superadmin validan el borrador: lo convierte en nota oficial
    reutilizando actualizar_nota_modulo (mismo recálculo de promedio y beca).
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError("Inscripción no encontrada")
    if modulo_index < 0 or modulo_index >= len(enrollment.modulos):
        raise ValueError(f"Índice de módulo {modulo_index} fuera de rango")

    modulo = enrollment.modulos[modulo_index]
    if modulo.estado_validacion_nota != "pendiente_validacion" or modulo.nota_borrador is None:
        raise ValueError("Este módulo no tiene un borrador pendiente de validación")

    nota_a_oficializar = modulo.nota_borrador
    nombre_modulo = modulo.nombre

    enrollment_actualizado = await actualizar_nota_modulo(
        enrollment_id=enrollment_id,
        modulo_index=modulo_index,
        nota=nota_a_oficializar,
        evaluador_username=evaluador_username
    )

    # actualizar_nota_modulo no toca nota_borrador/estado_validacion_nota; lo hacemos aquí
    enrollment_actualizado.modulos[modulo_index].estado_validacion_nota = "validada"
    enrollment_actualizado.modulos[modulo_index].nota_borrador = None
    await enrollment_actualizado.save()

    # ISSUE-Q-CORREO-NOTA (2026-07-08, reunión de postgrado contaduría): notificar
    # al estudiante por correo cuando CPD valida su nota. No bloqueante: si el
    # estudiante no tiene email o el envío falla, no revierte la validación.
    try:
        from models.student import Student
        from models.course import Course
        from core.email_utils import send_email, build_nota_validada_email
        from core.config import settings
        from services.notification_service import create_notification

        student = await Student.get(enrollment_actualizado.estudiante_id)
        course = await Course.get(enrollment_actualizado.curso_id)

        if student:
            try:
                await create_notification(
                    destinatario_id=student.id,
                    tipo_destinatario="student",
                    titulo="Calificación Validada",
                    mensaje=f"Tu nota del módulo '{nombre_modulo}' ya fue validada oficialmente por CPD.",
                    tipo_alerta="success",
                    ruta="/app/enrollments",
                    referencia_tipo="enrollment",
                    referencia_id=enrollment_actualizado.id
                )
            except Exception as e:
                print(f"Error notificando validación de nota al estudiante: {str(e)}")

            if student.email and course:
                portal_link = f"{settings.FRONTEND_URL.rstrip('/')}/app/enrollments"
                html = build_nota_validada_email(
                    nombre=student.nombre or student.registro,
                    curso_nombre=course.nombre_programa,
                    modulo_nombre=nombre_modulo,
                    nota=nota_a_oficializar,
                    portal_link=portal_link
                )
                from services import email_service
                await email_service.enviar(
                    destinatario=student.email,
                    asunto=f"Nota validada: {nombre_modulo} · Posgrado UAGRM",
                    html=html,
                    tipo=email_service.TipoEmail.NOTA_VALIDADA,
                    destinatario_id=getattr(student, "id", None),
                    destinatario_nombre=getattr(student, "nombre", None),
                )
    except Exception as e:
        print(f"Error enviando correo de nota validada: {str(e)}")

    return enrollment_actualizado


async def rechazar_nota_borrador(
    enrollment_id: PydanticObjectId,
    modulo_index: int
) -> Enrollment:
    """
    CPD rechaza el borrador propuesto por el docente. No toca la nota oficial
    (que puede mantener un valor validado previamente, si existía).

    F-050 (2026-07-22, audios viejos): antes NO recalculaba `nota_final`
    (promedio) tras rechazar, así que el promedio podía seguir sumando el
    borrador rechazado. Fix: recalcular SIEMPRE el promedio, igual que
    `validar_nota_borrador` y `actualizar_nota_modulo`.
    """
    enrollment = await Enrollment.get(enrollment_id)
    if not enrollment:
        raise ValueError("Inscripción no encontrada")
    if modulo_index < 0 or modulo_index >= len(enrollment.modulos):
        raise ValueError(f"Índice de módulo {modulo_index} fuera de rango")

    modulo = enrollment.modulos[modulo_index]
    if modulo.estado_validacion_nota != "pendiente_validacion":
        raise ValueError("Este módulo no tiene un borrador pendiente de validación")

    modulo.nota_borrador = None
    modulo.estado_validacion_nota = "sin_borrador"

    # F-050 FIX: recalcular el promedio igual que en actualizar_nota_modulo
    # y validar_nota_borrador. Sin esto, el promedio podía seguir sumando
    # la nota rechazada en alguna vista de UI / reportes.
    notas_evaluadas = [m.nota for m in enrollment.modulos if m.nota is not None]
    if notas_evaluadas:
        promedio = sum(notas_evaluadas) / len(notas_evaluadas)
        enrollment.nota_final = round(promedio, 2)
    else:
        # Sin notas válidas, no hay promedio. Mantener como None.
        enrollment.nota_final = None

    enrollment.updated_at = utcnow_naive()
    await enrollment.save()
    return enrollment
